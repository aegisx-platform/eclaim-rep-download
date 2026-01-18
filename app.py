"""Flask Web UI for E-Claim Downloader"""

import os
from flask import Flask, render_template, jsonify, request, send_from_directory, redirect, url_for, Response, stream_with_context, g
from werkzeug.utils import secure_filename
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
import humanize
import psycopg2
from pathlib import Path
import subprocess
import sys
import traceback
import yaml
from utils import FileManager, DownloaderRunner
from utils.history_manager_db import HistoryManagerDB
from utils.import_runner import ImportRunner
from utils.stm_import_runner import STMImportRunner
from utils.unified_import_runner import unified_import_runner
from utils.log_stream import log_streamer
from utils.settings_manager import SettingsManager
from utils.scheduler import download_scheduler
from utils.job_history_manager import job_history_manager
from utils.alert_manager import alert_manager
from utils.logging_config import setup_logger, safe_format_exception
from utils.auth import auth_manager, User, require_role, require_admin
from utils.audit_logger import audit_logger
from utils.rate_limiter import rate_limiter, limit_login, limit_api, limit_download, limit_export
from utils.security_headers import setup_security_headers
from utils.license_middleware import require_license_write_access, get_license_status_banner
from utils.fiscal_year import (
    get_fiscal_year_sql_filter_gregorian,
    get_fiscal_year_range_gregorian,
    get_fiscal_year_range_be
)
from config.database import get_db_config, DB_TYPE
from config.db_pool import init_pool, close_pool, get_connection as get_pooled_connection, return_connection, get_pool_status

# Import blueprints
from routes.settings import settings_api_bp
from routes.settings_pages import settings_pages_bp
from routes.external_api import external_api_bp
from routes.api_keys_management import api_keys_mgmt_bp
from routes.analytics_api import analytics_api_bp
from routes.downloads_api import downloads_api_bp
from routes.imports_api import imports_api_bp
from routes.files_api import files_api_bp

# Flask-Login
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_bcrypt import Bcrypt

# Flask-WTF CSRF Protection
from flask_wtf.csrf import CSRFProtect, CSRFError

# Swagger/OpenAPI Documentation
from flasgger import Swagger


# Database-specific SQL helpers for PostgreSQL/MySQL compatibility
# Note: MySQL % in DATE_FORMAT must be escaped as %% when used with cursor.execute()
def sql_date_trunc_month(column: str) -> str:
    """Generate SQL for truncating date to month start"""
    if DB_TYPE == 'mysql':
        return f"DATE_FORMAT({column}, '%%Y-%%m-01')"
    return f"DATE_TRUNC('month', {column})"


def sql_count_distinct_months(column: str) -> str:
    """Generate SQL for counting distinct months"""
    if DB_TYPE == 'mysql':
        return f"COUNT(DISTINCT DATE_FORMAT({column}, '%%Y-%%m'))"
    return f"COUNT(DISTINCT DATE_TRUNC('month', {column}))"


def sql_current_month_start() -> str:
    """Generate SQL for start of current month"""
    if DB_TYPE == 'mysql':
        return "DATE_FORMAT(CURRENT_DATE, '%%Y-%%m-01')"
    return "DATE_TRUNC('month', CURRENT_DATE)"


def sql_interval_months(months: int) -> str:
    """Generate SQL for interval in months"""
    if DB_TYPE == 'mysql':
        return f"INTERVAL {months} MONTH"
    return f"INTERVAL '{months} months'"


def sql_format_year_month(column: str) -> str:
    """Generate SQL for formatting date as YYYY-MM"""
    if DB_TYPE == 'mysql':
        return f"DATE_FORMAT({column}, '%%Y-%%m')"
    return f"TO_CHAR({column}, 'YYYY-MM')"


def sql_format_month(column: str) -> str:
    """Generate SQL for formatting date as YYYY-MM (alias for sql_format_year_month)"""
    return sql_format_year_month(column)


def sql_cast_numeric(expr: str) -> str:
    """Generate SQL for casting to numeric type"""
    if DB_TYPE == 'mysql':
        return f"CAST({expr} AS DECIMAL(15,2))"
    return f"({expr})::numeric"


def sql_interval_days(days: int) -> str:
    """Generate SQL for interval in days"""
    if DB_TYPE == 'mysql':
        return f"INTERVAL {days} DAY"
    return f"INTERVAL '{days} days'"


def sql_regex_match(column: str, pattern: str) -> str:
    """Generate SQL for regex matching (is numeric check)"""
    if DB_TYPE == 'mysql':
        return f"{column} REGEXP '{pattern}'"
    return f"{column} ~ '{pattern}'"


def sql_coalesce_numeric(column: str, default: int = 0) -> str:
    """Generate SQL for COALESCE with numeric cast"""
    if DB_TYPE == 'mysql':
        return f"CAST(COALESCE({column}, {default}) AS DECIMAL(15,2))"
    return f"COALESCE({column}, {default})::numeric"


def sql_cast_int(expr: str) -> str:
    """Generate SQL for casting to integer"""
    if DB_TYPE == 'mysql':
        return f"CAST({expr} AS SIGNED)"
    return f"({expr})::int"


def sql_extract_year(column: str) -> str:
    """Generate SQL for extracting year as integer"""
    if DB_TYPE == 'mysql':
        return f"YEAR({column})"
    return f"EXTRACT(YEAR FROM {column})::int"


def sql_extract_month(column: str) -> str:
    """Generate SQL for extracting month"""
    if DB_TYPE == 'mysql':
        return f"MONTH({column})"
    return f"EXTRACT(MONTH FROM {column})"


def sql_full_outer_join(left_table: str, right_table: str, left_alias: str, right_alias: str, join_condition: str) -> str:
    """
    Generate SQL for FULL OUTER JOIN.
    MySQL doesn't support FULL OUTER JOIN, so we simulate it with UNION of LEFT and RIGHT JOINs.
    For PostgreSQL, use native FULL OUTER JOIN.

    Note: This returns a subquery that can be used in FROM clause.
    """
    if DB_TYPE == 'mysql':
        return f"""(
            SELECT * FROM {left_table} {left_alias}
            LEFT JOIN {right_table} {right_alias} ON {join_condition}
            UNION
            SELECT * FROM {left_table} {left_alias}
            RIGHT JOIN {right_table} {right_alias} ON {join_condition}
            WHERE {left_alias}.{join_condition.split('=')[0].strip().split('.')[-1]} IS NULL
        )"""
    return f"{left_table} {left_alias} FULL OUTER JOIN {right_table} {right_alias} ON {join_condition}"


def sql_ilike(column: str, pattern: str) -> str:
    """Generate SQL for case-insensitive LIKE"""
    if DB_TYPE == 'mysql':
        return f"{column} LIKE {pattern}"  # MySQL LIKE is case-insensitive by default with utf8mb4
    return f"{column} ILIKE {pattern}"


# Thailand timezone
TZ_BANGKOK = ZoneInfo('Asia/Bangkok')

app = Flask(__name__)

# Set up secure logging with credential masking
logger = setup_logger('eclaim_app', enable_masking=True)

# Load SECRET_KEY from environment variable
# CRITICAL: Must be set in production for session security
secret_key = os.environ.get('SECRET_KEY')
if not secret_key:
    if os.environ.get('FLASK_ENV') == 'production':
        raise RuntimeError('SECRET_KEY environment variable must be set in production')
    # Development fallback (not for production use)
    secret_key = 'dev-only-secret-key-do-not-use-in-production'
app.config['SECRET_KEY'] = secret_key

# CSRF Protection Configuration
app.config['WTF_CSRF_ENABLED'] = True
app.config['WTF_CSRF_TIME_LIMIT'] = None  # No time limit (rely on session expiry)

# HTTPS-specific security settings
# Automatically detect HTTPS mode from environment variable
is_https = os.environ.get('WTF_CSRF_SSL_STRICT', 'false').lower() == 'true'
app.config['WTF_CSRF_SSL_STRICT'] = is_https  # Enforce HTTPS for CSRF cookies when True

# Session cookie security (HTTPS mode)
if is_https:
    app.config['SESSION_COOKIE_SECURE'] = True  # HTTPS-only session cookies
    app.config['SESSION_COOKIE_HTTPONLY'] = True  # Prevent JavaScript access
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # CSRF protection
    logger.info("HTTPS mode enabled - Secure cookies configured")
else:
    # Development mode - allow HTTP
    app.config['SESSION_COOKIE_SECURE'] = False
    logger.warning("HTTP mode - Cookies not secured (development only)")

app.config['WTF_CSRF_CHECK_DEFAULT'] = True  # Enable CSRF protection by default

# Initialize CSRF Protection
csrf = CSRFProtect(app)

# CSRF Error Handler
@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    """Handle CSRF validation errors."""
    logger.warning(f"CSRF validation failed: {e.description} - IP: {request.remote_addr}")

    # Log security event
    audit_logger.log(
        action='DATA_ACCESS',
        resource_type='csrf_validation',
        user_id=current_user.username if current_user.is_authenticated else 'anonymous',
        ip_address=request.remote_addr,
        status='denied',
        error_message=f'CSRF validation failed: {e.description}'
    )

    # Return JSON for AJAX requests
    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({
            'success': False,
            'error': 'CSRF validation failed. Please refresh the page and try again.',
            'csrf_error': True
        }), 400

    # Return error page for regular requests
    return render_template(
        'error.html',
        error_title='Security Validation Failed',
        error_message='CSRF validation failed. This usually happens when your session has expired or the page has been open for too long.',
        suggestion='Please refresh the page and try again.'
    ), 400

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'กรุณาเข้าสู่ระบบก่อนใช้งาน'
login_manager.login_message_category = 'info'

# Initialize Bcrypt
bcrypt = Bcrypt(app)

# Setup Security Headers
# Auto-detect mode: strict for HTTPS, permissive for HTTP (development)
security_mode = 'strict' if is_https else 'permissive'
setup_security_headers(app, mode=security_mode)
logger.info(f"Security headers configured ({security_mode} mode)")

# Register blueprints
app.register_blueprint(settings_api_bp)  # API routes
logger.info("✓ Settings API blueprint registered")

app.register_blueprint(settings_pages_bp)  # Page routes
logger.info("✓ Settings Pages blueprint registered")

app.register_blueprint(external_api_bp)  # External API v1 routes
logger.info("✓ External API blueprint registered at /api/v1")

app.register_blueprint(api_keys_mgmt_bp)  # API Keys management routes
logger.info("✓ API Keys Management blueprint registered")

app.register_blueprint(analytics_api_bp)  # Analytics and Predictive API routes
logger.info("✓ Analytics API blueprint registered")

app.register_blueprint(downloads_api_bp)  # Downloads and History API routes
logger.info("✓ Downloads API blueprint registered")

app.register_blueprint(imports_api_bp)  # Imports API routes (REP, STM, SMT)
logger.info("✓ Imports API blueprint registered")

app.register_blueprint(files_api_bp)  # File management API routes
logger.info("✓ Files API blueprint registered")

# Initialize Swagger UI for API documentation
# Load OpenAPI spec from YAML file
openapi_spec_path = os.path.join(app.root_path, 'static', 'swagger', 'openapi.yaml')

swagger_config = {
    "headers": [],
    "specs": [
        {
            "endpoint": 'apispec',
            "route": '/api/v1/apispec.json',
            "rule_filter": lambda _: True,
            "model_filter": lambda _: True,
        }
    ],
    "static_url_path": "/flasgger_static",
    "swagger_ui": True,
    "specs_route": "/api/docs",
    "openapi": "3.0.3"
}

# Load OpenAPI spec from YAML file
try:
    with open(openapi_spec_path, 'r', encoding='utf-8') as f:
        swagger_template = yaml.safe_load(f)
    logger.info("✓ Loaded OpenAPI spec from openapi.yaml")
except FileNotFoundError:
    logger.warning("⚠ OpenAPI spec file not found, using minimal template")
    swagger_template = {
        "openapi": "3.0.3",
        "info": {
            "title": "E-Claim HIS Integration API",
            "description": "REST API for Hospital Information System (HIS) integration with E-Claim data",
            "version": "1.0.0"
        }
    }

swagger = Swagger(app, config=swagger_config, template=swagger_template)
logger.info("✓ Swagger UI initialized at /api/docs")

@login_manager.user_loader
def load_user(user_id):
    """Load user for Flask-Login."""
    return auth_manager.get_user_by_id(int(user_id))

# Inject current_user and license info into all templates
@app.context_processor
def inject_user():
    """Inject user info and license info into all templates."""
    license_info = settings_manager.get_license_info()
    return dict(
        current_user=current_user,
        license=license_info
    )

# Initialize managers - now using database-backed history
history_manager = HistoryManagerDB(download_type='rep')  # REP files
stm_history_manager = HistoryManagerDB(download_type='stm')  # Statement files
file_manager = FileManager()
downloader_runner = DownloaderRunner()
import_runner = ImportRunner()
stm_import_runner = STMImportRunner()
settings_manager = SettingsManager()

# Make managers accessible to blueprints
app.config['downloader_runner'] = downloader_runner
app.config['settings_manager'] = settings_manager
app.config['history_manager'] = history_manager
app.config['stm_history_manager'] = stm_history_manager
app.config['file_manager'] = file_manager
app.config['FILE_MANAGER'] = file_manager  # Alias for backward compatibility
app.config['UNIFIED_IMPORT_RUNNER'] = import_runner
app.config['import_runner'] = import_runner
app.config['stm_import_runner'] = stm_import_runner


def check_license_status():
    """Check and log license status on application startup"""
    try:
        license_info = settings_manager.get_license_info()

        if not license_info['is_valid']:
            logger.warning("╔══════════════════════════════════════════════════════════╗")
            logger.warning("║  NO VALID LICENSE - RUNNING IN TRIAL MODE                ║")
            logger.warning("║  Limited to 1,000 records per import                     ║")
            logger.warning("║  Advanced features disabled                              ║")
            logger.warning("║  Contact: https://github.com/aegisx-platform            ║")
            logger.warning("╚══════════════════════════════════════════════════════════╝")
        else:
            tier = license_info.get('tier', 'unknown')
            expires_at = license_info.get('expires_at')
            days_left = license_info.get('days_until_expiry')

            if license_info.get('grace_period'):
                logger.warning("╔══════════════════════════════════════════════════════════╗")
                logger.warning(f"║  LICENSE EXPIRED - GRACE PERIOD                          ║")
                logger.warning(f"║  {license_info.get('grace_days_left', 0)} days remaining before restrictions     ║")
                logger.warning("║  Please renew your license                               ║")
                logger.warning("╚══════════════════════════════════════════════════════════╝")
            elif days_left is not None and days_left <= 30:
                logger.warning(f"License expires in {days_left} days - Tier: {tier}")
            else:
                logger.info(f"✓ Valid license - Tier: {tier} - Expires: {expires_at}")

    except Exception as e:
        logger.error(f"Error checking license status: {e}")


def init_scheduler():
    """Initialize unified scheduler with saved settings for all data types"""
    try:
        # Get unified schedule settings
        schedule_settings = settings_manager.get_schedule_settings()

        # Clear all existing scheduled jobs
        for job in download_scheduler.get_all_jobs():
            if job['id'].startswith('download_'):
                download_scheduler.remove_scheduled_download(job['id'])
        download_scheduler.remove_stm_jobs()
        download_scheduler.remove_smt_jobs()

        if not schedule_settings['schedule_enabled']:
            log_streamer.write_log(
                "⏸ Scheduler disabled",
                'info',
                'system'
            )
            return

        # Get common settings
        times = schedule_settings['schedule_times']
        auto_import = schedule_settings['schedule_auto_import']
        schemes = schedule_settings.get('schedule_schemes', ['ucs', 'ofc', 'sss', 'lgo'])
        type_rep = schedule_settings.get('schedule_type_rep', True)
        type_stm = schedule_settings.get('schedule_type_stm', False)
        type_smt = schedule_settings.get('schedule_type_smt', False)
        smt_vendor_id = schedule_settings.get('schedule_smt_vendor_id', '')

        if not times:
            log_streamer.write_log(
                "⚠ Scheduler enabled but no times configured",
                'warning',
                'system'
            )
            return

        # Schedule REP jobs if enabled
        if type_rep:
            for time_config in times:
                hour = time_config.get('hour', 0)
                minute = time_config.get('minute', 0)
                download_scheduler.add_scheduled_download(hour, minute, auto_import)

            log_streamer.write_log(
                f"✓ REP Scheduler initialized with {len(times)} jobs",
                'success',
                'system'
            )

        # Schedule Statement (STM) jobs if enabled - UCS only
        if type_stm:
            for time_config in times:
                hour = time_config.get('hour', 0)
                minute = time_config.get('minute', 0)
                download_scheduler.add_stm_scheduled_download(hour, minute, auto_import)

            log_streamer.write_log(
                f"✓ UC Statement Scheduler initialized with {len(times)} jobs",
                'success',
                'system'
            )

        # Schedule SMT Budget jobs if enabled
        if type_smt and smt_vendor_id:
            for time_config in times:
                hour = time_config.get('hour', 0)
                minute = time_config.get('minute', 0)
                download_scheduler.add_smt_scheduled_fetch(hour, minute, smt_vendor_id, auto_import)

            log_streamer.write_log(
                f"✓ SMT Budget Scheduler initialized with {len(times)} jobs (vendor: {smt_vendor_id})",
                'success',
                'system'
            )

        # Log summary
        data_types = []
        if type_rep:
            data_types.append('REP')
        if type_stm:
            data_types.append('Statement')
        if type_smt:
            data_types.append('SMT')
        log_streamer.write_log(
            f"📅 Unified scheduler active: {len(times)} times, types=[{', '.join(data_types)}]",
            'info',
            'system'
        )

    except Exception as e:
        app.logger.error(f"Error initializing scheduler: {e}")


def get_db_connection():
    """Get database connection from pool"""
    try:
        conn = get_pooled_connection()
        if conn is None:
            app.logger.error("Failed to get connection from pool")
        return conn
    except Exception as e:
        app.logger.error(f"Database connection error: {e}")
        return None


def get_import_status_map():
    """
    Get mapping of filename -> import status from database

    Returns:
        dict: {filename: {'imported': bool, 'file_id': int, 'imported_at': str}}
    """
    status_map = {}
    conn = get_db_connection()

    if not conn:
        return status_map

    try:
        cursor = conn.cursor()
        query = """
            SELECT filename, id, status, import_completed_at, imported_records, total_records
            FROM eclaim_imported_files
            WHERE status = 'completed'
            ORDER BY import_completed_at DESC
        """
        cursor.execute(query)
        rows = cursor.fetchall()

        for row in rows:
            filename, file_id, status, completed_at, imported_records, total_records = row
            status_map[filename] = {
                'imported': True,
                'file_id': file_id,
                'imported_at': completed_at.isoformat() if completed_at else None,
                'imported_records': imported_records or 0,
                'total_records': total_records or 0
            }

        cursor.close()
        conn.close()

    except Exception as e:
        app.logger.error(f"Error getting import status: {e}")
        if conn:
            conn.close()

    return status_map


# =============================================================================
# AUTHENTICATION ROUTES
# =============================================================================

@app.route('/login', methods=['GET', 'POST'])
@limit_login  # Rate limit: 5 requests per 5 minutes per IP (brute force protection)
def login():
    """User login page."""
    # If already logged in, redirect to dashboard
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        remember = request.form.get('remember') == 'on'

        if not username or not password:
            return render_template('login.html', error='กรุณากรอกชื่อผู้ใช้และรหัสผ่าน')

        # Authenticate user
        user = auth_manager.authenticate(
            username=username,
            password=password,
            ip_address=request.remote_addr
        )

        if user:
            # Login successful
            login_user(user, remember=remember)

            # Store user info in Flask's g object for audit logging
            g.user_id = user.username
            g.user_email = user.email

            # Check if must change password
            if user.must_change_password:
                return redirect(url_for('change_password'))

            # Redirect to next page or dashboard
            next_page = request.args.get('next')
            if next_page:
                return redirect(next_page)
            return redirect(url_for('dashboard'))
        else:
            # Login failed (error already logged in authenticate())
            return render_template('login.html', error='ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง')

    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    """User logout."""
    # Log logout
    audit_logger.log_logout(
        user_id=current_user.username,
        ip_address=request.remote_addr
    )

    logout_user()
    return redirect(url_for('login'))


@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    """Change password page."""
    if request.method == 'POST':
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')

        # Validate inputs
        if not new_password or not confirm_password:
            return render_template(
                'change_password.html',
                error='กรุณากรอกรหัสผ่านใหม่และยืนยันรหัสผ่าน',
                must_change=current_user.must_change_password
            )

        if new_password != confirm_password:
            return render_template(
                'change_password.html',
                error='รหัสผ่านใหม่ไม่ตรงกัน',
                must_change=current_user.must_change_password
            )

        if len(new_password) < 8:
            return render_template(
                'change_password.html',
                error='รหัสผ่านต้องมีความยาวอย่างน้อย 8 ตัวอักษร',
                must_change=current_user.must_change_password
            )

        # If not must_change, verify current password
        if not current_user.must_change_password:
            user_data = auth_manager.get_user_by_username(current_user.username)
            if not user_data or not auth_manager.verify_password(current_password, user_data['password_hash']):
                return render_template(
                    'change_password.html',
                    error='รหัสผ่านปัจจุบันไม่ถูกต้อง',
                    must_change=False
                )

        # Change password
        success = auth_manager.change_password(
            user_id=current_user.id,
            new_password=new_password,
            changed_by=current_user.username
        )

        if success:
            # Update current user's must_change_password flag
            current_user.must_change_password = False

            return render_template(
                'change_password.html',
                success='เปลี่ยนรหัสผ่านสำเร็จ กรุณาเข้าสู่ระบบใหม่',
                must_change=False
            )
        else:
            return render_template(
                'change_password.html',
                error='ไม่สามารถเปลี่ยนรหัสผ่านได้ กรุณาลองใหม่อีกครั้ง',
                must_change=current_user.must_change_password
            )

    return render_template(
        'change_password.html',
        must_change=current_user.must_change_password
    )


@app.route('/')
def index():
    """Redirect to dashboard or setup if needed"""
    # Check if setup is needed
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            # Quick check for essential seed data
            cursor.execute("SELECT COUNT(*) FROM health_offices")
            health_count = cursor.fetchone()[0]
            cursor.close()
            conn.close()

            # Check hospital code setting
            hospital_code = settings_manager.get_hospital_code()

            # Redirect to setup if:
            # 1. Seed data not complete (health_offices < 1000 records)
            # 2. Hospital code not configured
            if health_count < 1000 or not hospital_code:
                return redirect(url_for('setup'))
    except:
        pass

    return redirect(url_for('dashboard'))


@app.route('/setup')
def setup():
    """System setup and initialization page"""
    return render_template('setup.html')


@app.route('/settings/hospital')
@login_required
def hospital_settings():
    """Hospital settings page for managing hospital code and information"""
    current_hospital_code = settings_manager.get_hospital_code()
    return render_template('settings/hospital.html', hospital_code=current_hospital_code)


@app.route('/master-data')
@login_required
def master_data_management():
    """Master data management page with tabs"""
    tab = request.args.get('tab', 'offices')  # Default to health offices tab

    # Map tabs to templates
    tab_templates = {
        'offices': 'master_data/health_offices.html',
        'error-codes': 'master_data/error_codes.html',
        'fund-types': 'master_data/fund_types.html',
        'service-types': 'master_data/service_types.html',
        'dim-date': 'master_data/dim_date.html'
    }

    template = tab_templates.get(tab, 'master_data/health_offices.html')
    return render_template(template, active_tab=tab, settings=settings_manager.load_settings())


@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard view with statistics"""
    stats = history_manager.get_statistics()
    latest_files = history_manager.get_latest(5)

    # Format file sizes and dates for display
    for file in latest_files:
        file['size_formatted'] = humanize.naturalsize(file.get('file_size') or 0)
        try:
            # Parse datetime (stored as UTC in Docker container)
            dt = datetime.fromisoformat(file.get('download_date', ''))
            # If naive datetime, assume it's UTC and convert to Bangkok time
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc).astimezone(TZ_BANGKOK)
            # Use current Bangkok time for relative time calculation
            now = datetime.now(TZ_BANGKOK)
            file['date_formatted'] = humanize.naturaltime(dt, when=now)
        except (ValueError, TypeError, AttributeError):
            file['date_formatted'] = file.get('download_date', 'Unknown')

    # Check if downloader is running
    downloader_status = downloader_runner.get_status()

    # Get schedule settings and jobs
    schedule_settings = settings_manager.get_schedule_settings()
    schedule_jobs = download_scheduler.get_all_jobs()

    return render_template(
        'dashboard.html',
        stats=stats,
        latest_files=latest_files,
        downloader_running=downloader_status['running'],
        schedule_settings=schedule_settings,
        schedule_jobs=schedule_jobs
    )


@app.route('/upload')
@login_required
def upload():
    """Manual file upload page with downloaded files tab"""
    # Get same data as files() route for the Files tab
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    filter_month = request.args.get('month', type=int)
    filter_year = request.args.get('year', type=int)
    filter_type = request.args.get('type', type=str)

    # Default to current month/year if not specified
    now = datetime.now(TZ_BANGKOK)
    if filter_month is None:
        filter_month = now.month
    if filter_year is None:
        filter_year = now.year + 543  # Convert to Buddhist Era

    all_files = history_manager.get_all_downloads()

    # Get import status from database
    import_status_map = get_import_status_map()

    # Filter by month/year and type
    filtered_files = []
    for file in all_files:
        file_month = file.get('month')
        file_year = file.get('year')

        # If month/year not in file metadata, try to extract from filename
        if file_month is None or file_year is None:
            import re
            match = re.search(r'_(\d{4})(\d{2})\d{2}_', file.get('filename', ''))
            if match:
                file_year = int(match.group(1))
                file_month = int(match.group(2))

        # Apply month/year filter
        if file_month != filter_month or file_year != filter_year:
            continue

        # Apply type filter if specified
        if filter_type:
            file_path = file.get('file_path', '')
            if filter_type == 'rep' and '/rep/' not in file_path:
                continue
            elif filter_type == 'stm' and '/stm/' not in file_path:
                continue
            elif filter_type == 'smt' and '/smt/' not in file_path:
                continue

        filtered_files.append(file)

    # Sort by download date (most recent first)
    filtered_files = sorted(
        filtered_files,
        key=lambda d: d.get('download_date') or '',
        reverse=True
    )

    # Calculate pagination
    total_files = len(filtered_files)
    total_pages = (total_files + per_page - 1) // per_page
    page = max(1, min(page, total_pages if total_pages > 0 else 1))

    # Paginate
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated_files = filtered_files[start_idx:end_idx]

    # Format for display and add import status
    for file in paginated_files:
        file['size_formatted'] = humanize.naturalsize(file.get('file_size') or 0)
        try:
            dt = datetime.fromisoformat(file.get('download_date', ''))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc).astimezone(TZ_BANGKOK)
            file['date_formatted'] = dt.strftime('%Y-%m-%d %H:%M:%S')
            now = datetime.now(TZ_BANGKOK)
            file['date_relative'] = humanize.naturaltime(dt, when=now)
        except (ValueError, TypeError, AttributeError):
            file['date_formatted'] = file.get('download_date', 'Unknown')
            file['date_relative'] = 'Unknown'

        # Add import status
        filename = file.get('filename', '')
        if filename in import_status_map:
            file['import_status'] = import_status_map[filename]
            file['imported'] = True
        else:
            file['import_status'] = None
            file['imported'] = False

    # Count imported vs not imported
    imported_count = sum(1 for f in filtered_files if f.get('imported', False))
    not_imported_count = len(filtered_files) - imported_count

    # Get available months/years for filter dropdown
    available_dates = history_manager.get_available_dates()

    # Get schedule settings and jobs
    schedule_settings = settings_manager.get_schedule_settings()
    schedule_jobs = download_scheduler.get_all_jobs()

    return render_template(
        'upload.html',
        files=paginated_files,
        imported_count=imported_count,
        not_imported_count=not_imported_count,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        total_files=total_files,
        filter_month=filter_month,
        filter_year=filter_year,
        filter_type=filter_type,
        available_dates=available_dates,
        schedule_settings=schedule_settings,
        schedule_jobs=schedule_jobs
    )


@app.route('/files')
def files():
    """File list view with all downloads"""
    # Get filter parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    filter_type = request.args.get('type', type=str)  # Filter by file type (rep, stm, smt)
    filter_search = request.args.get('search', type=str)  # Search filename

    # Date range filter with defaults
    now = datetime.now(TZ_BANGKOK)
    default_date_from = now.replace(day=1).strftime('%Y-%m-%d')  # 1st day of current month
    default_date_to = now.strftime('%Y-%m-%d')  # Today

    filter_date_from = request.args.get('date_from', default_date_from, type=str)
    filter_date_to = request.args.get('date_to', default_date_to, type=str)

    all_files = history_manager.get_all_downloads()

    # Get import status from database
    import_status_map = get_import_status_map()

    # Filter by date range, type, and search
    filtered_files = []
    for file in all_files:
        # Extract date from filename (format: eclaim_10670_OP_25690106_xxx.xls → 2569-01-06)
        import re
        match = re.search(r'_(\d{4})(\d{2})(\d{2})_', file.get('filename', ''))
        if match:
            file_year = int(match.group(1))
            file_month = int(match.group(2))
            file_day = int(match.group(3))
            # Convert Buddhist Era to Gregorian (2569 → 2026)
            file_date_str = f"{file_year - 543:04d}-{file_month:02d}-{file_day:02d}"
        else:
            # Fallback: use download_date
            download_date = file.get('download_date', '')
            if download_date:
                try:
                    dt = datetime.fromisoformat(download_date)
                    file_date_str = dt.strftime('%Y-%m-%d')
                except:
                    file_date_str = None
            else:
                file_date_str = None

        # Apply date range filter
        if filter_date_from and file_date_str:
            if file_date_str < filter_date_from:
                continue

        if filter_date_to and file_date_str:
            if file_date_str > filter_date_to:
                continue

        # Apply type filter if specified
        if filter_type:
            file_path = file.get('file_path') or ''
            filename = file.get('filename', '').lower()

            # Check both file_path and filename for type matching
            if filter_type == 'rep':
                if '/rep/' not in file_path and 'eclaim_' not in filename:
                    continue
            elif filter_type == 'stm':
                if '/stm/' not in file_path and 'stm_' not in filename:
                    continue
            elif filter_type == 'smt':
                if '/smt/' not in file_path and 'smt_budget_' not in filename:
                    continue

        # Apply search filter if specified
        if filter_search:
            filename = file.get('filename', '').lower()
            if filter_search.lower() not in filename:
                continue

        filtered_files.append(file)

    # Sort by download date (most recent first)
    filtered_files = sorted(
        filtered_files,
        key=lambda d: d.get('download_date') or '',
        reverse=True
    )

    # Calculate pagination
    total_files = len(filtered_files)
    total_pages = (total_files + per_page - 1) // per_page  # Ceiling division
    page = max(1, min(page, total_pages if total_pages > 0 else 1))

    # Paginate
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated_files = filtered_files[start_idx:end_idx]

    # Format for display and add import status
    for file in paginated_files:
        file['size_formatted'] = humanize.naturalsize(file.get('file_size') or 0)
        try:
            # Parse datetime (MySQL stores in Bangkok time due to SYSTEM timezone)
            dt = datetime.fromisoformat(file.get('download_date', ''))
            # If naive datetime, add Bangkok timezone
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=TZ_BANGKOK)
            file['date_formatted'] = dt.strftime('%Y-%m-%d %H:%M:%S')
            # Use current Bangkok time for relative time calculation
            now = datetime.now(TZ_BANGKOK)
            file['date_relative'] = humanize.naturaltime(dt, when=now)
        except Exception as e:
            # Log error for debugging
            logger.error(f"Error parsing date for {file.get('filename')}: {e}")
            file['date_formatted'] = file.get('download_date', 'Unknown')
            file['date_relative'] = 'Unknown'

        # Add import status
        filename = file.get('filename', '')
        if filename in import_status_map:
            file['import_status'] = import_status_map[filename]
            file['imported'] = True
        else:
            file['import_status'] = None
            file['imported'] = False

    # Count imported vs not imported (from filtered files)
    imported_count = sum(1 for f in filtered_files if f.get('imported', False))
    not_imported_count = len(filtered_files) - imported_count

    # Get available months/years for filter dropdown
    available_dates = history_manager.get_available_dates()

    # Get schedule settings and jobs
    schedule_settings = settings_manager.get_schedule_settings()
    schedule_jobs = download_scheduler.get_all_jobs()

    return render_template(
        'files.html',
        files=paginated_files,
        imported_count=imported_count,
        not_imported_count=not_imported_count,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        total_files=total_files,
        filter_type=filter_type,
        filter_date_from=filter_date_from,
        filter_date_to=filter_date_to,
        filter_search=filter_search,
        available_dates=available_dates,
        schedule_settings=schedule_settings,
        schedule_jobs=schedule_jobs
    )


# ==================== OpenAPI Documentation Routes ====================

@app.route('/api/v1/openapi.yaml')
def serve_openapi_yaml():
    """Serve OpenAPI 3.0 specification in YAML format"""
    return send_from_directory(
        os.path.join(app.root_path, 'static', 'swagger'),
        'openapi.yaml',
        mimetype='application/x-yaml'
    )


@app.route('/api/v1/openapi.json')
def serve_openapi_json():
    """Serve OpenAPI 3.0 specification in JSON format"""
    yaml_path = os.path.join(app.root_path, 'static', 'swagger', 'openapi.yaml')

    try:
        with open(yaml_path, 'r', encoding='utf-8') as f:
            spec = yaml.safe_load(f)
        return jsonify(spec)
    except FileNotFoundError:
        return jsonify({
            'error': 'OpenAPI specification not found'
        }), 404
    except Exception as e:
        return jsonify({
            'error': f'Failed to load OpenAPI specification: {str(e)}'
        }), 500


# ==================== User Management Routes ====================

@app.route('/api/user/change-password', methods=['POST'])
@login_required
def api_change_password():
    """API endpoint for changing password"""
    try:
        data = request.get_json() or {}
        current_password = data.get('current_password', '')
        new_password = data.get('new_password', '')

        # Validate inputs
        if not current_password or not new_password:
            return jsonify({
                'success': False,
                'error': 'กรุณากรอกรหัสผ่านปัจจุบันและรหัสผ่านใหม่'
            }), 400

        if len(new_password) < 6:
            return jsonify({
                'success': False,
                'error': 'รหัสผ่านต้องมีความยาวอย่างน้อย 6 ตัวอักษร'
            }), 400

        # Verify current password
        user_data = auth_manager.get_user_by_username(current_user.username)
        if not user_data or not auth_manager.verify_password(current_password, user_data['password_hash']):
            return jsonify({
                'success': False,
                'error': 'รหัสผ่านปัจจุบันไม่ถูกต้อง'
            }), 401

        # Change password
        success = auth_manager.change_password(
            user_id=current_user.id,
            new_password=new_password,
            changed_by=current_user.username
        )

        if success:
            # Log password change
            audit_logger.log_password_change(
                user_id=current_user.id,
                username=current_user.username,
                changed_by=current_user.username
            )

            return jsonify({
                'success': True,
                'message': 'เปลี่ยนรหัสผ่านสำเร็จ'
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถเปลี่ยนรหัสผ่านได้'
            }), 500

    except Exception as e:
        logger.error(f"Error changing password: {e}")
        return jsonify({
            'success': False,
            'error': 'เกิดข้อผิดพลาดในการเปลี่ยนรหัสผ่าน'
        }), 500
@app.route('/files/<filename>/download')
def download_file(filename):
    """Download file to user's computer"""
    try:
        # Validate filename
        file_path = file_manager.get_file_path(filename)

        if not file_path.exists():
            return jsonify({'error': 'File not found'}), 404

        return send_from_directory(
            file_manager.download_dir,
            filename,
            as_attachment=True
        )

    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': f'Error: {str(e)}'}), 500
@app.route('/api/stats')
def api_stats():
    """JSON API for statistics (for AJAX refresh)"""
    stats = history_manager.get_statistics()
    downloader_status = downloader_runner.get_status()

    return jsonify({
        'stats': stats,
        'downloader_running': downloader_status['running']
    })


@app.route('/download-config')
def download_config():
    """Download configuration page with date selection"""
    return render_template('download_config.html')


@app.route('/data-management')
def data_management():
    """
    Combined Data Management page with tabs for:
    - Download (single month, bulk download, scheduler)
    - Files (file list with import status)
    - SMT Budget sync
    - Settings (credentials, database info)
    """
    # Get filter parameters for files tab
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)

    # Fiscal year parameter (Thai Buddhist Era)
    # Fiscal year 2569 = Oct 2568 - Sep 2569
    fiscal_year = request.args.get('fiscal_year', type=int)

    # Support both old (month/year) and new (start_month/end_month) parameters
    start_month = request.args.get('start_month', type=int) or request.args.get('month', type=int)
    start_year = request.args.get('start_year', type=int) or request.args.get('year', type=int)
    end_month = request.args.get('end_month', type=int)
    end_year = request.args.get('end_year', type=int)

    # If fiscal year is set, calculate start/end years
    if fiscal_year:
        # Fiscal year X runs from Oct (year X-1) to Sep (year X)
        if start_month and start_month >= 10:
            start_year = start_year or (fiscal_year - 1)
        else:
            start_year = start_year or fiscal_year
        if end_month and end_month <= 9:
            end_year = end_year or fiscal_year
        else:
            end_year = end_year or (fiscal_year - 1)

    filter_scheme = request.args.get('scheme', '').strip().lower()
    filter_file_type = request.args.get('file_type', '').strip().lower()  # op, ip, orf, appeal
    filter_status = request.args.get('status', '').strip().lower()  # imported, pending, or empty for all

    # Default to show all if no date specified
    now = datetime.now(TZ_BANGKOK)
    show_all_dates = start_month is None and end_month is None and fiscal_year is None

    # For backward compatibility
    filter_month = start_month or now.month
    filter_year = start_year or (now.year + 543)

    all_files = history_manager.get_all_downloads()

    # Get import status from database
    import_status_map = get_import_status_map()

    # Helper function to convert month/year to comparable number
    def date_to_num(m, y):
        return y * 12 + m

    # Filter files
    filtered_files = []
    for file in all_files:
        file_month = file.get('month')
        file_year = file.get('year')

        # If month/year not in file metadata, try to extract from filename
        if file_month is None or file_year is None:
            import re
            match = re.search(r'_(\d{4})(\d{2})\d{2}_', file.get('filename', ''))
            if match:
                file_year = int(match.group(1))
                file_month = int(match.group(2))

        # Skip if still no date info
        if file_month is None or file_year is None:
            if not show_all_dates:
                continue
        else:
            # Date range filter
            if not show_all_dates:
                file_date_num = date_to_num(file_month, file_year)

                if start_month and start_year:
                    start_num = date_to_num(start_month, start_year)
                    if file_date_num < start_num:
                        continue

                if end_month and end_year:
                    end_num = date_to_num(end_month, end_year)
                    if file_date_num > end_num:
                        continue

        # Scheme filter (check filename)
        if filter_scheme:
            filename_lower = file.get('filename', '').lower()
            if filter_scheme not in filename_lower:
                continue

        # File type filter (op, ip, orf, appeal)
        if filter_file_type:
            filename_upper = file.get('filename', '').upper()
            if filter_file_type == 'op':
                # Match OP but not OPLGO, OPSSS, OP_APPEAL
                if '_OP_' not in filename_upper and not filename_upper.endswith('_OP.xls'):
                    continue
            elif filter_file_type == 'ip':
                # Match IP but not IPLGO, IPSSS, IP_APPEAL
                if '_IP_' not in filename_upper and not filename_upper.endswith('_IP.xls'):
                    continue
            elif filter_file_type == 'orf':
                if '_ORF_' not in filename_upper:
                    continue
            elif filter_file_type == 'appeal':
                if 'APPEAL' not in filename_upper:
                    continue

        # Add import status early for status filtering
        filename = file.get('filename', '')
        file['imported'] = filename in import_status_map
        if filename in import_status_map:
            file['import_status'] = import_status_map[filename]
        else:
            file['import_status'] = None

        # Status filter
        if filter_status:
            if filter_status == 'imported' and not file['imported']:
                continue
            if filter_status == 'pending' and file['imported']:
                continue

        filtered_files.append(file)

    # Sort by download date (most recent first)
    filtered_files = sorted(
        filtered_files,
        key=lambda d: d.get('download_date') or '',
        reverse=True
    )

    # Calculate pagination
    total_files_filtered = len(filtered_files)
    total_pages = (total_files_filtered + per_page - 1) // per_page
    page = max(1, min(page, total_pages if total_pages > 0 else 1))

    # Paginate
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated_files = filtered_files[start_idx:end_idx]

    # Format for display (import status already added during filtering)
    for file in paginated_files:
        file['size_formatted'] = humanize.naturalsize(file.get('file_size') or 0)
        try:
            dt = datetime.fromisoformat(file.get('download_date', ''))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc).astimezone(TZ_BANGKOK)
            file['date_formatted'] = dt.strftime('%Y-%m-%d %H:%M:%S')
            now_time = datetime.now(TZ_BANGKOK)
            file['date_relative'] = humanize.naturaltime(dt, when=now_time)
        except (ValueError, TypeError, AttributeError):
            file['date_formatted'] = file.get('download_date', 'Unknown')
            file['date_relative'] = 'Unknown'

    # Count imported vs not imported (from filtered files, not all files)
    filtered_imported_count = sum(1 for f in filtered_files if f.get('imported', False))
    filtered_not_imported_count = len(filtered_files) - filtered_imported_count

    # Get available months/years for filter
    available_dates = history_manager.get_available_dates()

    # Get settings
    current_settings = settings_manager.load_settings()

    # Get database info
    db_type_display = 'MySQL' if DB_TYPE == 'mysql' else 'PostgreSQL'
    db_info = {
        'type': db_type_display,
        'claims_count': 0,
        'budget_count': 0
    }
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM claim_rep_opip_nhso_item")
            db_info['claims_count'] = cursor.fetchone()[0]
            try:
                cursor.execute("SELECT COUNT(*) FROM smt_budget_transfers")
                db_info['budget_count'] = cursor.fetchone()[0]
            except Exception:
                # Table may not exist in all deployments
                pass
            cursor.close()
            conn.close()
        except Exception as e:
            app.logger.error(f"Error getting db stats: {e}")
            if conn:
                conn.close()

    # Calculate file types count
    file_types = {}
    for f in filtered_files:
        ftype = f.get('file_type', 'unknown').upper()
        file_types[ftype] = file_types.get(ftype, 0) + 1

    # Get last download time
    last_run = 'Never'
    if filtered_files:
        try:
            latest = max(filtered_files, key=lambda x: x.get('download_date', ''))
            last_run = latest.get('download_date', 'Unknown')[:19]  # Truncate to datetime
        except (ValueError, TypeError):
            pass

    # Calculate stats (from filtered files to match filter selection)
    stats = {
        'total_files': len(filtered_files),
        'total_size': humanize.naturalsize(sum((f.get('file_size') or 0) for f in filtered_files)),
        'imported_count': filtered_imported_count,
        'not_imported_count': filtered_not_imported_count,
        'file_types': file_types,
        'last_run': last_run
    }

    # Get schedule settings for display
    schedule_settings = {
        'schedule_enabled': current_settings.get('schedule_enabled', False),
        'schedule_auto_import': current_settings.get('schedule_auto_import', False),
        'schedule_times': current_settings.get('schedule_times', [])
    }

    # Get downloader status
    downloader_status = downloader_runner.get_status()

    # Get next scheduled jobs
    schedule_jobs = download_scheduler.get_all_jobs() if download_scheduler else []

    return render_template(
        'data_management.html',
        files=paginated_files,
        stats=stats,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        total_files=total_files_filtered,
        filter_month=filter_month,
        filter_year=filter_year,
        available_dates=available_dates,
        settings=current_settings,
        db_info=db_info,
        schedule_settings=schedule_settings,
        downloader_running=downloader_status.get('running', False),
        schedule_jobs=schedule_jobs
    )


@app.route('/data-analysis')
def data_analysis():
    """
    Data Analysis page for viewing linked data across:
    - REP (E-Claim Reimbursement)
    - Statement (stm_claim_item)
    - SMT Budget
    """
    return render_template('data_analysis.html')


# ==============================================================================
# Data Analysis API Endpoints
# ==============================================================================

@app.route('/api/stm/download', methods=['POST'])
def trigger_stm_download():
    """Trigger STM (Statement) download with optional auto-import"""
    try:
        data = request.get_json()

        year = data.get('year')
        month = data.get('month')  # Optional
        scheme = data.get('scheme', 'ucs')
        person_type = data.get('person_type', 'all')
        auto_import = data.get('auto_import', False)

        # Validate inputs
        if not year:
            return jsonify({'success': False, 'error': 'Year is required'}), 400

        year = int(year)

        # Validate scheme
        valid_schemes = ['ucs', 'ofc', 'sss', 'lgo']
        if scheme not in valid_schemes:
            return jsonify({'success': False, 'error': f'Invalid scheme. Valid: {valid_schemes}'}), 400

        # Validate person_type
        valid_types = ['ip', 'op', 'all']
        if person_type not in valid_types:
            return jsonify({'success': False, 'error': f'Invalid person_type. Valid: {valid_types}'}), 400

        # Start STM download in background
        import subprocess
        import threading

        def run_stm_download():
            # Start job tracking
            job_id = None
            try:
                job_id = job_history_manager.start_job(
                    job_type='download',
                    job_subtype='statement',
                    parameters={
                        'year': year,
                        'month': month,
                        'scheme': scheme,
                        'person_type': person_type,
                        'auto_import': auto_import
                    },
                    triggered_by='manual'
                )
            except Exception as e:
                app.logger.warning(f"Could not start job tracking: {e}")

            cmd = ['python3', 'stm_downloader_http.py', '--year', str(year), '--scheme', scheme, '--type', person_type]
            if month:
                cmd.extend(['--month', str(int(month))])

            log_file = Path('logs') / f"stm_download_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
            log_file.parent.mkdir(exist_ok=True)

            try:
                with open(log_file, 'w') as f:
                    result = subprocess.run(cmd, stdout=f, stderr=subprocess.STDOUT)

                # Auto-import if enabled
                import_results = None
                if auto_import:
                    import_log_file = Path('logs') / f"stm_import_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
                    with open(import_log_file, 'w') as f:
                        subprocess.run(['python3', 'stm_import.py', 'downloads/stm/'], stdout=f, stderr=subprocess.STDOUT)

                # Complete job
                if job_id:
                    try:
                        job_history_manager.complete_job(
                            job_id=job_id,
                            status='completed',
                            results={
                                'scheme': scheme,
                                'year': year,
                                'month': month,
                                'person_type': person_type,
                                'auto_import': auto_import
                            }
                        )
                    except Exception as e:
                        app.logger.warning(f"Could not complete job tracking: {e}")

            except Exception as e:
                if job_id:
                    try:
                        job_history_manager.complete_job(
                            job_id=job_id,
                            status='failed',
                            error_message=str(e)
                        )
                    except Exception:
                        pass
                raise

        thread = threading.Thread(target=run_stm_download)
        thread.start()

        return jsonify({
            'success': True,
            'message': f'STM download started for {scheme.upper()} year {year}' + (' with auto-import' if auto_import else '')
        })

    except Exception as e:
        app.logger.error(f"STM download error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/stm/history')
def get_stm_history():
    """Get STM download history from database"""
    try:
        from utils.download_history_db import DownloadHistoryDB

        db = DownloadHistoryDB()
        db.connect()

        # Get recent downloads (last 50)
        downloads = db.get_recent_downloads('stm', limit=50)
        stats = db.get_statistics('stm')

        db.disconnect()

        return jsonify({
            'success': True,
            'downloads': downloads,
            'total': stats.get('total_downloads', 0),
            'last_download': downloads[0]['downloaded_at'] if downloads else None
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/api/stm/stats')
def get_stm_stats():
    """Get Statement files statistics with optional filtering"""
    import re

    # Get filter params
    fiscal_year = request.args.get('fiscal_year', type=int)
    start_month = request.args.get('start_month', type=int)
    end_month = request.args.get('end_month', type=int)
    filter_status = request.args.get('status', '').strip().lower()

    # Calculate start/end year from fiscal year
    start_year = None
    end_year = None
    if fiscal_year:
        if start_month and start_month >= 10:
            start_year = fiscal_year - 1
        else:
            start_year = fiscal_year
        if end_month and end_month <= 9:
            end_year = fiscal_year
        else:
            end_year = fiscal_year - 1

    try:
        download_dir = Path('downloads/stm')
        stm_files = []
        total_size = 0
        imported_count = 0
        pending_count = 0

        # Get imported file list from database
        imported_filenames = set()
        import_info = {}
        try:
            
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute(
                """SELECT filename, status, imported_records, total_records,
                          import_completed_at, file_type, scheme
                   FROM stm_imported_files"""
            )
            for row in cursor.fetchall():
                imported_filenames.add(row[0])
                import_info[row[0]] = {
                    'status': row[1],
                    'imported_records': row[2],
                    'total_records': row[3],
                    'import_completed_at': row[4].isoformat() if row[4] else None,
                    'file_type': row[5],
                    'scheme': row[6]
                }
            cursor.close()
            conn.close()
        except Exception as e:
            app.logger.warning(f"Could not fetch STM import status: {e}")

        # Helper to convert month/year to comparable number
        def date_to_num(m, y):
            return y * 12 + m

        if download_dir.exists():
            for f in download_dir.glob('STM_*.xls'):
                # Parse date from filename: STM_10670_OPUCS256812_02.xls
                # 256812 = year 2568, month 12
                file_year = None
                file_month = None
                match = re.search(r'(\d{4})(\d{2})_\d+\.xls$', f.name)
                if match:
                    file_year = int(match.group(1))
                    file_month = int(match.group(2))

                # Apply date filter
                if fiscal_year and file_year and file_month:
                    file_date_num = date_to_num(file_month, file_year)
                    if start_month and start_year:
                        start_num = date_to_num(start_month, start_year)
                        if file_date_num < start_num:
                            continue
                    if end_month and end_year:
                        end_num = date_to_num(end_month, end_year)
                        if file_date_num > end_num:
                            continue

                stat = f.stat()
                file_size = stat.st_size

                is_imported = f.name in imported_filenames and import_info.get(f.name, {}).get('status') == 'completed'

                # Apply status filter
                if filter_status:
                    if filter_status == 'imported' and not is_imported:
                        continue
                    if filter_status == 'pending' and is_imported:
                        continue

                total_size += file_size

                if is_imported:
                    imported_count += 1
                else:
                    pending_count += 1

                file_info = {
                    'filename': f.name,
                    'size': file_size,
                    'size_formatted': humanize.naturalsize(file_size),
                    'modified': datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    'is_imported': is_imported,
                    'file_year': file_year,
                    'file_month': file_month
                }

                # Add import details if available
                if f.name in import_info:
                    file_info.update(import_info[f.name])

                stm_files.append(file_info)

        # Sort by modified date desc
        stm_files.sort(key=lambda x: x['modified'], reverse=True)

        return jsonify({
            'success': True,
            'total_files': len(stm_files),
            'imported_count': imported_count,
            'pending_count': pending_count,
            'total_size': humanize.naturalsize(total_size),
            'total_size_bytes': total_size,
            'files': stm_files[:100]
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/api/stm/clear', methods=['POST'])
def clear_stm_files():
    """Clear all Statement files"""
    try:
        download_dir = Path('downloads/stm')
        deleted_count = 0

        # Delete all STM import records first
        try:
            
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM stm_imported_files")
            conn.commit()
            cursor.close()
            conn.close()
        except Exception as e:
            app.logger.warning(f"Could not clear STM import records: {e}")

        # Delete files
        for f in download_dir.glob('STM_*.xls'):
            f.unlink()
            deleted_count += 1

        return jsonify({
            'success': True,
            'message': f'Deleted {deleted_count} Statement files'
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/stm/records')
def get_stm_records():
    """Get Statement database records with reconciliation status using optimized view"""
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 50))
        offset = (page - 1) * limit

        view_mode = request.args.get('view_mode', 'rep')  # 'rep' or 'tran'
        fiscal_year = request.args.get('fiscal_year', '')
        rep_no = request.args.get('rep_no', '')
        status = request.args.get('status', '')

        conn = get_db_connection()
        cursor = conn.cursor()

        # Build WHERE clause using the v_stm_rep_reconciliation view
        where_clauses = []
        params = []

        if fiscal_year:
            # Fiscal year filter
            fy = int(fiscal_year)
            where_clauses.append("""
                ((statement_year = %s AND statement_month >= 10)
                 OR (statement_year = %s AND statement_month <= 9))
            """)
            params.extend([fy - 1, fy])

        if rep_no:
            where_clauses.append("rep_repno LIKE %s")
            params.append(f"%{rep_no}%")

        if status:
            where_clauses.append("reconcile_status = %s")
            params.append(status)

        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

        if view_mode == 'rep':
            # Group by REP No - use aggregated query on the view
            count_sql = f"""
                SELECT COUNT(DISTINCT rep_repno)
                FROM v_stm_rep_reconciliation
                WHERE {where_sql}
            """
            cursor.execute(count_sql, params)
            total = cursor.fetchone()[0]

            query = f"""
                SELECT
                    rep_repno as rep_no,
                    COUNT(*) as count,
                    SUM(stm_compensation) as stm_amount,
                    SUM(rep_reimb_nhso) as rep_amount,
                    CASE
                        WHEN SUM(CASE WHEN reconcile_status = 'matched' THEN 1 ELSE 0 END) = COUNT(*) THEN 'matched'
                        WHEN SUM(CASE WHEN reconcile_status IN ('amount_diff', 'diff_amount') THEN 1 ELSE 0 END) > 0 THEN 'diff_amount'
                        ELSE 'stm_only'
                    END as status
                FROM v_stm_rep_reconciliation
                WHERE {where_sql}
                GROUP BY rep_repno
                ORDER BY rep_repno DESC
                LIMIT %s OFFSET %s
            """
            cursor.execute(query, params + [limit, offset])
        else:
            # Individual transactions from view
            count_sql = f"""
                SELECT COUNT(*)
                FROM v_stm_rep_reconciliation
                WHERE {where_sql}
            """
            cursor.execute(count_sql, params)
            total = cursor.fetchone()[0]

            query = f"""
                SELECT
                    tran_id,
                    rep_repno as rep_no,
                    patient_name,
                    hn,
                    stm_compensation as stm_amount,
                    rep_reimb_nhso as rep_amount,
                    reconcile_status as status
                FROM v_stm_rep_reconciliation
                WHERE {where_sql}
                ORDER BY rep_repno DESC, tran_id
                LIMIT %s OFFSET %s
            """
            cursor.execute(query, params + [limit, offset])

        columns = [desc[0] for desc in cursor.description]
        records = [dict(zip(columns, row)) for row in cursor.fetchall()]

        # Convert Decimal to float for JSON serialization
        for rec in records:
            for key in ['stm_amount', 'rep_amount', 'count']:
                if key in rec and rec[key] is not None:
                    rec[key] = float(rec[key])

        # Get stats from view (much faster than correlated subquery)
        stats_sql = """
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN reconcile_status = 'matched' THEN 1 ELSE 0 END) as matched,
                SUM(CASE WHEN reconcile_status IN ('amount_diff', 'diff_amount') THEN 1 ELSE 0 END) as diff_amount,
                SUM(CASE WHEN reconcile_status = 'stm_only' OR reconcile_status IS NULL THEN 1 ELSE 0 END) as stm_only
            FROM v_stm_rep_reconciliation
        """
        cursor.execute(stats_sql)
        stats_row = cursor.fetchone()
        stats = {
            'total': int(stats_row[0] or 0),
            'matched': int(stats_row[1] or 0),
            'diff_amount': int(stats_row[2] or 0),
            'stm_only': int(stats_row[3] or 0)
        }

        cursor.close()
        conn.close()

        total_pages = (total + limit - 1) // limit

        return jsonify({
            'success': True,
            'records': records,
            'stats': stats,
            'pagination': {
                'page': page,
                'limit': limit,
                'total': total,
                'total_pages': total_pages
            }
        })

    except Exception as e:
        app.logger.error(f"Error fetching STM records: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/stm/clear-database', methods=['POST'])
def clear_stm_database():
    """Clear all Statement records from database (keep files)"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Delete claim items first (foreign key constraint)
        cursor.execute("DELETE FROM stm_claim_item")
        cursor.execute("DELETE FROM stm_rep_summary")
        cursor.execute("DELETE FROM stm_receivable_summary")
        cursor.execute("DELETE FROM stm_imported_files")

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'message': 'Statement database cleared successfully'
        })

    except Exception as e:
        app.logger.error(f"Error clearing STM database: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/rep/file-type-stats')
def get_file_type_stats():
    """Get file statistics grouped by file type"""
    try:
        import re
        from collections import defaultdict

        # Get all files from downloads directory
        all_files = file_manager.history_manager.get_all_downloads()

        # Get import status from database
        import_status_map = {}
        try:
            with get_db_connection() as conn:
                if conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT filename, status FROM eclaim_imported_files")
                    for row in cursor.fetchall():
                        import_status_map[row[0]] = row[1]
                    cursor.close()
        except Exception as e:
            app.logger.warning(f"Could not get import status: {e}")

        # File type definitions with descriptions
        file_type_info = {
            'OP': {'name': 'OP', 'description': 'ผู้ป่วยนอก (Outpatient)', 'category': 'ucs', 'icon': '🏥'},
            'IP': {'name': 'IP', 'description': 'ผู้ป่วยใน (Inpatient)', 'category': 'ucs', 'icon': '🛏️'},
            'OPLGO': {'name': 'OP-LGO', 'description': 'ผู้ป่วยนอก อปท.', 'category': 'lgo', 'icon': '🏛️'},
            'IPLGO': {'name': 'IP-LGO', 'description': 'ผู้ป่วยใน อปท.', 'category': 'lgo', 'icon': '🏛️'},
            'OPSSS': {'name': 'OP-SSS', 'description': 'ผู้ป่วยนอก ประกันสังคม', 'category': 'sss', 'icon': '👷'},
            'IPSSS': {'name': 'IP-SSS', 'description': 'ผู้ป่วยใน ประกันสังคม', 'category': 'sss', 'icon': '👷'},
            'ORF': {'name': 'ORF', 'description': 'Outpatient Referral', 'category': 'special', 'icon': '🔄'},
            'IP_APPEAL': {'name': 'IP Appeal', 'description': 'อุทธรณ์ผู้ป่วยใน', 'category': 'appeal', 'icon': '📝'},
            'IP_APPEAL_NHSO': {'name': 'IP Appeal NHSO', 'description': 'อุทธรณ์ ผป.ใน (ฝั่ง สปสช.)', 'category': 'appeal', 'icon': '📝'},
            'OP_APPEAL': {'name': 'OP Appeal', 'description': 'อุทธรณ์ผู้ป่วยนอก', 'category': 'appeal', 'icon': '📋'},
            'OP_APPEAL_CD': {'name': 'OP Appeal CD', 'description': 'อุทธรณ์ ผป.นอก (โรคเรื้อรัง)', 'category': 'appeal', 'icon': '📋'},
        }

        # Parse file types from filenames
        type_stats = defaultdict(lambda: {'count': 0, 'imported': 0, 'pending': 0, 'size': 0})
        pattern = re.compile(r'eclaim_\d+_([A-Z_]+)_\d{8}_\d+\.xls')

        for file_info in all_files:
            filename = file_info.get('filename', '')
            match = pattern.match(filename)
            if match:
                file_type = match.group(1)
                type_stats[file_type]['count'] += 1
                type_stats[file_type]['size'] += file_info.get('size', 0)

                # Check import status
                status = import_status_map.get(filename, 'pending')
                if status == 'completed':
                    type_stats[file_type]['imported'] += 1
                else:
                    type_stats[file_type]['pending'] += 1

        # Build response with descriptions
        result = []
        for file_type, stats in sorted(type_stats.items(), key=lambda x: x[1]['count'], reverse=True):
            info = file_type_info.get(file_type, {
                'name': file_type,
                'description': f'Unknown type: {file_type}',
                'category': 'unknown',
                'icon': '📄'
            })
            result.append({
                'type': file_type,
                'name': info['name'],
                'description': info['description'],
                'category': info['category'],
                'icon': info['icon'],
                'count': stats['count'],
                'imported': stats['imported'],
                'pending': stats['pending'],
                'size': stats['size'],
                'size_formatted': humanize.naturalsize(stats['size'])
            })

        return jsonify({
            'success': True,
            'file_types': result,
            'total_types': len(result)
        })

    except Exception as e:
        app.logger.error(f"Error getting file type stats: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/rep/records')
def get_rep_records():
    """Get REP database records with reconciliation status"""
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 50))
        offset = (page - 1) * limit

        view_mode = request.args.get('view_mode', 'rep')  # 'rep' or 'tran'
        fiscal_year = request.args.get('fiscal_year', '')
        rep_no = request.args.get('rep_no', '')
        tran_id = request.args.get('tran_id', '')
        status = request.args.get('status', '')

        conn = get_db_connection()
        cursor = conn.cursor()

        # Build WHERE clause
        where_clauses = []
        params = []

        if fiscal_year:
            # Fiscal year filter based on dateadm (admission date)
            # FY 2569 = Oct 2568 to Sep 2569 in Thai calendar = Oct 2025 to Sep 2026 in Gregorian
            fy = int(fiscal_year)
            start_date = f"{fy - 544}-10-01"  # Convert BE to CE: 2569-544=2025
            end_date = f"{fy - 543}-09-30"
            where_clauses.append("c.dateadm BETWEEN %s AND %s")
            params.extend([start_date, end_date])

        if rep_no:
            where_clauses.append("c.rep_no LIKE %s")
            params.append(f"%{rep_no}%")

        if tran_id:
            where_clauses.append("c.tran_id LIKE %s")
            params.append(f"%{tran_id}%")

        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

        if view_mode == 'rep':
            # Group by REP No
            count_sql = f"""
                SELECT COUNT(DISTINCT c.rep_no)
                FROM claim_rep_opip_nhso_item c
                WHERE {where_sql}
            """
            cursor.execute(count_sql, params)
            total = cursor.fetchone()[0]

            query = f"""
                SELECT
                    c.rep_no,
                    COUNT(*) as count,
                    SUM(COALESCE(c.reimb_nhso, 0)) as rep_amount,
                    (
                        SELECT SUM(COALESCE(s.paid_after_deduction, 0))
                        FROM stm_claim_item s
                        WHERE s.rep_no = c.rep_no
                    ) as stm_amount,
                    CASE
                        WHEN (SELECT COUNT(*) FROM stm_claim_item s WHERE s.rep_no = c.rep_no) = 0 THEN 'rep_only'
                        WHEN ABS(SUM(COALESCE(c.reimb_nhso, 0)) - COALESCE((SELECT SUM(COALESCE(s.paid_after_deduction, 0)) FROM stm_claim_item s WHERE s.rep_no = c.rep_no), 0)) < 1 THEN 'matched'
                        ELSE 'diff_amount'
                    END as status
                FROM claim_rep_opip_nhso_item c
                WHERE {where_sql}
                GROUP BY c.rep_no
                ORDER BY c.rep_no DESC
                LIMIT %s OFFSET %s
            """
            cursor.execute(query, params + [limit, offset])
        else:
            # Individual transactions
            count_sql = f"""
                SELECT COUNT(*)
                FROM claim_rep_opip_nhso_item c
                WHERE {where_sql}
            """
            cursor.execute(count_sql, params)
            total = cursor.fetchone()[0]

            query = f"""
                SELECT
                    c.tran_id,
                    c.rep_no,
                    c.hn,
                    COALESCE(c.reimb_nhso, 0) as rep_amount,
                    (
                        SELECT COALESCE(s.paid_after_deduction, 0)
                        FROM stm_claim_item s
                        WHERE s.tran_id = c.tran_id
                        LIMIT 1
                    ) as stm_amount,
                    CASE
                        WHEN NOT EXISTS (SELECT 1 FROM stm_claim_item s WHERE s.tran_id = c.tran_id) THEN 'rep_only'
                        WHEN ABS(COALESCE(c.reimb_nhso, 0) - COALESCE((SELECT s.paid_after_deduction FROM stm_claim_item s WHERE s.tran_id = c.tran_id LIMIT 1), 0)) < 1 THEN 'matched'
                        ELSE 'diff_amount'
                    END as status
                FROM claim_rep_opip_nhso_item c
                WHERE {where_sql}
                ORDER BY c.rep_no DESC, c.tran_id
                LIMIT %s OFFSET %s
            """
            cursor.execute(query, params + [limit, offset])

        columns = [desc[0] for desc in cursor.description]
        records = [dict(zip(columns, row)) for row in cursor.fetchall()]

        # Convert Decimal to float for JSON serialization
        for rec in records:
            for key in ['rep_amount', 'stm_amount', 'count']:
                if key in rec and rec[key] is not None:
                    rec[key] = float(rec[key])

        # Apply status filter after fetch (for complex status calculation)
        if status:
            records = [r for r in records if r.get('status') == status]

        # Get stats - count by checking if STM exists
        stats_sql = """
            SELECT
                COUNT(*) as total,
                COUNT(CASE WHEN EXISTS (
                    SELECT 1 FROM stm_claim_item s
                    WHERE s.tran_id = c.tran_id
                    AND ABS(COALESCE(s.paid_after_deduction, 0) - COALESCE(c.reimb_nhso, 0)) < 1
                ) THEN 1 END) as matched,
                COUNT(CASE WHEN EXISTS (
                    SELECT 1 FROM stm_claim_item s
                    WHERE s.tran_id = c.tran_id
                    AND ABS(COALESCE(s.paid_after_deduction, 0) - COALESCE(c.reimb_nhso, 0)) >= 1
                ) THEN 1 END) as diff_amount
            FROM claim_rep_opip_nhso_item c
        """
        cursor.execute(stats_sql)
        stats_row = cursor.fetchone()
        total_all = stats_row[0] or 0
        matched = stats_row[1] or 0
        diff_amount = stats_row[2] or 0
        rep_only = total_all - matched - diff_amount

        stats = {
            'total': total_all,
            'matched': matched,
            'diff_amount': diff_amount,
            'rep_only': rep_only
        }

        cursor.close()
        conn.close()

        total_pages = (total + limit - 1) // limit

        return jsonify({
            'success': True,
            'records': records,
            'stats': stats,
            'pagination': {
                'page': page,
                'limit': limit,
                'total': total,
                'total_pages': total_pages
            }
        })

    except Exception as e:
        app.logger.error(f"Error fetching REP records: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/rep/clear-database', methods=['POST'])
def clear_rep_database():
    """Clear all REP records from database (keep files)"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Delete REP records
        cursor.execute("DELETE FROM claim_rep_opip_nhso_item")
        cursor.execute("DELETE FROM claim_rep_orf_nhso_item")
        cursor.execute("DELETE FROM eclaim_imported_files")

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'message': 'REP database cleared successfully'
        })

    except Exception as e:
        app.logger.error(f"Error clearing REP database: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
def naturaltime_filter(value):
    """Template filter for human-readable timestamps"""
    try:
        dt = datetime.fromisoformat(value)
        return humanize.naturaltime(dt)
    except (ValueError, TypeError, AttributeError):
        return value


@app.template_filter('number_format')
def number_format_filter(value):
    """Template filter for number formatting with commas"""
    try:
        return "{:,}".format(int(value))
    except (ValueError, TypeError):
        return value


@app.route('/api/clear-all', methods=['POST'])
def clear_all_data():
    """Clear all data: files, history, and database (DANGER!)"""
    try:
        # 1. Delete all files in downloads subdirectories
        deleted_files = 0
        for subdir in ['rep', 'stm', 'smt']:
            subdir_path = Path('downloads') / subdir
            if subdir_path.exists():
                for file in subdir_path.glob('*.*'):
                    if file.is_file():
                        file.unlink()
                        deleted_files += 1

        # 2. Reset download history
        history_manager.save_history({'last_run': None, 'downloads': []})

        # 3. Clear database tables
        conn = get_db_connection()
        if conn:
            try:
                cursor = conn.cursor()
                cursor.execute("TRUNCATE TABLE claim_rep_opip_nhso_item, claim_rep_orf_nhso_item, eclaim_imported_files RESTART IDENTITY CASCADE;")
                conn.commit()
                cursor.close()
                conn.close()
            except Exception as e:
                app.logger.error(f"Database clear error: {e}")
                if conn:
                    conn.close()
                return jsonify({
                    'success': False,
                    'error': f'Database clear failed: {str(e)}'
                }), 500

        # 4. Clear realtime logs
        log_streamer.clear_logs()

        return jsonify({
            'success': True,
            'deleted_files': deleted_files,
            'message': 'All data cleared successfully'
        }), 200

    except Exception as e:
        app.logger.error(f"Clear all data error: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/logs/stream')
def stream_logs():
    """Stream real-time logs via Server-Sent Events (SSE)"""
    def generate():
        try:
            for log_entry in log_streamer.stream_logs(tail=50):
                yield log_entry
        except GeneratorExit:
            pass

    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',
            'Connection': 'keep-alive'
        }
    )


@app.route('/api/logs/clear', methods=['POST'])
def clear_logs():
    """Clear realtime log file"""
    try:
        log_streamer.clear_logs()
        return jsonify({'success': True, 'message': 'Logs cleared'}), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== Job History API ====================

@app.route('/api/jobs')
def get_jobs():
    """Get recent job history"""
    try:
        job_type = request.args.get('type')  # download, import, schedule
        status = request.args.get('status')  # running, completed, failed
        limit = request.args.get('limit', 50, type=int)
        date_from = request.args.get('date_from')  # YYYY-MM-DD
        date_to = request.args.get('date_to')  # YYYY-MM-DD

        jobs = job_history_manager.get_recent_jobs(
            job_type=job_type,
            status=status,
            limit=min(limit, 200),
            date_from=date_from,
            date_to=date_to
        )

        return jsonify({
            'success': True,
            'jobs': jobs,
            'total': len(jobs)
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/jobs/stats')
def get_job_stats():
    """Get job statistics"""
    try:
        days = request.args.get('days', 7, type=int)
        stats = job_history_manager.get_job_stats(days=min(days, 30))

        return jsonify({
            'success': True,
            'stats': stats,
            'period_days': days
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/jobs/<job_id>')
def get_job_detail(job_id):
    """Get specific job details"""
    try:
        jobs = job_history_manager.get_recent_jobs(limit=500)
        job = next((j for j in jobs if j['job_id'] == job_id), None)

        if not job:
            return jsonify({'success': False, 'error': 'Job not found'}), 404

        return jsonify({
            'success': True,
            'job': job
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== System Health Dashboard ====================

@app.route('/api/system/health')
def get_system_health():
    """
    Comprehensive system health dashboard
    Returns status of all system components
    """
    import psutil
    import shutil
    from pathlib import Path
    from config.database import DOWNLOADS_DIR

    health = {
        'success': True,
        'timestamp': datetime.now(TZ_BANGKOK).isoformat(),
        'overall_status': 'healthy',  # healthy, warning, critical
        'components': {}
    }

    issues = []

    # === 1. Database Connection ===
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("SELECT 1")

            # Get database type
            db_type_display = 'PostgreSQL' if DB_TYPE == 'postgresql' else 'MySQL'

            cursor.close()
            conn.close()

            health['components']['database'] = {
                'status': 'healthy',
                'message': 'Connected',
                'type': db_type_display
            }
        else:
            health['components']['database'] = {
                'status': 'critical',
                'message': 'Cannot connect to database'
            }
            issues.append('database')
    except Exception as e:
        health['components']['database'] = {
            'status': 'critical',
            'message': f'Database error: {str(e)}'
        }
        issues.append('database')

    # === 2. Disk Space ===
    try:
        downloads_dir = Path(DOWNLOADS_DIR)
        if downloads_dir.exists():
            disk = shutil.disk_usage(str(downloads_dir))
            free_gb = disk.free / (1024**3)
            total_gb = disk.total / (1024**3)
            used_percent = (disk.used / disk.total) * 100

            disk_status = 'healthy'
            if used_percent > 90:
                disk_status = 'critical'
                issues.append('disk')
            elif used_percent > 80:
                disk_status = 'warning'

            health['components']['disk'] = {
                'status': disk_status,
                'free_gb': round(free_gb, 2),
                'total_gb': round(total_gb, 2),
                'used_percent': round(used_percent, 1),
                'message': f'{round(free_gb, 1)} GB free ({round(100-used_percent, 1)}%)'
            }
        else:
            health['components']['disk'] = {
                'status': 'warning',
                'message': 'Downloads directory not found'
            }
    except Exception as e:
        health['components']['disk'] = {
            'status': 'warning',
            'message': f'Cannot check disk: {str(e)}'
        }

    # === 3. Running Processes ===
    pid_files = {
        'downloader': Path('/tmp/eclaim_downloader.pid'),
        'import': Path('/tmp/eclaim_import.pid'),
        'parallel': Path('/tmp/eclaim_parallel_download.pid'),
        'stm': Path('/tmp/eclaim_stm_downloader.pid'),
        'smt': Path('/tmp/eclaim_smt_fetch.pid')
    }

    running_processes = []
    stale_processes = []

    for name, pid_file in pid_files.items():
        if pid_file.exists():
            try:
                pid = int(pid_file.read_text().strip())
                if psutil.pid_exists(pid):
                    proc = psutil.Process(pid)
                    running_processes.append({
                        'name': name,
                        'pid': pid,
                        'status': proc.status(),
                        'started': datetime.fromtimestamp(proc.create_time()).isoformat()
                    })
                else:
                    stale_processes.append(name)
            except (ValueError, psutil.NoSuchProcess):
                stale_processes.append(name)

    process_status = 'healthy'
    if stale_processes:
        process_status = 'warning'

    health['components']['processes'] = {
        'status': process_status,
        'running': running_processes,
        'running_count': len(running_processes),
        'stale_pids': stale_processes,
        'message': f'{len(running_processes)} active' + (f', {len(stale_processes)} stale PIDs' if stale_processes else '')
    }

    # === 4. Recent Jobs (last 24 hours) ===
    try:
        recent_jobs = job_history_manager.get_recent_jobs(limit=100)
        now = datetime.now()

        jobs_24h = {
            'total': 0,
            'running': 0,
            'completed': 0,
            'failed': 0
        }

        for job in recent_jobs:
            try:
                started = datetime.fromisoformat(job['started_at'].replace('Z', '+00:00')) if isinstance(job['started_at'], str) else job['started_at']
                if started.tzinfo:
                    started = started.replace(tzinfo=None)
                if (now - started).total_seconds() < 86400:  # 24 hours
                    jobs_24h['total'] += 1
                    if job['status'] == 'running':
                        jobs_24h['running'] += 1
                    elif job['status'] in ('completed', 'completed_with_errors'):
                        jobs_24h['completed'] += 1
                    elif job['status'] == 'failed':
                        jobs_24h['failed'] += 1
            except (ValueError, TypeError):
                pass

        jobs_status = 'healthy'
        if jobs_24h['failed'] > jobs_24h['completed']:
            jobs_status = 'critical'
            issues.append('jobs')
        elif jobs_24h['failed'] > 0:
            jobs_status = 'warning'

        health['components']['jobs'] = {
            'status': jobs_status,
            'last_24h': jobs_24h,
            'message': f"{jobs_24h['completed']} completed, {jobs_24h['failed']} failed in 24h"
        }
    except Exception as e:
        health['components']['jobs'] = {
            'status': 'warning',
            'message': f'Cannot check jobs: {str(e)}'
        }

    # === 5. Files Statistics ===
    try:
        downloads_dir = Path(DOWNLOADS_DIR)
        if downloads_dir.exists():
            # Search recursively in subdirectories (rep/, stm/, smt/)
            xls_files = list(downloads_dir.glob('**/*.xls'))
            total_size = sum(f.stat().st_size for f in xls_files if f.exists())

            health['components']['files'] = {
                'status': 'healthy',
                'count': len(xls_files),
                'total_size_mb': round(total_size / (1024**2), 2),
                'message': f'{len(xls_files)} files ({round(total_size / (1024**2), 1)} MB)'
            }
        else:
            health['components']['files'] = {
                'status': 'warning',
                'count': 0,
                'message': 'Downloads directory not found'
            }
    except Exception as e:
        health['components']['files'] = {
            'status': 'warning',
            'message': f'Cannot check files: {str(e)}'
        }

    # === 6. Memory Usage ===
    try:
        memory = psutil.virtual_memory()
        mem_status = 'healthy'
        if memory.percent > 90:
            mem_status = 'critical'
            issues.append('memory')
        elif memory.percent > 80:
            mem_status = 'warning'

        health['components']['memory'] = {
            'status': mem_status,
            'used_percent': round(memory.percent, 1),
            'available_gb': round(memory.available / (1024**3), 2),
            'message': f'{round(memory.percent, 1)}% used, {round(memory.available / (1024**3), 1)} GB available'
        }
    except Exception as e:
        health['components']['memory'] = {
            'status': 'warning',
            'message': f'Cannot check memory: {str(e)}'
        }

    # === Determine Overall Status ===
    if 'database' in issues or 'memory' in issues:
        health['overall_status'] = 'critical'
    elif 'disk' in issues or 'jobs' in issues:
        health['overall_status'] = 'warning'
    elif any(c.get('status') == 'warning' for c in health['components'].values()):
        health['overall_status'] = 'warning'

    health['issues'] = issues

    return jsonify(health)


# ==================== Alert System API ====================

@app.route('/api/alerts')
def get_alerts():
    """Get system alerts"""
    try:
        include_dismissed = request.args.get('include_dismissed', 'false').lower() == 'true'
        alert_type = request.args.get('type')
        severity = request.args.get('severity')
        limit = request.args.get('limit', 50, type=int)

        alerts = alert_manager.get_alerts(
            include_dismissed=include_dismissed,
            alert_type=alert_type,
            severity=severity,
            limit=min(limit, 200)
        )

        return jsonify({
            'success': True,
            'alerts': alerts,
            'total': len(alerts)
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/alerts/unread-count')
def get_alerts_unread_count():
    """Get count of unread alerts"""
    try:
        count = alert_manager.get_unread_count()
        return jsonify({
            'success': True,
            'count': count
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/alerts/<int:alert_id>/read', methods=['POST'])
def mark_alert_read(alert_id):
    """Mark an alert as read"""
    try:
        success = alert_manager.mark_as_read(alert_id)
        return jsonify({'success': success})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/alerts/read-all', methods=['POST'])
def mark_all_alerts_read():
    """Mark all alerts as read"""
    try:
        affected = alert_manager.mark_all_as_read()
        return jsonify({
            'success': True,
            'affected': affected
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/alerts/<int:alert_id>/dismiss', methods=['POST'])
def dismiss_alert(alert_id):
    """Dismiss an alert"""
    try:
        success = alert_manager.dismiss_alert(alert_id)
        return jsonify({'success': success})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/alerts/dismiss-all', methods=['POST'])
def dismiss_all_alerts():
    """Dismiss all alerts"""
    try:
        affected = alert_manager.dismiss_all()
        return jsonify({
            'success': True,
            'affected': affected
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/alerts/check-health', methods=['POST'])
def check_health_and_alert():
    """
    Check system health and create alerts for any issues found.
    This can be called periodically or manually to update alerts.
    """
    import psutil
    import shutil
    from pathlib import Path
    from config.database import DOWNLOADS_DIR

    alerts_created = []

    try:
        # Check disk space
        downloads_dir = Path(DOWNLOADS_DIR)
        if downloads_dir.exists():
            disk = shutil.disk_usage(str(downloads_dir))
            used_percent = (disk.used / disk.total) * 100
            free_gb = disk.free / (1024**3)

            if used_percent > 80:
                alert_id = alert_manager.alert_disk_warning(used_percent, free_gb)
                if alert_id:
                    alerts_created.append({'type': 'disk_warning', 'id': alert_id})

        # Check memory
        memory = psutil.virtual_memory()
        if memory.percent > 80:
            available_gb = memory.available / (1024**3)
            alert_id = alert_manager.alert_memory_warning(memory.percent, available_gb)
            if alert_id:
                alerts_created.append({'type': 'memory_warning', 'id': alert_id})

        # Check stale processes
        pid_files = {
            'downloader': Path('/tmp/eclaim_downloader.pid'),
            'import': Path('/tmp/eclaim_import.pid'),
            'parallel': Path('/tmp/eclaim_parallel_download.pid'),
        }

        stale_processes = []
        for name, pid_file in pid_files.items():
            if pid_file.exists():
                try:
                    pid = int(pid_file.read_text().strip())
                    if not psutil.pid_exists(pid):
                        stale_processes.append(name)
                except (ValueError, psutil.NoSuchProcess):
                    stale_processes.append(name)

        if stale_processes:
            alert_id = alert_manager.alert_stale_process(stale_processes)
            if alert_id:
                alerts_created.append({'type': 'stale_process', 'id': alert_id})

        return jsonify({
            'success': True,
            'alerts_created': len(alerts_created),
            'details': alerts_created
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== Seed Data Initialization ====================

# Global progress tracking for seed initialization
seed_progress = {
    'running': False,
    'current_task': None,
    'tasks': [],
    'completed': 0,
    'total': 0,
    'error': None
}


@app.route('/api/system/seed-status')
def get_seed_status():
    """
    Check if seed data initialization is needed.
    Returns status of each seed data table.
    """
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Check each seed table
        seed_tables = {
            'dim_date': {'name': 'Dimension Date', 'min_records': 100},
            'health_offices': {'name': 'Health Offices', 'min_records': 1000},
            'nhso_error_codes': {'name': 'NHSO Error Codes', 'min_records': 10},
            'fund_types': {'name': 'Fund Types', 'min_records': 5},
            'service_types': {'name': 'Service Types', 'min_records': 5}
        }

        results = {}
        needs_init = False

        for table, info in seed_tables.items():
            try:
                cursor.execute(f"SELECT COUNT(*) FROM {table}")
                count = cursor.fetchone()[0]
                is_empty = count < info['min_records']
                results[table] = {
                    'name': info['name'],
                    'count': count,
                    'is_empty': is_empty,
                    'min_required': info['min_records']
                }
                if is_empty:
                    needs_init = True
            except Exception as e:
                results[table] = {
                    'name': info['name'],
                    'count': 0,
                    'is_empty': True,
                    'error': str(e)
                }
                needs_init = True

        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'needs_initialization': needs_init,
            'tables': results,
            'seed_running': seed_progress['running'],
            'seed_progress': seed_progress if seed_progress['running'] else None
        })

    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/system/seed-init', methods=['POST'])
def run_seed_initialization():
    """
    Start seed data initialization process.
    Runs in background and updates progress.
    """
    global seed_progress

    if seed_progress['running']:
        return jsonify({
            'success': False,
            'error': 'Seed initialization already running',
            'progress': seed_progress
        }), 400

    # Start seed process in background thread
    import threading

    def run_seeds():
        global seed_progress
        seed_progress = {
            'running': True,
            'current_task': None,
            'tasks': [
                {'id': 'dim', 'name': 'Dimension Tables', 'status': 'pending', 'records': 0},
                {'id': 'health', 'name': 'Health Offices', 'status': 'pending', 'records': 0},
                {'id': 'errors', 'name': 'NHSO Error Codes', 'status': 'pending', 'records': 0}
            ],
            'completed': 0,
            'total': 3,
            'error': None,
            'started_at': datetime.now(TZ_BANGKOK).isoformat()
        }

        try:
            # Task 1: Dimension tables (migrate.py --seed)
            seed_progress['current_task'] = 'dim'
            seed_progress['tasks'][0]['status'] = 'running'

            import subprocess
            import os
            cwd = os.environ.get('APP_ROOT', os.path.dirname(os.path.abspath(__file__)))
            result = subprocess.run(
                ['python', 'database/migrate.py', '--seed'],
                capture_output=True,
                text=True,
                cwd=cwd
            )
            if result.returncode != 0:
                raise Exception(f"Dimension seed failed: {result.stderr}")

            seed_progress['tasks'][0]['status'] = 'completed'
            seed_progress['tasks'][0]['records'] = 2600  # Approximate
            seed_progress['completed'] = 1

            # Task 2: Health Offices
            seed_progress['current_task'] = 'health'
            seed_progress['tasks'][1]['status'] = 'running'

            result = subprocess.run(
                ['python', 'database/seeds/health_offices_importer.py'],
                capture_output=True,
                text=True,
                cwd=cwd
            )
            if result.returncode != 0:
                raise Exception(f"Health offices seed failed: {result.stderr}")

            # Parse record count from output
            import re
            match = re.search(r'Imported: (\d+)', result.stdout)
            records = int(match.group(1)) if match else 43884

            seed_progress['tasks'][1]['status'] = 'completed'
            seed_progress['tasks'][1]['records'] = records
            seed_progress['completed'] = 2

            # Task 3: NHSO Error Codes
            seed_progress['current_task'] = 'errors'
            seed_progress['tasks'][2]['status'] = 'running'

            # Pass the correct path to the error codes SQL file
            sql_file = os.path.join(cwd, 'database/seeds/nhso_error_codes.sql')
            result = subprocess.run(
                ['python', 'database/seeds/nhso_error_codes_importer.py', sql_file],
                capture_output=True,
                text=True,
                cwd=cwd
            )
            if result.returncode != 0:
                raise Exception(f"Error codes seed failed: {result.stderr}")

            seed_progress['tasks'][2]['status'] = 'completed'
            seed_progress['tasks'][2]['records'] = 200  # Approximate
            seed_progress['completed'] = 3

            seed_progress['current_task'] = None
            seed_progress['finished_at'] = datetime.now(TZ_BANGKOK).isoformat()

        except Exception as e:
            seed_progress['error'] = str(e)
            # Mark current task as failed
            for task in seed_progress['tasks']:
                if task['status'] == 'running':
                    task['status'] = 'failed'
                    task['error'] = str(e)

        finally:
            seed_progress['running'] = False

    thread = threading.Thread(target=run_seeds, daemon=True)
    thread.start()

    return jsonify({
        'success': True,
        'message': 'Seed initialization started',
        'progress': seed_progress
    })


@app.route('/api/system/seed-progress')
def get_seed_progress():
    """Get current seed initialization progress"""
    return jsonify({
        'success': True,
        'progress': seed_progress
    })


@app.route('/api/system/sync-status', methods=['POST'])
def sync_system_status():
    """
    Comprehensive system status check and sync:
    1. Check if download processes are actually running
    2. Reset stuck progress files if processes are dead
    3. Sync files in folders with history records
    4. Return detailed status report
    """
    import json
    import os
    import psutil
    from pathlib import Path

    report = {
        'success': True,
        'timestamp': datetime.now(TZ_BANGKOK).isoformat(),
        'actions': [],
        'summary': {
            'processes_checked': 0,
            'processes_reset': 0,
            'files_synced': 0,
            'files_added': 0,
            'files_removed': 0
        }
    }

    # PID files to check
    pid_files = {
        'downloader': Path('/tmp/eclaim_downloader.pid'),
        'import': Path('/tmp/eclaim_import.pid'),
        'parallel': Path('/tmp/eclaim_parallel_download.pid'),
        'stm': Path('/tmp/eclaim_stm_downloader.pid'),
        'smt': Path('/tmp/eclaim_smt_fetch.pid')
    }

    # === STEP 1: Check PID files and reset if processes are dead ===
    for name, pid_file in pid_files.items():
        report['summary']['processes_checked'] += 1
        if pid_file.exists():
            try:
                pid = int(pid_file.read_text().strip())
                try:
                    process = psutil.Process(pid)
                    if process.is_running() and process.status() != psutil.STATUS_ZOMBIE:
                        report['actions'].append({
                            'type': 'process_check',
                            'name': name,
                            'pid': pid,
                            'status': 'running',
                            'action': 'none'
                        })
                    else:
                        # Process is zombie, cleanup
                        pid_file.unlink()
                        report['actions'].append({
                            'type': 'process_check',
                            'name': name,
                            'pid': pid,
                            'status': 'zombie',
                            'action': 'pid_file_removed'
                        })
                        report['summary']['processes_reset'] += 1
                except psutil.NoSuchProcess:
                    # Process doesn't exist, cleanup
                    pid_file.unlink()
                    report['actions'].append({
                        'type': 'process_check',
                        'name': name,
                        'pid': pid,
                        'status': 'dead',
                        'action': 'pid_file_removed'
                    })
                    report['summary']['processes_reset'] += 1
            except (ValueError, IOError) as e:
                pid_file.unlink()
                report['actions'].append({
                    'type': 'process_check',
                    'name': name,
                    'status': 'invalid_pid_file',
                    'action': 'pid_file_removed',
                    'error': str(e)
                })
                report['summary']['processes_reset'] += 1

    # === STEP 2: Sync REP files with database ===
    # (Progress tracking now uses database - no JSON files to reset)
    # REP files are stored in downloads/rep/ subfolder
    rep_dir = Path('downloads/rep')

    if rep_dir.exists():
        try:
            conn = get_pooled_connection()
            cursor = conn.cursor()

            # Get existing filenames in database
            cursor.execute("SELECT filename FROM download_history WHERE download_type = 'rep'")
            db_filenames = {row[0] for row in cursor.fetchall()}

            # Get actual files on disk
            disk_files = {f.name for f in rep_dir.glob('*.xls') if f.is_file() and f.stat().st_size > 0}

            # Files on disk but not in database -> add to database
            missing_from_db = disk_files - db_filenames
            db_type = os.environ.get('DB_TYPE', 'postgresql').lower()
            for filename in missing_from_db:
                file_path = rep_dir / filename
                if db_type == 'mysql':
                    cursor.execute("""
                        INSERT INTO download_history
                        (download_type, filename, file_size, file_path, file_exists)
                        VALUES ('rep', %s, %s, %s, TRUE)
                        ON DUPLICATE KEY UPDATE file_size = VALUES(file_size)
                    """, (filename, file_path.stat().st_size, str(file_path)))
                else:
                    cursor.execute("""
                        INSERT INTO download_history
                        (download_type, filename, file_size, file_path, file_exists)
                        VALUES ('rep', %s, %s, %s, TRUE)
                        ON CONFLICT (download_type, filename) DO NOTHING
                    """, (filename, file_path.stat().st_size, str(file_path)))
                report['summary']['files_added'] += 1

            # Files in database but not on disk -> mark as file_exists=FALSE
            missing_from_disk = db_filenames - disk_files
            if missing_from_disk:
                for filename in missing_from_disk:
                    cursor.execute("""
                        UPDATE download_history
                        SET file_exists = FALSE, updated_at = CURRENT_TIMESTAMP
                        WHERE download_type = 'rep' AND filename = %s
                    """, (filename,))
                report['summary']['files_removed'] += len(missing_from_disk)

            conn.commit()
            cursor.close()
            return_connection(conn)

            if missing_from_db or missing_from_disk:
                report['actions'].append({
                    'type': 'history_sync',
                    'target': 'REP files (database)',
                    'added': len(missing_from_db),
                    'removed': len(missing_from_disk)
                })

            report['summary']['files_synced'] = len(disk_files)

        except Exception as e:
            report['actions'].append({
                'type': 'history_sync',
                'target': 'REP files',
                'status': 'error',
                'error': str(e)
            })

    # === STEP 4: Sync Statement files with database ===
    # Statement files are stored in downloads/stm/ subfolder
    stm_dir = Path('downloads/stm')

    if stm_dir.exists():
        try:
            conn = get_pooled_connection()
            cursor = conn.cursor()

            # Get existing filenames in database
            cursor.execute("SELECT filename FROM download_history WHERE download_type = 'stm'")
            stm_db_filenames = {row[0] for row in cursor.fetchall()}

            # Get actual files on disk
            stm_disk_files = {f.name for f in stm_dir.glob('*.xls') if f.is_file() and f.stat().st_size > 0}

            # Files on disk but not in database -> add to database
            stm_missing_from_db = stm_disk_files - stm_db_filenames
            db_type = os.environ.get('DB_TYPE', 'postgresql').lower()
            for filename in stm_missing_from_db:
                file_path = stm_dir / filename
                if db_type == 'mysql':
                    cursor.execute("""
                        INSERT INTO download_history
                        (download_type, filename, file_size, file_path, file_exists)
                        VALUES ('stm', %s, %s, %s, TRUE)
                        ON DUPLICATE KEY UPDATE file_size = VALUES(file_size)
                    """, (filename, file_path.stat().st_size, str(file_path)))
                else:
                    cursor.execute("""
                        INSERT INTO download_history
                        (download_type, filename, file_size, file_path, file_exists)
                        VALUES ('stm', %s, %s, %s, TRUE)
                        ON CONFLICT (download_type, filename) DO NOTHING
                    """, (filename, file_path.stat().st_size, str(file_path)))
                report['summary']['files_added'] += 1

            # Files in database but not on disk -> mark as file_exists=FALSE
            stm_missing_from_disk = stm_db_filenames - stm_disk_files
            if stm_missing_from_disk:
                for filename in stm_missing_from_disk:
                    cursor.execute("""
                        UPDATE download_history
                        SET file_exists = FALSE, updated_at = CURRENT_TIMESTAMP
                        WHERE download_type = 'stm' AND filename = %s
                    """, (filename,))
                report['summary']['files_removed'] += len(stm_missing_from_disk)

            conn.commit()
            cursor.close()
            return_connection(conn)

            if stm_missing_from_db or stm_missing_from_disk:
                report['actions'].append({
                    'type': 'history_sync',
                    'target': 'Statement files (database)',
                    'added': len(stm_missing_from_db),
                    'removed': len(stm_missing_from_disk)
                })

        except Exception as e:
            report['actions'].append({
                'type': 'history_sync',
                'target': 'Statement files',
                'status': 'error',
                'error': str(e)
            })

    # Generate summary message
    actions_taken = report['summary']['processes_reset'] + report['summary']['files_added'] + report['summary']['files_removed']
    if actions_taken > 0:
        report['message'] = f"Sync completed: Reset {report['summary']['processes_reset']} processes, " \
                           f"Added {report['summary']['files_added']} files, " \
                           f"Removed {report['summary']['files_removed']} orphaned records"
    else:
        report['message'] = "All systems are in sync. No actions needed."

    return jsonify(report), 200


def _check_pid_alive(pid_file_path):
    """Helper to check if PID in file is alive"""
    import psutil
    try:
        pid_file = Path(pid_file_path)
        if not pid_file.exists():
            return False
        pid = int(pid_file.read_text().strip())
        process = psutil.Process(pid)
        return process.is_running() and process.status() != psutil.STATUS_ZOMBIE
    except (ValueError, IOError, psutil.NoSuchProcess):
        return False


# ===== Schedule API routes moved to routes/settings.py =====


# ============================================
# SMT Budget Routes
# ============================================

@app.route('/smt-budget')
@login_required
def smt_budget():
    """SMT Budget Report page"""
    smt_settings = settings_manager.get_smt_settings()
    smt_jobs = [j for j in download_scheduler.get_all_jobs() if j['id'].startswith('smt_')]

    # Get latest SMT data from database
    smt_summary = get_smt_summary()

    return render_template(
        'smt_budget.html',
        smt_settings=smt_settings,
        smt_jobs=smt_jobs,
        smt_summary=smt_summary
    )


def get_smt_summary():
    """Get SMT budget summary from database"""
    conn = get_db_connection()
    if not conn:
        return None

    try:
        cursor = conn.cursor()

        # Check if table exists
        cursor.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables
                WHERE table_name = 'smt_budget_transfers'
            );
        """)
        table_exists = cursor.fetchone()[0]

        if not table_exists:
            cursor.close()
            conn.close()
            return None

        # Get summary
        cursor.execute("""
            SELECT
                COUNT(*) as total_records,
                COALESCE(SUM(amount), 0) as total_amount,
                MIN(posting_date) as earliest_date,
                MAX(posting_date) as latest_date,
                MAX(created_at) as last_updated
            FROM smt_budget_transfers
        """)
        row = cursor.fetchone()

        # Get summary by fund group
        cursor.execute("""
            SELECT
                fund_group_desc,
                COUNT(*) as record_count,
                SUM(amount) as total_amount
            FROM smt_budget_transfers
            GROUP BY fund_group_desc
            ORDER BY SUM(amount) DESC
            LIMIT 10
        """)
        fund_groups = cursor.fetchall()

        cursor.close()
        conn.close()

        return {
            'total_records': row[0] or 0,
            'total_amount': float(row[1] or 0),
            'earliest_date': row[2],
            'latest_date': row[3],
            'last_updated': row[4].isoformat() if row[4] else None,
            'fund_groups': [
                {'name': r[0], 'count': r[1], 'amount': float(r[2] or 0)}
                for r in fund_groups
            ]
        }

    except Exception as e:
        app.logger.error(f"Error getting SMT summary: {e}")
        if conn:
            conn.close()
        return None


@app.route('/api/smt/fetch', methods=['POST'])
def smt_fetch():
    """Trigger SMT budget fetch"""
    # Check license feature access
    if not settings_manager.check_feature_access('smt_budget'):
        return jsonify({
            'success': False,
            'error': 'SMT Budget feature requires Basic tier or higher license',
            'upgrade_required': True,
            'current_tier': settings_manager.get_license_info().get('tier', 'trial')
        }), 403

    job_id = None
    try:
        data = request.get_json() or {}
        vendor_id = data.get('vendor_id')
        start_date = data.get('start_date')  # dd/mm/yyyy BE format
        end_date = data.get('end_date')  # dd/mm/yyyy BE format
        budget_year = data.get('budget_year')  # Buddhist Era year
        save_db = data.get('save_db', True)
        export_format = data.get('export_format')  # 'json' or 'csv' or None

        if not vendor_id:
            # Try to get from global hospital_code setting (falls back to legacy smt_vendor_id)
            vendor_id = settings_manager.get_hospital_code()

        if not vendor_id:
            return jsonify({'success': False, 'error': 'Vendor ID is required. Please configure Hospital Code in Settings.'}), 400

        # Start job tracking
        try:
            job_id = job_history_manager.start_job(
                job_type='download',
                job_subtype='smt_fetch',
                parameters={
                    'vendor_id': vendor_id,
                    'budget_year': budget_year,
                    'start_date': start_date,
                    'end_date': end_date,
                    'save_db': save_db
                },
                triggered_by='manual'
            )
        except Exception as e:
            app.logger.warning(f"Could not start job tracking: {e}")

        # Import and run fetcher
        from smt_budget_fetcher import SMTBudgetFetcher

        date_info = ""
        if start_date and end_date:
            date_info = f" ({start_date} - {end_date})"
        elif budget_year:
            date_info = f" (FY {budget_year})"

        log_streamer.write_log(
            f"Starting SMT fetch for vendor {vendor_id}{date_info}...",
            'info',
            'smt'
        )

        fetcher = SMTBudgetFetcher(vendor_id=vendor_id)
        result = fetcher.fetch_budget_summary(
            budget_year=int(budget_year) if budget_year else None,
            start_date=start_date,
            end_date=end_date
        )
        records = result.get('datas', [])

        if not records:
            log_streamer.write_log(
                f"No records found for vendor {vendor_id}",
                'warning',
                'smt'
            )
            return jsonify({
                'success': True,
                'message': 'No records found',
                'records': 0
            })

        # Calculate summary
        summary = fetcher.calculate_summary(records)

        # Debug: show vendor_no format from first record
        if records:
            first_record = records[0]
            print(f"[DEBUG] SMT fetch - vendor_no format from API: vndrNo='{first_record.get('vndrNo')}'")

        # Save to database if requested
        saved_count = 0
        if save_db:
            saved_count = fetcher.save_to_database(records)
            log_streamer.write_log(
                f"Saved {saved_count} records to database",
                'success',
                'smt'
            )

        # Export if requested
        export_path = None
        if export_format == 'json':
            export_path = fetcher.export_to_json(records)
        elif export_format == 'csv':
            export_path = fetcher.export_to_csv(records)

        log_streamer.write_log(
            f"✓ SMT fetch completed: {len(records)} records, {summary['total_amount']:,.2f} Baht",
            'success',
            'smt'
        )

        # Complete job tracking
        if job_id:
            try:
                job_history_manager.complete_job(
                    job_id=job_id,
                    status='completed',
                    results={
                        'records': len(records),
                        'saved': saved_count,
                        'total_amount': summary['total_amount'],
                        'vendor_id': vendor_id,
                        'budget_year': budget_year
                    }
                )
            except Exception as e:
                app.logger.warning(f"Could not complete job tracking: {e}")

        return jsonify({
            'success': True,
            'records': len(records),
            'saved': saved_count,
            'total_amount': summary['total_amount'],
            'export_path': export_path
        })

    except Exception as e:
        # Mark job as failed
        if job_id:
            try:
                job_history_manager.complete_job(
                    job_id=job_id,
                    status='failed',
                    error_message=str(e)
                )
            except Exception:
                pass
        log_streamer.write_log(
            f"✗ SMT fetch failed: {str(e)}",
            'error',
            'smt'
        )
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== SMT settings route moved to routes/settings.py =====


@app.route('/api/smt/fiscal-years')
def api_smt_fiscal_years():
    """Get available fiscal years in database"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Get distinct fiscal years from posting_date
        cursor.execute("""
            SELECT DISTINCT fiscal_year
            FROM smt_budget
            WHERE fiscal_year IS NOT NULL
            ORDER BY fiscal_year DESC
        """)

        rows = cursor.fetchall()
        fiscal_years = [r[0] for r in rows if r[0] and r[0] > 2500]

        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'fiscal_years': fiscal_years
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e), 'fiscal_years': []}), 200


@app.route('/api/smt/data')
def api_smt_data():
    """Get SMT budget data from database with optional date filtering"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        fund_group = request.args.get('fund_group')
        start_date = request.args.get('start_date')  # Format: dd/mm/yyyy BE
        end_date = request.args.get('end_date')      # Format: dd/mm/yyyy BE

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Helper function to convert dd/mm/yyyy to sortable yyyymmdd format
        def to_sortable(date_str):
            if not date_str or len(date_str) < 10:
                return None
            parts = date_str.split('/')
            if len(parts) == 3:
                return f"{parts[2]}{parts[1]}{parts[0]}"
            return None

        # Build query with parameterized values
        conditions = []
        params = []

        if fund_group:
            conditions.append("fund_group_desc = %s")
            params.append(fund_group)

        # Convert dates to sortable format for string comparison
        # posting_date can be stored as:
        # - "25671227" (yyyymmdd) - 8 digits, already sortable
        # - "27/12/2567" (dd/mm/yyyy) - 10 chars, need transformation
        # We use CASE to handle both formats
        sortable_posting_date_expr = """
            CASE
                WHEN LENGTH(posting_date) = 8 THEN posting_date
                WHEN LENGTH(posting_date) = 10 THEN CONCAT(RIGHT(posting_date, 4), SUBSTRING(posting_date, 4, 2), SUBSTRING(posting_date, 1, 2))
                ELSE posting_date
            END
        """

        if start_date:
            sortable_start = to_sortable(start_date)
            if sortable_start:
                conditions.append(f"({sortable_posting_date_expr}) >= %s")
                params.append(sortable_start)

        if end_date:
            sortable_end = to_sortable(end_date)
            if sortable_end:
                conditions.append(f"({sortable_posting_date_expr}) <= %s")
                params.append(sortable_end)

        where_clause = ""
        if conditions:
            where_clause = "WHERE " + " AND ".join(conditions)

        # Get total count
        count_query = "SELECT COUNT(*) FROM smt_budget_transfers " + where_clause
        cursor.execute(count_query, params)
        total = cursor.fetchone()[0]

        # Get paginated data - order by sortable date format
        offset = (page - 1) * per_page
        sortable_order_expr = """
            CASE
                WHEN LENGTH(posting_date) = 8 THEN posting_date
                WHEN LENGTH(posting_date) = 10 THEN CONCAT(RIGHT(posting_date, 4), SUBSTRING(posting_date, 4, 2), SUBSTRING(posting_date, 1, 2))
                ELSE posting_date
            END
        """
        select_query = """
            SELECT id, run_date, posting_date, ref_doc_no, vendor_no,
                   fund_name, fund_group_desc, amount, total_amount,
                   bank_name, payment_status, created_at
            FROM smt_budget_transfers
            """ + where_clause + f"""
            ORDER BY ({sortable_order_expr}) DESC, id DESC
            LIMIT %s OFFSET %s
        """
        cursor.execute(select_query, params + [per_page, offset])

        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        data = [
            {
                'id': r[0],
                'run_date': str(r[1]) if r[1] else None,
                'posting_date': r[2],
                'ref_doc_no': r[3],
                'vendor_no': r[4],
                'fund_name': r[5],
                'fund_group_desc': r[6],
                'amount': float(r[7] or 0),
                'total_amount': float(r[8] or 0),
                'bank_name': r[9],
                'payment_status': r[10],
                'created_at': r[11].isoformat() if r[11] else None
            }
            for r in rows
        ]

        return jsonify({
            'success': True,
            'data': data,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': (total + per_page - 1) // per_page
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/smt/stats')
def api_smt_stats():
    """Get SMT budget statistics (record count and last sync time)"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': True, 'record_count': 0, 'last_sync': None})

        cursor = conn.cursor()

        # Get total record count
        cursor.execute("SELECT COUNT(*) FROM smt_budget_transfers")
        record_count = cursor.fetchone()[0]

        # Get last sync time (most recent created_at)
        cursor.execute("SELECT MAX(created_at) FROM smt_budget_transfers")
        last_sync_result = cursor.fetchone()[0]
        last_sync = str(last_sync_result) if last_sync_result else None

        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'record_count': record_count,
            'last_sync': last_sync
        })

    except Exception as e:
        return jsonify({'success': True, 'record_count': 0, 'last_sync': None})


@app.route('/api/smt/clear', methods=['POST'])
def api_smt_clear():
    """Clear all SMT budget data from database"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Delete all SMT budget transfers
        cursor.execute("DELETE FROM smt_budget_transfers")
        deleted_count = cursor.rowcount

        conn.commit()
        cursor.close()
        conn.close()

        log_streamer.write_log(f"Cleared {deleted_count} SMT budget records", 'info', 'smt')

        return jsonify({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'Deleted {deleted_count} records'
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/api/smt/download', methods=['POST'])
def api_smt_download():
    """Download SMT budget data and export to CSV, with optional auto-import to database"""
    job_id = None
    try:
        data = request.get_json() or {}
        vendor_id = data.get('vendor_id')
        start_date = data.get('start_date')  # dd/mm/yyyy BE format
        end_date = data.get('end_date')      # dd/mm/yyyy BE format
        budget_source = data.get('budget_source', '')  # UC, OF, SS, LG, or empty for all
        budget_type = data.get('budget_type', '')      # OP, IP, PP, or empty for all
        auto_import = data.get('auto_import', False)   # Auto-import to database after download

        # Vendor ID is optional - empty means all in region
        if not vendor_id:
            # Try to get from global hospital_code setting (falls back to legacy smt_vendor_id)
            default_vendor = settings_manager.get_hospital_code()
            # Only use default if vendor_id is not explicitly provided as empty string
            if vendor_id is None and default_vendor:
                vendor_id = default_vendor

        # Start job tracking
        try:
            job_id = job_history_manager.start_job(
                job_type='download',
                job_subtype='smt',
                parameters={
                    'vendor_id': vendor_id,
                    'start_date': start_date,
                    'end_date': end_date,
                    'budget_source': budget_source,
                    'auto_import': auto_import
                },
                triggered_by='manual'
            )
        except Exception as e:
            app.logger.warning(f"Could not start job tracking: {e}")

        # Import and run fetcher
        from smt_budget_fetcher import SMTBudgetFetcher

        vendor_display = vendor_id if vendor_id else 'ทั้งหมด (All)'
        log_streamer.write_log(
            f"Starting SMT download for vendor {vendor_display} ({start_date} - {end_date})...",
            'info',
            'smt'
        )

        fetcher = SMTBudgetFetcher(vendor_id=vendor_id if vendor_id else None)
        result = fetcher.fetch_budget_summary(
            start_date=start_date,
            end_date=end_date,
            budget_source=budget_source
        )
        records = result.get('datas', [])

        if not records:
            return jsonify({
                'success': True,
                'message': 'No records found',
                'records': 0
            })

        # Calculate summary
        summary = fetcher.calculate_summary(records)

        # Auto-import to database if requested
        imported_count = 0
        new_records = 0
        export_path = None

        if auto_import:
            # Count records before import to detect new records
            count_before = 0
            try:
                conn = get_db_connection()
                if conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT COUNT(*) FROM smt_budget_transfers")
                    count_before = cursor.fetchone()[0]
                    cursor.close()
                    conn.close()
            except Exception:
                pass

            log_streamer.write_log(
                f"Auto-importing {len(records)} records to database...",
                'info',
                'smt'
            )
            imported_count = fetcher.save_to_database(records)

            # Count after import to detect new records
            count_after = 0
            try:
                conn = get_db_connection()
                if conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT COUNT(*) FROM smt_budget_transfers")
                    count_after = cursor.fetchone()[0]
                    cursor.close()
                    conn.close()
            except Exception:
                pass

            new_records = count_after - count_before

            # Only create CSV if there are new records
            if new_records > 0:
                export_path = fetcher.export_to_csv(records)
                log_streamer.write_log(
                    f"✓ {new_records} new records, exported to {export_path}",
                    'success',
                    'smt'
                )
            else:
                log_streamer.write_log(
                    f"✓ No new records (all {len(records)} already exist), skipping CSV export",
                    'info',
                    'smt'
                )

            message = f'Fetched {len(records)} records: {new_records} new, {len(records) - new_records} existing'
        else:
            # Without auto-import, always create CSV
            export_path = fetcher.export_to_csv(records)
            message = f'Downloaded {len(records)} records'

        log_streamer.write_log(
            f"✓ SMT download completed",
            'success',
            'smt'
        )

        # Complete job tracking
        if job_id:
            try:
                job_history_manager.complete_job(
                    job_id=job_id,
                    status='completed',
                    results={
                        'records': len(records),
                        'new_records': new_records,
                        'imported': imported_count,
                        'total_amount': summary['total_amount'],
                        'export_path': export_path
                    }
                )
            except Exception as e:
                app.logger.warning(f"Could not complete job tracking: {e}")

        return jsonify({
            'success': True,
            'message': message,
            'records': len(records),
            'new_records': new_records,
            'imported': imported_count,
            'total_amount': summary['total_amount'],
            'export_path': export_path
        })

    except Exception as e:
        # Mark job as failed
        if job_id:
            try:
                job_history_manager.complete_job(
                    job_id=job_id,
                    status='failed',
                    error_message=str(e)
                )
            except Exception:
                pass
        log_streamer.write_log(f"✗ SMT download failed: {str(e)}", 'error', 'smt')
        return jsonify({'success': False, 'error': str(e)}), 500
def init_smt_scheduler():
    """Initialize SMT scheduler with saved settings"""
    try:
        smt_settings = settings_manager.get_smt_settings()

        # Clear existing SMT jobs
        download_scheduler.remove_smt_jobs()

        # Get vendor ID from settings or global hospital_code
        vendor_id = smt_settings.get('smt_vendor_id') or settings_manager.get_hospital_code()
        if smt_settings['smt_schedule_enabled'] and vendor_id:
            auto_save_db = smt_settings['smt_auto_save_db']

            for time_config in smt_settings['smt_schedule_times']:
                hour = time_config.get('hour', 0)
                minute = time_config.get('minute', 0)
                download_scheduler.add_smt_scheduled_fetch(hour, minute, vendor_id, auto_save_db)

            log_streamer.write_log(
                f"✓ SMT scheduler initialized with {len(smt_settings['smt_schedule_times'])} jobs",
                'success',
                'system'
            )
    except Exception as e:
        app.logger.error(f"Error initializing SMT scheduler: {e}")


# ============================================
# Analytics Dashboard Routes
# ============================================

def _validate_date_param(date_str):
    """
    Validate date parameter format (YYYY-MM-DD).
    Returns validated date string or None if invalid.
    """
    if not date_str:
        return None
    import re
    # Strict format: YYYY-MM-DD with valid ranges
    if not re.match(r'^\d{4}-(?:0[1-9]|1[0-2])-(?:0[1-9]|[12]\d|3[01])$', date_str):
        return None
    return date_str


def get_analytics_date_filter():
    """
    Get date filter parameters from request args.
    Returns tuple: (where_clause, params, filter_info)

    Supports:
    - fiscal_year: Buddhist Era fiscal year (e.g., 2569 = Oct 2025 - Sep 2026)
    - start_date: Start date in YYYY-MM-DD format
    - end_date: End date in YYYY-MM-DD format

    All parameters are validated and passed via parameterized queries.
    """
    fiscal_year = request.args.get('fiscal_year', type=int)
    start_date = _validate_date_param(request.args.get('start_date'))
    end_date = _validate_date_param(request.args.get('end_date'))

    where_clauses = []
    params = []
    filter_info = {}

    if fiscal_year:
        # Use standardized fiscal year calculation from utils/fiscal_year.py
        # FY 2569 BE = Oct 2025 CE - Sep 2026 CE (1 Oct 2025 - 30 Sep 2026)
        where_clause, where_params = get_fiscal_year_sql_filter_gregorian(fiscal_year, 'dateadm')
        where_clauses.append(where_clause)
        params.extend(where_params)
        filter_info['fiscal_year'] = fiscal_year
        filter_info['date_range'] = f"{where_params[0]} to {where_params[1]}"
    elif start_date or end_date:
        if start_date:
            where_clauses.append("dateadm >= %s")
            params.append(start_date)
            filter_info['start_date'] = start_date
        if end_date:
            where_clauses.append("dateadm <= %s")
            params.append(end_date)
            filter_info['end_date'] = end_date

    where_clause = " AND ".join(where_clauses) if where_clauses else ""
    return where_clause, params, filter_info


def get_available_fiscal_years(cursor):
    """Get list of available fiscal years from data"""
    year_expr = sql_extract_year('dateadm')
    month_expr = sql_extract_month('dateadm')
    cursor.execute(f"""
        SELECT DISTINCT
            CASE
                WHEN {month_expr} >= 10 THEN {year_expr} + 544
                ELSE {year_expr} + 543
            END as fiscal_year
        FROM claim_rep_opip_nhso_item
        WHERE dateadm IS NOT NULL
        ORDER BY fiscal_year DESC
    """)
    return [row[0] for row in cursor.fetchall()]


@app.route('/analytics')
def analytics():
    """Analytics Dashboard - Comprehensive claim analysis"""
    return render_template('analytics.html')


@app.route('/claims')
def claims():
    """Claims Viewer - Detailed claim list with filters and drill-down"""
    return render_template('claims.html')


@app.route('/denial')
def denial():
    """Denial Root Cause Analysis - Charts and recommendations"""
    return render_template('denial.html')


@app.route('/strategic')
def strategic():
    """Strategic Analytics - Forecasting and YoY comparison"""
    return render_template('strategic.html')


@app.route('/benchmark')
def benchmark_page():
    """Benchmark Comparison page"""
    return render_template('benchmark.html')


@app.route('/api/benchmark/hospitals')
def api_benchmark_hospitals():
    """Get list of hospitals from SMT data for comparison"""
    try:
        fiscal_year = request.args.get('fiscal_year')

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Build date filter based on fiscal year
        # Thai fiscal year: Oct (year-1) to Sep (year)
        where_clause = ""
        params = []

        if fiscal_year:
            fiscal_year_int = int(fiscal_year)
            # Use fiscal year utility for Gregorian date filtering (run_date is Gregorian)
            sql_filter, filter_params = get_fiscal_year_sql_filter_gregorian(fiscal_year_int, 's.run_date')
            where_clause = f"WHERE {sql_filter}"
            params = filter_params

        # Get summary by vendor from smt_budget_transfers with hospital name lookup
        ltrim_expr = "TRIM(LEADING '0' FROM s.vendor_no)" if DB_TYPE == 'mysql' else "LTRIM(s.vendor_no, '0')"

        # Fix collation mismatch for MySQL
        if DB_TYPE == 'mysql':
            join_condition = f"""
                h.hcode5 COLLATE utf8mb4_unicode_ci = {ltrim_expr} COLLATE utf8mb4_unicode_ci
                OR h.hcode5 COLLATE utf8mb4_unicode_ci = s.vendor_no COLLATE utf8mb4_unicode_ci
            """
        else:
            join_condition = f"""
                h.hcode5 = {ltrim_expr}
                OR h.hcode5 = s.vendor_no
            """

        query = f"""
            SELECT
                s.vendor_no,
                COUNT(*) as records,
                COALESCE(SUM(s.total_amount), 0) as total_amount,
                COALESCE(SUM(s.wait_amount), 0) as wait_amount,
                COALESCE(SUM(s.debt_amount), 0) as debt_amount,
                COALESCE(SUM(s.bond_amount), 0) as bond_amount,
                MIN(s.run_date) as first_date,
                MAX(s.run_date) as last_date,
                h.name as hospital_name
            FROM smt_budget_transfers s
            LEFT JOIN health_offices h ON (
                {join_condition}
            )
            {where_clause}
            GROUP BY s.vendor_no, h.name
            ORDER BY total_amount DESC
        """
        cursor.execute(query, params)

        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        hospitals = []
        for row in rows:
            vendor_no = row[0]
            hospital_name = row[8] if row[8] else None
            hospitals.append({
                'vendor_no': vendor_no,
                'records': row[1],
                'total_amount': float(row[2]) if row[2] else 0,
                'wait_amount': float(row[3]) if row[3] else 0,
                'debt_amount': float(row[4]) if row[4] else 0,
                'bond_amount': float(row[5]) if row[5] else 0,
                'first_date': row[6].strftime('%Y-%m-%d') if row[6] else None,
                'last_date': row[7].strftime('%Y-%m-%d') if row[7] else None,
                'hospital_name': hospital_name
            })

        return jsonify({
            'success': True,
            'hospitals': hospitals,
            'count': len(hospitals)
        })

    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/benchmark/timeseries')
def api_benchmark_timeseries():
    """Get time-series data for hospital comparison charts"""
    try:
        fiscal_year = request.args.get('fiscal_year')
        start_month = request.args.get('start_month')
        end_month = request.args.get('end_month')

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Build date filter based on fiscal year and month range
        # Thai fiscal year: Oct (year-1) to Sep (year)
        where_clause = ""
        params = []

        if fiscal_year:
            fiscal_year_int = int(fiscal_year)
            # Use fiscal year utility for Gregorian date filtering (run_date is Gregorian)
            sql_filter, filter_params = get_fiscal_year_sql_filter_gregorian(fiscal_year_int, 'run_date')
            where_clause = f"WHERE {sql_filter}"
            params = filter_params

            # If specific month range is specified
            if start_month and end_month:
                start_m = int(start_month)
                end_m = int(end_month)
                # Adjust dates based on month in fiscal year
                # Convert Buddhist Era to Gregorian for month range calculation
                gregorian_year = fiscal_year_int - 543
                if start_m >= 10:
                    start_date = f"{gregorian_year - 1}-{start_m:02d}-01"
                else:
                    start_date = f"{gregorian_year}-{start_m:02d}-01"
                if end_m >= 10:
                    end_date = f"{gregorian_year - 1}-{end_m:02d}-28"
                else:
                    end_date = f"{gregorian_year}-{end_m:02d}-28"
                # Override with month-specific range
                where_clause = "WHERE run_date >= %s AND run_date <= %s"
                params = [start_date, end_date]

        # Get monthly summary by vendor
        year_expr = sql_extract_year('s.run_date')
        month_expr = sql_extract_month('s.run_date')
        # LTRIM syntax differs between databases
        ltrim_expr = "TRIM(LEADING '0' FROM s.vendor_no)" if DB_TYPE == 'mysql' else "LTRIM(s.vendor_no, '0')"
        query = f"""
            SELECT
                s.vendor_no,
                h.name as hospital_name,
                {year_expr} as year,
                {month_expr} as month,
                COUNT(*) as records,
                COALESCE(SUM(s.total_amount), 0) as total_amount,
                COALESCE(SUM(s.wait_amount), 0) as wait_amount,
                COALESCE(SUM(s.debt_amount), 0) as debt_amount
            FROM smt_budget_transfers s
            LEFT JOIN health_offices h ON (
                h.hcode5 = {ltrim_expr}
                OR h.hcode5 = s.vendor_no
            )
            {where_clause}
            GROUP BY s.vendor_no, h.name, {year_expr}, {month_expr}
            ORDER BY s.vendor_no, year, month
        """

        cursor.execute(query, params)
        rows = cursor.fetchall()

        # Also get available fiscal years
        fy_year_expr = sql_extract_year('run_date')
        fy_month_expr = sql_extract_month('run_date')
        cursor.execute(f"""
            SELECT DISTINCT
                CASE
                    WHEN {fy_month_expr} >= 10 THEN {fy_year_expr} + 544
                    ELSE {fy_year_expr} + 543
                END as fiscal_year
            FROM smt_budget_transfers
            ORDER BY fiscal_year DESC
        """)
        fiscal_years = [int(r[0]) for r in cursor.fetchall()]

        cursor.close()
        conn.close()

        # Organize data by vendor
        vendors = {}
        months = set()

        for row in rows:
            vendor_no = row[0]
            hospital_name = row[1]
            year = int(row[2])
            month = int(row[3])
            month_key = f"{year}-{month:02d}"
            months.add(month_key)

            if vendor_no not in vendors:
                vendors[vendor_no] = {
                    'vendor_no': vendor_no,
                    'hospital_name': hospital_name or f'รพ. {vendor_no.lstrip("0")}',
                    'data': {}
                }

            vendors[vendor_no]['data'][month_key] = {
                'records': row[4],
                'total_amount': float(row[5]) if row[5] else 0,
                'wait_amount': float(row[6]) if row[6] else 0,
                'debt_amount': float(row[7]) if row[7] else 0
            }

        return jsonify({
            'success': True,
            'vendors': list(vendors.values()),
            'months': sorted(list(months)),
            'fiscal_years': fiscal_years
        })

    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/benchmark/hospital-years')
def api_benchmark_hospital_years():
    """Get which fiscal years have data for a specific hospital"""
    try:
        vendor_id = request.args.get('vendor_id')

        if not vendor_id:
            return jsonify({'success': False, 'error': 'vendor_id is required'}), 400

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Normalize vendor_id
        vendor_id_10 = vendor_id.zfill(10)
        vendor_id_5 = vendor_id.lstrip('0') or vendor_id  # Keep original if all zeros

        print(f"[DEBUG] hospital-years: vendor_id={vendor_id}, vendor_id_10={vendor_id_10}, vendor_id_5={vendor_id_5}")

        # First check what vendor_no values exist in DB for debugging
        cursor.execute("""
            SELECT DISTINCT vendor_no FROM smt_budget_transfers
            WHERE vendor_no LIKE %s OR vendor_no LIKE %s
            LIMIT 5
        """, (f'%{vendor_id_5}%', f'%{vendor_id}%'))
        debug_vendors = cursor.fetchall()
        print(f"[DEBUG] Found vendor_nos in DB matching pattern: {debug_vendors}")

        # Get all fiscal years that have data for this hospital
        # Use flexible matching: cast to numeric and compare to handle different padding
        fy_year_expr = sql_extract_year('run_date')
        fy_month_expr = sql_extract_month('run_date')
        # MySQL and PostgreSQL have different REGEXP_REPLACE syntax
        if DB_TYPE == 'mysql':
            vendor_cast = "CAST(REGEXP_REPLACE(vendor_no, '[^0-9]', '') AS UNSIGNED)"
        else:
            vendor_cast = "CAST(NULLIF(REGEXP_REPLACE(vendor_no, '[^0-9]', '', 'g'), '') AS BIGINT)"
        cursor.execute(f"""
            SELECT
                CASE
                    WHEN {fy_month_expr} >= 10 THEN {fy_year_expr} + 544
                    ELSE {fy_year_expr} + 543
                END as fiscal_year,
                COUNT(*) as records,
                COALESCE(SUM(total_amount), 0) as total_amount
            FROM smt_budget_transfers
            WHERE {vendor_cast} = %s
            GROUP BY fiscal_year
            ORDER BY fiscal_year DESC
        """, (int(vendor_id_5),))
        rows = cursor.fetchall()
        print(f"[DEBUG] Found {len(rows)} years with data for vendor {vendor_id_5}")

        # Get all available years in the system (from any hospital)
        cursor.execute(f"""
            SELECT DISTINCT
                CASE
                    WHEN {fy_month_expr} >= 10 THEN {fy_year_expr} + 544
                    ELSE {fy_year_expr} + 543
                END as fiscal_year
            FROM smt_budget_transfers
            WHERE run_date IS NOT NULL
            ORDER BY fiscal_year DESC
        """)
        all_years = [int(r[0]) for r in cursor.fetchall()]

        cursor.close()
        conn.close()

        # Build response with status for each year
        hospital_years = {}
        for row in rows:
            year = int(row[0])
            hospital_years[year] = {
                'year': year,
                'has_data': True,
                'records': row[1],
                'total_amount': float(row[2])
            }

        # Add years that don't have data for this hospital
        # Include years from 2565 to current fiscal year
        today = datetime.now(TZ_BANGKOK)
        current_fiscal = today.year + 544 if today.month >= 10 else today.year + 543

        for year in range(2565, current_fiscal + 1):
            if year not in hospital_years:
                hospital_years[year] = {
                    'year': year,
                    'has_data': False,
                    'records': 0,
                    'total_amount': 0
                }

        # Sort by year descending
        years_list = sorted(hospital_years.values(), key=lambda x: x['year'], reverse=True)

        return jsonify({
            'success': True,
            'vendor_id': vendor_id,
            'years': years_list,
            'all_system_years': all_years
        })

    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/benchmark/my-hospital')
def api_benchmark_my_hospital():
    """Get detailed analytics for a single hospital"""
    try:
        vendor_id = request.args.get('vendor_id')
        fiscal_year = request.args.get('fiscal_year')

        if not vendor_id:
            return jsonify({'success': False, 'error': 'vendor_id is required'}), 400

        # Default to current fiscal year
        if not fiscal_year:
            today = datetime.now(TZ_BANGKOK)
            fiscal_year = today.year + 543 if today.month >= 10 else today.year + 542

        fiscal_year = int(fiscal_year)

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Convert fiscal year to date range using standardized calculation
        start_date, end_date = get_fiscal_year_range_gregorian(fiscal_year)

        # Previous year for YoY comparison
        prev_start_date, prev_end_date = get_fiscal_year_range_gregorian(fiscal_year - 1)

        # Normalize vendor_id (can be 5 or 10 digits)
        vendor_id_10 = vendor_id.zfill(10)
        vendor_id_5 = vendor_id.lstrip('0')

        # Get hospital info from health_offices (including bed count)
        cursor.execute("""
            SELECT name, hospital_level, province, health_region, hcode5, COALESCE(actual_beds, 0) as actual_beds
            FROM health_offices
            WHERE hcode5 = %s OR hcode5 = %s
            LIMIT 1
        """, (vendor_id_5, vendor_id))
        hospital_row = cursor.fetchone()

        actual_beds = int(hospital_row[5]) if hospital_row and hospital_row[5] else 0
        hospital_info = {
            'vendor_no': vendor_id_10,
            'name': hospital_row[0] if hospital_row else f'รพ. {vendor_id_5}',
            'level': hospital_row[1] if hospital_row else None,
            'province': hospital_row[2] if hospital_row else None,
            'health_region': hospital_row[3] if hospital_row else None,
            'actual_beds': actual_beds
        }

        # Get current year summary
        cursor.execute("""
            SELECT
                COUNT(*) as records,
                COALESCE(SUM(total_amount), 0) as total_amount,
                COALESCE(SUM(wait_amount), 0) as wait_amount,
                COALESCE(SUM(debt_amount), 0) as debt_amount,
                COALESCE(SUM(bond_amount), 0) as bond_amount
            FROM smt_budget_transfers
            WHERE (vendor_no = %s OR vendor_no = %s)
              AND run_date >= %s AND run_date <= %s
        """, (vendor_id_10, vendor_id_5, start_date, end_date))
        summary_row = cursor.fetchone()

        # Get previous year total for YoY
        cursor.execute("""
            SELECT COALESCE(SUM(total_amount), 0) as prev_total
            FROM smt_budget_transfers
            WHERE (vendor_no = %s OR vendor_no = %s)
              AND run_date >= %s AND run_date <= %s
        """, (vendor_id_10, vendor_id_5, prev_start_date, prev_end_date))
        prev_row = cursor.fetchone()
        prev_total = float(prev_row[0]) if prev_row and prev_row[0] else 0

        total_amount = float(summary_row[1]) if summary_row else 0
        wait_amount = float(summary_row[2]) if summary_row else 0
        debt_amount = float(summary_row[3]) if summary_row else 0

        growth_yoy = ((total_amount - prev_total) / prev_total * 100) if prev_total > 0 else 0

        # Calculate per-bed metrics
        revenue_per_bed = (total_amount / actual_beds) if actual_beds > 0 else 0
        wait_per_bed = (wait_amount / actual_beds) if actual_beds > 0 else 0
        debt_per_bed = (debt_amount / actual_beds) if actual_beds > 0 else 0

        # Get fund breakdown by category
        # Categories: OPD, IPD, CR (Central Reimburse), PP (Prevention/Promotion), OTHER
        # Note: %% escapes % for Python string formatting in psycopg2
        like_op = "LIKE" if DB_TYPE == 'mysql' else "ILIKE"
        fund_case = f"""CASE
                    WHEN fund_name {like_op} '%%ผู้ป่วยใน%%' OR fund_name = 'IP_CF' THEN 'IPD'
                    WHEN fund_name {like_op} '%%ผู้ป่วยนอก%%' OR fund_name = 'OP_CF' THEN 'OPD'
                    WHEN fund_name {like_op} '%%CENTRAL REIMBURSE%%' THEN 'CR'
                    WHEN fund_name {like_op} '%%สร้างเสริมสุขภาพ%%'
                         OR fund_name {like_op} '%%ป้องกันโรค%%'
                         OR fund_name {like_op} '%%ควบคุม%%ป้องกัน%%' THEN 'PP'
                    ELSE 'OTHER'
                END"""
        cursor.execute(f"""
            SELECT
                {fund_case} as fund_category,
                COALESCE(SUM(total_amount), 0) as total_amount,
                COALESCE(SUM(wait_amount), 0) as wait_amount,
                COALESCE(SUM(debt_amount), 0) as debt_amount,
                COUNT(*) as records
            FROM smt_budget_transfers
            WHERE (vendor_no = %s OR vendor_no = %s)
              AND run_date >= %s AND run_date <= %s
            GROUP BY {fund_case}
        """, (vendor_id_10, vendor_id_5, start_date, end_date))

        # Initialize fund categories
        fund_categories = {
            'OPD': {'total_amount': 0, 'wait_amount': 0, 'debt_amount': 0, 'records': 0, 'label': 'ผู้ป่วยนอก'},
            'IPD': {'total_amount': 0, 'wait_amount': 0, 'debt_amount': 0, 'records': 0, 'label': 'ผู้ป่วยใน'},
            'CR': {'total_amount': 0, 'wait_amount': 0, 'debt_amount': 0, 'records': 0, 'label': 'Central Reimburse'},
            'PP': {'total_amount': 0, 'wait_amount': 0, 'debt_amount': 0, 'records': 0, 'label': 'ส่งเสริมป้องกัน'},
            'OTHER': {'total_amount': 0, 'wait_amount': 0, 'debt_amount': 0, 'records': 0, 'label': 'อื่นๆ'}
        }

        ipd_total = 0
        ipd_wait = 0
        ipd_debt = 0
        opd_total = 0
        for row in cursor.fetchall():
            cat = row[0]
            if cat in fund_categories:
                fund_categories[cat]['total_amount'] = float(row[1]) if row[1] else 0
                fund_categories[cat]['wait_amount'] = float(row[2]) if row[2] else 0
                fund_categories[cat]['debt_amount'] = float(row[3]) if row[3] else 0
                fund_categories[cat]['records'] = row[4]
            if cat == 'IPD':
                ipd_total = float(row[1]) if row[1] else 0
                ipd_wait = float(row[2]) if row[2] else 0
                ipd_debt = float(row[3]) if row[3] else 0
            elif cat == 'OPD':
                opd_total = float(row[1]) if row[1] else 0

        # Calculate per-bed metrics from IPD only (beds are for inpatients)
        ipd_revenue_per_bed = (ipd_total / actual_beds) if actual_beds > 0 else 0
        ipd_wait_per_bed = (ipd_wait / actual_beds) if actual_beds > 0 else 0
        ipd_debt_per_bed = (ipd_debt / actual_beds) if actual_beds > 0 else 0

        # Calculate ratios for each fund category
        for cat in fund_categories:
            cat_amount = fund_categories[cat]['total_amount']
            fund_categories[cat]['ratio'] = round((cat_amount / total_amount * 100) if total_amount > 0 else 0, 1)

        summary = {
            'total_amount': total_amount,
            'wait_amount': wait_amount,
            'debt_amount': debt_amount,
            'bond_amount': float(summary_row[4]) if summary_row else 0,
            'wait_ratio': (wait_amount / total_amount * 100) if total_amount > 0 else 0,
            'debt_ratio': (debt_amount / total_amount * 100) if total_amount > 0 else 0,
            'record_count': summary_row[0] if summary_row else 0,
            'growth_yoy': round(growth_yoy, 1),
            # Per-bed metrics (all from IPD since beds are for inpatients)
            'actual_beds': actual_beds,
            'revenue_per_bed': round(ipd_revenue_per_bed, 2),  # IPD revenue per bed
            'ipd_revenue_per_bed': round(ipd_revenue_per_bed, 2),  # same as above (for backward compat)
            'wait_per_bed': round(ipd_wait_per_bed, 2),  # IPD wait per bed
            'debt_per_bed': round(ipd_debt_per_bed, 2),  # IPD debt per bed
            # Fund category breakdown (OPD, IPD, CR, PP, OTHER)
            'fund_categories': fund_categories,
            # Legacy OPD/IPD fields for backward compatibility
            'ipd_amount': ipd_total,
            'opd_amount': opd_total,
            'ipd_ratio': round((ipd_total / total_amount * 100) if total_amount > 0 else 0, 1),
            'opd_ratio': round((opd_total / total_amount * 100) if total_amount > 0 else 0, 1)
        }

        # Get fund breakdown
        cursor.execute("""
            SELECT
                fund_name,
                fund_group,
                fund_group_desc,
                COALESCE(SUM(total_amount), 0) as amount,
                COUNT(*) as records
            FROM smt_budget_transfers
            WHERE (vendor_no = %s OR vendor_no = %s)
              AND run_date >= %s AND run_date <= %s
            GROUP BY fund_name, fund_group, fund_group_desc
            ORDER BY amount DESC
        """, (vendor_id_10, vendor_id_5, start_date, end_date))

        fund_rows = cursor.fetchall()
        fund_breakdown = []
        for row in fund_rows:
            amount = float(row[3]) if row[3] else 0
            fund_breakdown.append({
                'fund_name': row[0] or 'ไม่ระบุ',
                'fund_group': row[1],
                'fund_group_desc': row[2],
                'amount': amount,
                'percentage': (amount / total_amount * 100) if total_amount > 0 else 0,
                'records': row[4]
            })

        # Get monthly trend
        cursor.execute("""
            SELECT
                """ + sql_format_year_month('run_date') + """ as month,
                COALESCE(SUM(total_amount), 0) as total_amount,
                COALESCE(SUM(wait_amount), 0) as wait_amount,
                COALESCE(SUM(debt_amount), 0) as debt_amount,
                COUNT(*) as records
            FROM smt_budget_transfers
            WHERE (vendor_no = %s OR vendor_no = %s)
              AND run_date >= %s AND run_date <= %s
            GROUP BY """ + sql_format_year_month('run_date') + """
            ORDER BY month
        """, (vendor_id_10, vendor_id_5, start_date, end_date))

        monthly_rows = cursor.fetchall()
        monthly_trend = []
        for row in monthly_rows:
            monthly_trend.append({
                'month': row[0],
                'total_amount': float(row[1]) if row[1] else 0,
                'wait_amount': float(row[2]) if row[2] else 0,
                'debt_amount': float(row[3]) if row[3] else 0,
                'records': row[4]
            })

        # Calculate risk score
        wait_ratio = summary['wait_ratio']
        debt_ratio = summary['debt_ratio']

        wait_score = min(100, (wait_ratio / 20) * 100)  # 20% = max risk
        debt_score = min(100, (debt_ratio / 15) * 100)  # 15% = max risk
        # Growth score: positive growth = no risk, negative growth = risk (capped at 100)
        if growth_yoy >= 0:
            growth_score = 0
        else:
            growth_score = min(100, abs(growth_yoy) * 2)  # -50% growth = 100 risk

        risk_score = int(wait_score * 0.3 + debt_score * 0.4 + growth_score * 0.3)
        risk_level = 'low' if risk_score < 40 else ('medium' if risk_score < 70 else 'high')

        risk_assessment = {
            'score': risk_score,
            'level': risk_level,
            'indicators': [
                {'name': 'Wait Ratio', 'value': round(wait_ratio, 1), 'threshold': 10, 'status': 'pass' if wait_ratio < 10 else 'fail'},
                {'name': 'Debt Ratio', 'value': round(debt_ratio, 1), 'threshold': 5, 'status': 'pass' if debt_ratio < 5 else 'fail'},
                {'name': 'Growth YoY', 'value': round(growth_yoy, 1), 'threshold': 0, 'status': 'pass' if growth_yoy > 0 else 'fail'}
            ]
        }

        # Get ranking (national)
        cursor.execute("""
            WITH hospital_totals AS (
                SELECT
                    vendor_no,
                    SUM(total_amount) as total
                FROM smt_budget_transfers
                WHERE run_date >= %s AND run_date <= %s
                GROUP BY vendor_no
            )
            SELECT
                COUNT(*) as total_hospitals,
                SUM(CASE WHEN total > %s THEN 1 ELSE 0 END) as hospitals_above
            FROM hospital_totals
        """, (start_date, end_date, total_amount))
        rank_row = cursor.fetchone()
        total_hospitals = int(rank_row[0]) if rank_row and rank_row[0] else 0
        hospitals_above = int(rank_row[1]) if rank_row and rank_row[1] else 0
        national_rank = hospitals_above + 1

        ranking = {
            'national': {
                'rank': national_rank,
                'total': total_hospitals,
                'percentile': int((1 - national_rank / total_hospitals) * 100) if total_hospitals > 0 else 0
            }
        }

        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'fiscal_year': fiscal_year,
            'hospital': hospital_info,
            'summary': summary,
            'fund_breakdown': fund_breakdown,
            'monthly_trend': monthly_trend,
            'risk_score': risk_assessment,
            'ranking': ranking
        })

    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/benchmark/region-average')
def api_benchmark_region_average():
    """Get regional average for comparison"""
    try:
        health_region = request.args.get('health_region')
        fiscal_year = request.args.get('fiscal_year')

        if not health_region:
            return jsonify({'success': False, 'error': 'health_region is required'}), 400

        # Default to current fiscal year
        if not fiscal_year:
            today = datetime.now(TZ_BANGKOK)
            fiscal_year = today.year + 543 if today.month >= 10 else today.year + 542

        fiscal_year = int(fiscal_year)
        # health_region in DB is "เขตสุขภาพที่ X", build match pattern
        health_region_pattern = f"เขตสุขภาพที่ {health_region}"

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Convert fiscal year to date range using standardized calculation
        start_date, end_date = get_fiscal_year_range_gregorian(fiscal_year)

        # Get regional averages
        ltrim_expr = "TRIM(LEADING '0' FROM s.vendor_no)" if DB_TYPE == 'mysql' else "LTRIM(s.vendor_no, '0')"
        cursor.execute(f"""
            WITH hospital_totals AS (
                SELECT
                    s.vendor_no,
                    SUM(s.total_amount) as total_amount,
                    SUM(s.wait_amount) as wait_amount,
                    SUM(s.debt_amount) as debt_amount
                FROM smt_budget_transfers s
                JOIN health_offices h ON (
                    h.hcode5 = {ltrim_expr}
                    OR h.hcode5 = s.vendor_no
                )
                WHERE h.health_region = %s
                  AND s.run_date >= %s AND s.run_date <= %s
                GROUP BY s.vendor_no
            )
            SELECT
                COUNT(*) as hospital_count,
                COALESCE(AVG(total_amount), 0) as avg_total,
                COALESCE(AVG(wait_amount), 0) as avg_wait,
                COALESCE(AVG(debt_amount), 0) as avg_debt,
                COALESCE(SUM(total_amount), 0) as sum_total
            FROM hospital_totals
        """, (health_region_pattern, start_date, end_date))
        avg_row = cursor.fetchone()

        averages = {
            'hospital_count': avg_row[0] if avg_row else 0,
            'avg_total_amount': float(avg_row[1]) if avg_row else 0,
            'avg_wait_amount': float(avg_row[2]) if avg_row else 0,
            'avg_debt_amount': float(avg_row[3]) if avg_row else 0,
            'total_amount': float(avg_row[4]) if avg_row else 0
        }

        avg_total = averages['avg_total_amount']
        averages['avg_wait_ratio'] = (averages['avg_wait_amount'] / avg_total * 100) if avg_total > 0 else 0
        averages['avg_debt_ratio'] = (averages['avg_debt_amount'] / avg_total * 100) if avg_total > 0 else 0

        # Get fund breakdown averages for region
        cursor.execute(f"""
            WITH hospital_funds AS (
                SELECT
                    s.vendor_no,
                    s.fund_name,
                    s.fund_group,
                    SUM(s.total_amount) as amount
                FROM smt_budget_transfers s
                JOIN health_offices h ON (
                    h.hcode5 = {ltrim_expr}
                    OR h.hcode5 = s.vendor_no
                )
                WHERE h.health_region = %s
                  AND s.run_date >= %s AND s.run_date <= %s
                GROUP BY s.vendor_no, s.fund_name, s.fund_group
            )
            SELECT
                fund_name,
                fund_group,
                AVG(amount) as avg_amount,
                SUM(amount) as total_amount
            FROM hospital_funds
            GROUP BY fund_name, fund_group
            ORDER BY total_amount DESC
        """, (health_region_pattern, start_date, end_date))

        fund_rows = cursor.fetchall()
        fund_breakdown = []
        for row in fund_rows:
            fund_breakdown.append({
                'fund_name': row[0] or 'ไม่ระบุ',
                'fund_group': row[1],
                'avg_amount': float(row[2]) if row[2] else 0,
                'total_amount': float(row[3]) if row[3] else 0
            })

        # Get monthly trend for region
        month_format = sql_format_year_month('s.run_date')
        cursor.execute(f"""
            SELECT
                {month_format} as month,
                COALESCE(SUM(s.total_amount), 0) as total_amount,
                COALESCE(AVG(s.total_amount), 0) as avg_amount
            FROM smt_budget_transfers s
            JOIN health_offices h ON (
                h.hcode5 = {ltrim_expr}
                OR h.hcode5 = s.vendor_no
            )
            WHERE h.health_region = %s
              AND s.run_date >= %s AND s.run_date <= %s
            GROUP BY {month_format}
            ORDER BY month
        """, (health_region_pattern, start_date, end_date))

        monthly_rows = cursor.fetchall()
        monthly_trend = []
        for row in monthly_rows:
            monthly_trend.append({
                'month': row[0],
                'total_amount': float(row[1]) if row[1] else 0,
                'avg_amount': float(row[2]) if row[2] else 0
            })

        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'region': request.args.get('health_region'),
            'fiscal_year': fiscal_year,
            'averages': averages,
            'fund_breakdown': fund_breakdown,
            'monthly_trend': monthly_trend
        })

    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/benchmark/hospitals/<vendor_no>', methods=['DELETE'])
def api_benchmark_delete_hospital(vendor_no):
    """Delete SMT data for a specific hospital"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Delete all records for this vendor
        cursor.execute("""
            DELETE FROM smt_budget_transfers
            WHERE vendor_no = %s OR vendor_no = %s
        """, (vendor_no, vendor_no.zfill(10)))

        deleted = cursor.rowcount
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'message': f'Deleted {deleted} records for vendor {vendor_no}'
        })

    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/benchmark/available-years')
def api_benchmark_available_years():
    """Get list of fiscal years that have SMT data in the database"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Get distinct fiscal years from smt_budget_transfers
        # Fiscal year is determined by run_date: Oct-Dec = next year, Jan-Sep = current year
        year_expr = sql_extract_year('run_date')
        month_expr = sql_extract_month('run_date')
        cursor.execute(f"""
            SELECT DISTINCT
                CASE
                    WHEN {month_expr} >= 10 THEN {year_expr} + 544
                    ELSE {year_expr} + 543
                END as fiscal_year
            FROM smt_budget_transfers
            WHERE run_date IS NOT NULL
            ORDER BY fiscal_year DESC
        """)

        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        years = [int(row[0]) for row in rows]

        return jsonify({
            'success': True,
            'years': years
        })

    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# Phase 3: Predictive & AI Analytics
# ============================================

@app.route('/predictive')
def predictive():
    """Phase 3: Predictive Analytics page"""
    return render_template('predictive.html')


def calculate_performance_metrics(summary, monthly_data):
    """
    Calculate Claims Performance Metrics
    Returns dict with:
    - payment_accuracy: % of months with matched amounts
    - claim_to_payment_ratio: SMT/REP amount ratio
    - underpayment_rate: % of months underpaid
    - pending_rate: % of months pending payment
    - avg_difference: Average difference per claim
    - top_month: Month with highest claim amount
    - matched_months: Count of matched months
    - underpaid_months: Count of underpaid months
    - pending_months: Count of pending months
    - total_months: Total months with data
    - data_quality: Data quality metrics
    """
    if not summary or not monthly_data:
        return {
            'payment_accuracy': 0,
            'claim_to_payment_ratio': 0,
            'underpayment_rate': 0,
            'pending_rate': 0,
            'avg_difference': 0,
            'top_month': None,
            'matched_months': 0,
            'underpaid_months': 0,
            'pending_months': 0,
            'total_months': 0,
            'data_quality': {
                'coverage_rate': 0,
                'alert_level': 'unknown',
                'missing_amount': 0,
                'expected_claims': 0,
                'actual_claims': 0,
                'missing_claims': 0,
                'avg_claim_amount': 0
            }
        }

    total_months = len(monthly_data)
    matched_months = 0
    underpaid_months = 0
    pending_months = 0
    top_month = None
    max_claim_total = 0

    for month in monthly_data:
        has_rep = month.get('has_rep_data', False)
        has_smt = month.get('has_smt_data', False)
        difference = month.get('difference', 0)
        claim_total = month.get('claim_total', 0)

        # Count matched months (difference < 1% of claim total)
        if has_rep and has_smt and claim_total > 0:
            diff_percent = abs(difference / claim_total)
            if diff_percent < 0.01:  # Less than 1%
                matched_months += 1

        # Count underpaid months
        if has_rep and has_smt and difference < 0:
            underpaid_months += 1

        # Count pending months (has REP but no SMT)
        if has_rep and not has_smt:
            pending_months += 1

        # Find top month
        if has_rep and claim_total > max_claim_total:
            max_claim_total = claim_total
            top_month = month

    # Calculate ratios
    payment_accuracy = (matched_months / total_months * 100) if total_months > 0 else 0
    underpayment_rate = (underpaid_months / total_months * 100) if total_months > 0 else 0
    pending_rate = (pending_months / total_months * 100) if total_months > 0 else 0

    rep_total = summary.get('rep', {}).get('total_amount', 0)
    smt_total = summary.get('smt', {}).get('total_amount', 0)
    total_claims = summary.get('rep', {}).get('total_claims', 0)

    claim_to_payment_ratio = (smt_total / rep_total * 100) if rep_total > 0 else 0
    avg_difference = ((smt_total - rep_total) / total_claims) if total_claims > 0 else 0

    # Calculate Data Quality Metrics
    # Coverage Rate: REP / SMT (%)
    coverage_rate = (rep_total / smt_total * 100) if smt_total > 0 else 0

    # Determine alert level based on coverage
    if coverage_rate < 20:
        alert_level = 'critical'  # Red - Very low coverage
    elif coverage_rate < 50:
        alert_level = 'warning'   # Yellow - Low coverage
    elif coverage_rate < 80:
        alert_level = 'caution'   # Orange - Moderate coverage
    else:
        alert_level = 'good'      # Green - Good coverage

    # Estimate missing claims
    avg_claim_amount = rep_total / total_claims if total_claims > 0 else 0
    expected_claims = int(smt_total / avg_claim_amount) if avg_claim_amount > 0 else 0
    missing_claims = max(0, expected_claims - total_claims)
    missing_amount = smt_total - rep_total if smt_total > rep_total else 0

    data_quality = {
        'coverage_rate': coverage_rate,
        'alert_level': alert_level,
        'missing_amount': missing_amount,
        'expected_claims': expected_claims,
        'actual_claims': total_claims,
        'missing_claims': missing_claims,
        'avg_claim_amount': avg_claim_amount
    }

    return {
        'payment_accuracy': payment_accuracy,
        'claim_to_payment_ratio': claim_to_payment_ratio,
        'underpayment_rate': underpayment_rate,
        'pending_rate': pending_rate,
        'avg_difference': avg_difference,
        'top_month': top_month,
        'matched_months': matched_months,
        'underpaid_months': underpaid_months,
        'pending_months': pending_months,
        'total_months': total_months,
        'data_quality': data_quality
    }


@app.route('/reconciliation')
def reconciliation():
    """Reconciliation Report page - Compare REP claims vs SMT payments"""
    from utils.reconciliation import ReconciliationReport

    # Get fiscal year from query param
    fiscal_year = request.args.get('fy', type=int)

    # Get hospital code from settings
    hospital_code = settings_manager.get_hospital_code()

    conn = get_db_connection()
    if not conn:
        return render_template(
            'reconciliation.html',
            error='Database connection failed',
            summary=None,
            monthly_data=[],
            fund_data=[],
            fiscal_years=[],
            selected_fy=None,
            performance_metrics=None
        )

    try:
        report = ReconciliationReport(conn, hospital_code)

        # Get available fiscal years
        fiscal_years = report.get_available_fiscal_years()

        # Default to current fiscal year if not specified
        if not fiscal_year and fiscal_years:
            fiscal_year = fiscal_years[0]

        # Get data for selected fiscal year
        if fiscal_year:
            summary = report.get_summary_stats_by_fy(fiscal_year)
            monthly_data = report.get_monthly_reconciliation_by_fy(fiscal_year)
        else:
            summary = report.get_summary_stats()
            monthly_data = report.get_monthly_reconciliation()

        fund_data = report.get_fund_reconciliation()

        # Calculate performance metrics
        performance_metrics = calculate_performance_metrics(summary, monthly_data)

        conn.close()

        return render_template(
            'reconciliation.html',
            summary=summary,
            monthly_data=monthly_data,
            fund_data=fund_data,
            fiscal_years=fiscal_years,
            selected_fy=fiscal_year,
            performance_metrics=performance_metrics
        )
    except Exception as e:
        app.logger.error(f"Reconciliation error: {e}")
        if conn:
            conn.close()
        return render_template(
            'reconciliation.html',
            error=str(e),
            summary=None,
            monthly_data=[],
            fund_data=[],
            fiscal_years=[],
            selected_fy=None,
            performance_metrics=None
        )


@app.route('/api/reconciliation/fiscal-years')
def api_reconciliation_fiscal_years():
    """Get available fiscal years"""
    from utils.reconciliation import ReconciliationReport

    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        report = ReconciliationReport(conn, settings_manager.get_hospital_code())
        fiscal_years = report.get_available_fiscal_years()
        conn.close()

        return jsonify({
            'success': True,
            'data': fiscal_years
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reconciliation/monthly')
def api_reconciliation_monthly():
    """Get monthly reconciliation data"""
    from utils.reconciliation import ReconciliationReport

    try:
        fiscal_year = request.args.get('fy', type=int)

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        report = ReconciliationReport(conn, settings_manager.get_hospital_code())

        if fiscal_year:
            data = report.get_monthly_reconciliation_by_fy(fiscal_year)
        else:
            data = report.get_monthly_reconciliation()

        conn.close()

        return jsonify({
            'success': True,
            'fiscal_year': fiscal_year,
            'data': data
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reconciliation/fund')
def api_reconciliation_fund():
    """Get fund-based reconciliation data"""
    from utils.reconciliation import ReconciliationReport

    try:
        month_be = request.args.get('month')  # Optional: YYYYMM in Buddhist Era

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        report = ReconciliationReport(conn, settings_manager.get_hospital_code())
        data = report.get_fund_reconciliation(month_be)
        conn.close()

        return jsonify({
            'success': True,
            'data': data
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reconciliation/summary')
def api_reconciliation_summary():
    """Get overall reconciliation summary stats"""
    from utils.reconciliation import ReconciliationReport

    try:
        fiscal_year = request.args.get('fy', type=int)

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        report = ReconciliationReport(conn, settings_manager.get_hospital_code())

        if fiscal_year:
            summary = report.get_summary_stats_by_fy(fiscal_year)
        else:
            summary = report.get_summary_stats()

        conn.close()

        return jsonify({
            'success': True,
            'data': summary
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reconciliation/rep-monthly')
def api_rep_monthly():
    """Get REP monthly summary by fund"""
    from utils.reconciliation import ReconciliationReport

    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        report = ReconciliationReport(conn, settings_manager.get_hospital_code())
        data = report.get_rep_monthly_summary()
        conn.close()

        return jsonify({
            'success': True,
            'data': data
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reconciliation/smt-monthly')
def api_smt_monthly():
    """Get SMT monthly summary by fund"""
    from utils.reconciliation import ReconciliationReport

    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        report = ReconciliationReport(conn, settings_manager.get_hospital_code())
        data = report.get_smt_monthly_summary()
        conn.close()

        return jsonify({
            'success': True,
            'data': data
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/db/pool-status')
def api_db_pool_status():
    """Get database connection pool status"""
    try:
        status = get_pool_status()
        return jsonify({
            'success': True,
            'pool': status
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# Health Offices Master Data Management
# ============================================

@app.route('/health-offices')
def health_offices_page():
    """Redirect to Data Management - Health Offices tab"""
    return redirect(url_for('data_management') + '?tab=offices')


@app.route('/api/health-offices')
def api_health_offices_list():
    """Get health offices with filtering and pagination"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Get filter parameters
        search = request.args.get('search', '').strip()
        province = request.args.get('province', '')
        status = request.args.get('status', '')
        level = request.args.get('level', '')
        region = request.args.get('region', '')
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        offset = (page - 1) * per_page

        # Build query
        where_clauses = []
        params = []

        if search:
            like_op = "LIKE" if DB_TYPE == 'mysql' else "ILIKE"
            where_clauses.append(f"(name {like_op} %s OR hcode5 {like_op} %s OR hcode9 {like_op} %s)")
            params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])
        if province:
            where_clauses.append("province = %s")
            params.append(province)
        if status:
            where_clauses.append("status = %s")
            params.append(status)
        if level:
            where_clauses.append("hospital_level = %s")
            params.append(level)
        if region:
            where_clauses.append("health_region = %s")
            params.append(region)

        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

        # Get total count
        cursor.execute(f"SELECT COUNT(*) FROM health_offices WHERE {where_sql}", params)
        total = cursor.fetchone()[0]

        # Get data
        cursor.execute(f"""
            SELECT id, name, hcode5, hcode9, org_type, service_type, hospital_level,
                   actual_beds, status, health_region, province, district, address
            FROM health_offices
            WHERE {where_sql}
            ORDER BY name
            LIMIT %s OFFSET %s
        """, params + [per_page, offset])

        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        offices = []
        for row in rows:
            offices.append({
                'id': row[0],
                'name': row[1],
                'hcode5': row[2],
                'hcode9': row[3],
                'org_type': row[4],
                'service_type': row[5],
                'hospital_level': row[6],
                'actual_beds': row[7],
                'status': row[8],
                'health_region': row[9],
                'province': row[10],
                'district': row[11],
                'address': row[12]
            })

        return jsonify({
            'success': True,
            'data': offices,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': (total + per_page - 1) // per_page
        })

    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/health-offices/stats')
def api_health_offices_stats():
    """Get health offices statistics"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        stats = {}

        # Total count
        cursor.execute("SELECT COUNT(*) FROM health_offices")
        stats['total'] = cursor.fetchone()[0]

        # By status
        cursor.execute("""
            SELECT status, COUNT(*) FROM health_offices
            WHERE status IS NOT NULL
            GROUP BY status ORDER BY COUNT(*) DESC
        """)
        stats['by_status'] = [{'status': r[0], 'count': r[1]} for r in cursor.fetchall()]

        # By hospital level
        cursor.execute("""
            SELECT hospital_level, COUNT(*) FROM health_offices
            WHERE hospital_level IS NOT NULL
            GROUP BY hospital_level ORDER BY COUNT(*) DESC
        """)
        stats['by_level'] = [{'level': r[0], 'count': r[1]} for r in cursor.fetchall()]

        # By region
        cursor.execute("""
            SELECT health_region, COUNT(*) FROM health_offices
            WHERE health_region IS NOT NULL
            GROUP BY health_region ORDER BY health_region
        """)
        stats['by_region'] = [{'region': r[0], 'count': r[1]} for r in cursor.fetchall()]

        # By province (top 10)
        cursor.execute("""
            SELECT province, COUNT(*) FROM health_offices
            WHERE province IS NOT NULL
            GROUP BY province ORDER BY COUNT(*) DESC LIMIT 10
        """)
        stats['by_province_top10'] = [{'province': r[0], 'count': r[1]} for r in cursor.fetchall()]

        # Get distinct values for filters
        cursor.execute("SELECT DISTINCT province FROM health_offices WHERE province IS NOT NULL ORDER BY province")
        stats['provinces'] = [r[0] for r in cursor.fetchall()]

        cursor.execute("SELECT DISTINCT health_region FROM health_offices WHERE health_region IS NOT NULL ORDER BY health_region")
        stats['regions'] = [r[0] for r in cursor.fetchall()]

        cursor.execute("SELECT DISTINCT hospital_level FROM health_offices WHERE hospital_level IS NOT NULL ORDER BY hospital_level")
        stats['levels'] = [r[0] for r in cursor.fetchall()]

        # Last import info
        cursor.execute("""
            SELECT import_date, filename, total_records, imported, import_mode
            FROM health_offices_import_log
            ORDER BY import_date DESC LIMIT 1
        """)
        last_import = cursor.fetchone()
        if last_import:
            stats['last_import'] = {
                'date': last_import[0].strftime('%Y-%m-%d %H:%M') if last_import[0] else None,
                'filename': last_import[1],
                'total_records': last_import[2],
                'imported': last_import[3],
                'mode': last_import[4]
            }

        cursor.close()
        conn.close()

        return jsonify({'success': True, 'stats': stats})

    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/health-offices/import', methods=['POST'])
def api_health_offices_import():
    """Import health offices from uploaded Excel file"""
    import time
    import pandas as pd
    from openpyxl import load_workbook
    import re
    import io

    def parse_formula_value(val):
        """Parse Excel formula value like ='32045' or =\"32045\" to plain value"""
        if val is None:
            return None
        val_str = str(val)
        # Match ="xxx" or ='xxx' format
        match = re.match(r'^=["\'](.*)["\']\s*$', val_str)
        if match:
            return match.group(1)
        return val_str if val_str.lower() not in ('none', 'nan', '') else None

    try:
        start_time = time.time()

        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400

        import_mode = request.form.get('mode', 'upsert')  # 'upsert' or 'replace'

        # Read Excel file using openpyxl to handle formula values
        file_bytes = io.BytesIO(file.read())
        wb = load_workbook(file_bytes, data_only=False)
        ws = wb.active

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        # Read data rows with formula parsing for code columns
        data = []
        formula_columns = {'รหัส 9 หลักใหม่', 'รหัส 9 หลัก', 'รหัส 5 หลัก', 'เลขอนุญาตให้ประกอบสถานบริการสุขภาพ 11 หลัก'}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            row_data = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers):
                    header = headers[col_idx]
                    if header in formula_columns:
                        row_data[header] = parse_formula_value(cell.value)
                    else:
                        row_data[header] = cell.value
            data.append(row_data)

        df = pd.DataFrame(data)

        # Column mapping
        column_map = {
            'ชื่อ': 'name',
            'รหัส 9 หลักใหม่': 'hcode9_new',
            'รหัส 9 หลัก': 'hcode9',
            'รหัส 5 หลัก': 'hcode5',
            'เลขอนุญาตให้ประกอบสถานบริการสุขภาพ 11 หลัก': 'license_no',
            'ประเภทองค์กร': 'org_type',
            'ประเภทหน่วยบริการสุขภาพ': 'service_type',
            'สังกัด': 'affiliation',
            'แผนก/กรม': 'department',
            'ระดับโรงพยาบาล': 'hospital_level',
            'เตียงที่ใช้จริง': 'actual_beds',
            'สถานะการใช้งาน': 'status',
            'เขตบริการ': 'health_region',
            'ที่อยู่': 'address',
            'รหัสจังหวัด': 'province_code',
            'จังหวัด': 'province',
            'รหัสอำเภอ': 'district_code',
            'อำเภอ/เขต': 'district',
            'รหัสตำบล': 'subdistrict_code',
            'ตำบล/แขวง': 'subdistrict',
            'หมู่': 'moo',
            'รหัสไปรษณีย์': 'postal_code',
            'แม่ข่าย': 'parent_code',
            'วันที่ก่อตั้ง': 'established_date',
            'วันที่ปิดบริการ': 'closed_date',
            'อัพเดตล่าสุด(เริ่ม 05/09/2566)': 'source_updated_at'
        }

        # Rename columns
        df = df.rename(columns=column_map)

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Clear existing data if replace mode
        if import_mode == 'replace':
            cursor.execute("TRUNCATE TABLE health_offices RESTART IDENTITY")

        imported = 0
        updated = 0
        skipped = 0
        errors = 0
        total_records = len(df)

        for idx, row in df.iterrows():
            try:
                # Convert codes to string (handle both formula-parsed strings and numbers)
                def clean_code(val):
                    if pd.isna(val) or val is None or str(val).lower() in ('none', 'nan', ''):
                        return None
                    # Remove any whitespace and convert to string
                    return str(val).strip()

                hcode5 = clean_code(row.get('hcode5'))
                hcode9 = clean_code(row.get('hcode9'))
                hcode9_new = clean_code(row.get('hcode9_new'))

                # Parse dates
                def parse_date(val):
                    if pd.isna(val):
                        return None
                    if isinstance(val, str):
                        try:
                            # Try dd/mm/yyyy format
                            parts = val.split('/')
                            if len(parts) == 3:
                                day, month, year = parts
                                if int(year) > 2500:
                                    year = str(int(year) - 543)
                                return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                        except:
                            pass
                    return None

                established = parse_date(row.get('established_date'))
                closed = parse_date(row.get('closed_date'))
                source_updated = parse_date(row.get('source_updated_at'))

                # Prepare values
                values = {
                    'name': row.get('name'),
                    'hcode9_new': hcode9_new,
                    'hcode9': hcode9,
                    'hcode5': hcode5,
                    'license_no': clean_code(row.get('license_no')),
                    'org_type': row.get('org_type'),
                    'service_type': row.get('service_type'),
                    'affiliation': row.get('affiliation'),
                    'department': row.get('department'),
                    'hospital_level': row.get('hospital_level'),
                    'actual_beds': int(row['actual_beds']) if pd.notna(row.get('actual_beds')) else 0,
                    'status': row.get('status'),
                    'health_region': row.get('health_region'),
                    'address': row.get('address'),
                    'province_code': str(int(row['province_code'])) if pd.notna(row.get('province_code')) else None,
                    'province': row.get('province'),
                    'district_code': str(int(row['district_code'])) if pd.notna(row.get('district_code')) else None,
                    'district': row.get('district'),
                    'subdistrict_code': str(int(row['subdistrict_code'])) if pd.notna(row.get('subdistrict_code')) else None,
                    'subdistrict': row.get('subdistrict'),
                    'moo': row.get('moo'),
                    'postal_code': row.get('postal_code'),
                    'parent_code': row.get('parent_code'),
                    'established_date': established,
                    'closed_date': closed,
                    'source_updated_at': source_updated
                }

                # Skip if no name
                if not values['name'] or pd.isna(values['name']):
                    skipped += 1
                    continue

                # Insert or update
                if import_mode == 'replace' or not hcode5:
                    # Insert only
                    cursor.execute("""
                        INSERT INTO health_offices (
                            name, hcode9_new, hcode9, hcode5, license_no, org_type, service_type,
                            affiliation, department, hospital_level, actual_beds, status, health_region,
                            address, province_code, province, district_code, district, subdistrict_code,
                            subdistrict, moo, postal_code, parent_code, established_date, closed_date,
                            source_updated_at
                        ) VALUES (
                            %(name)s, %(hcode9_new)s, %(hcode9)s, %(hcode5)s, %(license_no)s, %(org_type)s,
                            %(service_type)s, %(affiliation)s, %(department)s, %(hospital_level)s,
                            %(actual_beds)s, %(status)s, %(health_region)s, %(address)s, %(province_code)s,
                            %(province)s, %(district_code)s, %(district)s, %(subdistrict_code)s,
                            %(subdistrict)s, %(moo)s, %(postal_code)s, %(parent_code)s, %(established_date)s,
                            %(closed_date)s, %(source_updated_at)s
                        )
                    """, values)
                    imported += 1
                else:
                    # Upsert by hcode5 - use appropriate syntax based on database type
                    if DB_TYPE == 'mysql':
                        cursor.execute("""
                            INSERT INTO health_offices (
                                name, hcode9_new, hcode9, hcode5, license_no, org_type, service_type,
                                affiliation, department, hospital_level, actual_beds, status, health_region,
                                address, province_code, province, district_code, district, subdistrict_code,
                                subdistrict, moo, postal_code, parent_code, established_date, closed_date,
                                source_updated_at
                            ) VALUES (
                                %(name)s, %(hcode9_new)s, %(hcode9)s, %(hcode5)s, %(license_no)s, %(org_type)s,
                                %(service_type)s, %(affiliation)s, %(department)s, %(hospital_level)s,
                                %(actual_beds)s, %(status)s, %(health_region)s, %(address)s, %(province_code)s,
                                %(province)s, %(district_code)s, %(district)s, %(subdistrict_code)s,
                                %(subdistrict)s, %(moo)s, %(postal_code)s, %(parent_code)s, %(established_date)s,
                                %(closed_date)s, %(source_updated_at)s
                            )
                            ON DUPLICATE KEY UPDATE
                                name = VALUES(name),
                                hcode9_new = VALUES(hcode9_new),
                                hcode9 = VALUES(hcode9),
                                license_no = VALUES(license_no),
                                org_type = VALUES(org_type),
                                service_type = VALUES(service_type),
                                affiliation = VALUES(affiliation),
                                department = VALUES(department),
                                hospital_level = VALUES(hospital_level),
                                actual_beds = VALUES(actual_beds),
                                status = VALUES(status),
                                health_region = VALUES(health_region),
                                address = VALUES(address),
                                province_code = VALUES(province_code),
                                province = VALUES(province),
                                district_code = VALUES(district_code),
                                district = VALUES(district),
                                subdistrict_code = VALUES(subdistrict_code),
                                subdistrict = VALUES(subdistrict),
                                moo = VALUES(moo),
                                postal_code = VALUES(postal_code),
                                parent_code = VALUES(parent_code),
                                established_date = VALUES(established_date),
                                closed_date = VALUES(closed_date),
                                source_updated_at = VALUES(source_updated_at),
                                updated_at = CURRENT_TIMESTAMP
                        """, values)
                    else:
                        cursor.execute("""
                            INSERT INTO health_offices (
                                name, hcode9_new, hcode9, hcode5, license_no, org_type, service_type,
                                affiliation, department, hospital_level, actual_beds, status, health_region,
                                address, province_code, province, district_code, district, subdistrict_code,
                                subdistrict, moo, postal_code, parent_code, established_date, closed_date,
                                source_updated_at
                            ) VALUES (
                                %(name)s, %(hcode9_new)s, %(hcode9)s, %(hcode5)s, %(license_no)s, %(org_type)s,
                                %(service_type)s, %(affiliation)s, %(department)s, %(hospital_level)s,
                                %(actual_beds)s, %(status)s, %(health_region)s, %(address)s, %(province_code)s,
                                %(province)s, %(district_code)s, %(district)s, %(subdistrict_code)s,
                                %(subdistrict)s, %(moo)s, %(postal_code)s, %(parent_code)s, %(established_date)s,
                                %(closed_date)s, %(source_updated_at)s
                            )
                            ON CONFLICT (hcode5) DO UPDATE SET
                                name = EXCLUDED.name,
                                hcode9_new = EXCLUDED.hcode9_new,
                                hcode9 = EXCLUDED.hcode9,
                                license_no = EXCLUDED.license_no,
                                org_type = EXCLUDED.org_type,
                                service_type = EXCLUDED.service_type,
                                affiliation = EXCLUDED.affiliation,
                                department = EXCLUDED.department,
                                hospital_level = EXCLUDED.hospital_level,
                                actual_beds = EXCLUDED.actual_beds,
                                status = EXCLUDED.status,
                                health_region = EXCLUDED.health_region,
                                address = EXCLUDED.address,
                                province_code = EXCLUDED.province_code,
                                province = EXCLUDED.province,
                                district_code = EXCLUDED.district_code,
                                district = EXCLUDED.district,
                                subdistrict_code = EXCLUDED.subdistrict_code,
                                subdistrict = EXCLUDED.subdistrict,
                                moo = EXCLUDED.moo,
                                postal_code = EXCLUDED.postal_code,
                                parent_code = EXCLUDED.parent_code,
                                established_date = EXCLUDED.established_date,
                                closed_date = EXCLUDED.closed_date,
                                source_updated_at = EXCLUDED.source_updated_at,
                                updated_at = CURRENT_TIMESTAMP
                        """, values)
                    if cursor.rowcount > 0:
                        imported += 1

            except Exception as e:
                errors += 1
                if errors <= 5:
                    app.logger.error(f"Error importing row {idx}: {e}")

        # Log import
        duration = time.time() - start_time
        cursor.execute("""
            INSERT INTO health_offices_import_log
            (filename, total_records, imported, updated, skipped, errors, import_mode, duration_seconds)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (file.filename, total_records, imported, updated, skipped, errors, import_mode, duration))

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'message': f'Import completed',
            'total': total_records,
            'imported': imported,
            'updated': updated,
            'skipped': skipped,
            'errors': errors,
            'duration': round(duration, 2)
        })

    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/health-offices/clear', methods=['POST'])
def api_health_offices_clear():
    """Clear all health offices data"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()
        cursor.execute("TRUNCATE TABLE health_offices RESTART IDENTITY")
        conn.commit()

        cursor.execute("SELECT COUNT(*) FROM health_offices")
        count = cursor.fetchone()[0]

        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'message': 'All health offices data cleared',
            'remaining': count
        })

    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/health-offices/lookup/<code>')
def api_health_offices_lookup(code):
    """Lookup health office by code (hcode5 or hcode9)"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, name, hcode5, hcode9, hcode9_new, org_type, service_type,
                   hospital_level, actual_beds, status, health_region, province,
                   district, address
            FROM health_offices
            WHERE hcode5 = %s OR hcode9 = %s OR hcode5 = %s
            LIMIT 1
        """, (code, code, code.zfill(5)))

        row = cursor.fetchone()
        cursor.close()
        conn.close()

        if not row:
            return jsonify({'success': False, 'error': 'Not found'}), 404

        return jsonify({
            'success': True,
            'data': {
                'id': row[0],
                'name': row[1],
                'hcode5': row[2],
                'hcode9': row[3],
                'hcode9_new': row[4],
                'org_type': row[5],
                'service_type': row[6],
                'hospital_level': row[7],
                'actual_beds': row[8],
                'status': row[9],
                'health_region': row[10],
                'province': row[11],
                'district': row[12],
                'address': row[13]
            }
        })

    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/settings/hospital-code', methods=['POST'])
def api_save_hospital_code():
    """Save hospital code setting"""
    try:
        data = request.get_json()
        if not data or 'hospital_code' not in data:
            return jsonify({
                'success': False,
                'error': 'hospital_code is required'
            }), 400

        hospital_code = data['hospital_code'].strip()

        # Validate format (5-digit code)
        if not hospital_code.isdigit() or len(hospital_code) != 5:
            return jsonify({
                'success': False,
                'error': 'Hospital code must be a 5-digit number'
            }), 400

        # Save to settings
        settings_manager.set_hospital_code(hospital_code)

        return jsonify({
            'success': True,
            'message': f'Hospital code {hospital_code} saved successfully',
            'hospital_code': hospital_code
        })

    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/settings/hospital-code', methods=['GET'])
def api_get_hospital_code():
    """Get current hospital code setting"""
    try:
        hospital_code = settings_manager.get_hospital_code()
        return jsonify({
            'success': True,
            'hospital_code': hospital_code
        })
    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== Master Data Count APIs ====================

@app.route('/api/master-data/health-offices/count')
@login_required
def api_master_data_health_offices_count():
    """Get count of health offices"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM health_offices")
        count = cursor.fetchone()[0]
        cursor.close()
        conn.close()

        return jsonify({'success': True, 'count': count})
    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/master-data/error-codes/count')
@login_required
def api_master_data_error_codes_count():
    """Get count of NHSO error codes"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM nhso_error_codes")
        count = cursor.fetchone()[0]
        cursor.close()
        conn.close()

        return jsonify({'success': True, 'count': count})
    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/master-data/fund-types/count')
@login_required
def api_master_data_fund_types_count():
    """Get count of fund types"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM fund_types")
        count = cursor.fetchone()[0]
        cursor.close()
        conn.close()

        return jsonify({'success': True, 'count': count})
    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/master-data/service-types/count')
@login_required
def api_master_data_service_types_count():
    """Get count of service types"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM service_types")
        count = cursor.fetchone()[0]
        cursor.close()
        conn.close()

        return jsonify({'success': True, 'count': count})
    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/master-data/dim-date/count')
@login_required
def api_master_data_dim_date_count():
    """Get count of date dimension records"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM dim_date")
        count = cursor.fetchone()[0]
        cursor.close()
        conn.close()

        return jsonify({'success': True, 'count': count})
    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== Master Data List APIs ====================

@app.route('/api/master-data/error-codes')
@login_required
def api_master_data_error_codes_list():
    """Get NHSO error codes with pagination"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        search = request.args.get('search', '').strip()
        category = request.args.get('category', '').strip()

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Build WHERE clause
        where_clauses = []
        params = []

        if search:
            like_op = "LIKE" if DB_TYPE == 'mysql' else "ILIKE"
            where_clauses.append(f"(code {like_op} %s OR type {like_op} %s OR description {like_op} %s)")
            params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])

        if category:
            where_clauses.append("type = %s")
            params.append(category)

        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

        # Get total count
        cursor.execute(f"SELECT COUNT(*) FROM nhso_error_codes WHERE {where_sql}", params)
        total = cursor.fetchone()[0]

        # Get paginated data
        offset = (page - 1) * per_page
        cursor.execute(f"""
            SELECT code, description, type, guide
            FROM nhso_error_codes
            WHERE {where_sql}
            ORDER BY code
            LIMIT %s OFFSET %s
        """, params + [per_page, offset])

        rows = cursor.fetchall()
        data = [{
            'error_code': row[0],
            'error_message': row[1] or '',
            'category': row[2],
            'description': row[3] or ''
        } for row in rows]

        cursor.close()
        conn.close()

        return jsonify({'success': True, 'data': data, 'total': total})
    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/master-data/error-codes/<code>')
@login_required
def api_master_data_error_code_by_code(code):
    """Get error code details by code"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        cursor.execute("""
            SELECT code, description, type, guide
            FROM nhso_error_codes
            WHERE code = %s
        """, [code])

        row = cursor.fetchone()
        cursor.close()
        conn.close()

        if not row:
            return jsonify({'success': False, 'error': 'Error code not found'}), 404

        data = {
            'error_code': row[0],
            'error_message': row[1] or '',
            'category': row[2] or '',
            'guide': row[3] or ''
        }

        return jsonify({'success': True, 'data': data})
    except Exception as e:
        logger.error(f"Error getting error code {code}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/master-data/fund-types')
@login_required
def api_master_data_fund_types_list():
    """Get fund types with pagination"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        search = request.args.get('search', '').strip()

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Build WHERE clause
        where_sql = "1=1"
        params = []

        if search:
            like_op = "LIKE" if DB_TYPE == 'mysql' else "ILIKE"
            where_sql = f"(fund_code {like_op} %s OR fund_name_th {like_op} %s)"
            params.extend([f'%{search}%', f'%{search}%'])

        # Get total count
        cursor.execute(f"SELECT COUNT(*) FROM fund_types WHERE {where_sql}", params)
        total = cursor.fetchone()[0]

        # Get paginated data
        offset = (page - 1) * per_page
        cursor.execute(f"""
            SELECT fund_code, fund_name_th, fund_name_en, description, is_active
            FROM fund_types
            WHERE {where_sql}
            ORDER BY fund_code
            LIMIT %s OFFSET %s
        """, params + [per_page, offset])

        rows = cursor.fetchall()
        data = [{
            'fund_code': row[0],
            'fund_name': row[1],
            'short_name': row[2] or '',
            'description': row[3] or '',
            'is_active': bool(row[4])
        } for row in rows]

        cursor.close()
        conn.close()

        return jsonify({'success': True, 'data': data, 'total': total})
    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/master-data/service-types')
@login_required
def api_master_data_service_types_list():
    """Get service types with pagination"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        search = request.args.get('search', '').strip()

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Build WHERE clause
        where_sql = "1=1"
        params = []

        if search:
            like_op = "LIKE" if DB_TYPE == 'mysql' else "ILIKE"
            where_sql = f"(service_code {like_op} %s OR service_name_th {like_op} %s)"
            params.extend([f'%{search}%', f'%{search}%'])

        # Get total count
        cursor.execute(f"SELECT COUNT(*) FROM service_types WHERE {where_sql}", params)
        total = cursor.fetchone()[0]

        # Get paginated data
        offset = (page - 1) * per_page
        cursor.execute(f"""
            SELECT service_code, service_name_th, service_name_en, description, is_active
            FROM service_types
            WHERE {where_sql}
            ORDER BY service_code
            LIMIT %s OFFSET %s
        """, params + [per_page, offset])

        rows = cursor.fetchall()
        data = [{
            'service_code': row[0],
            'service_name': row[1],
            'short_name': row[2] or '',
            'description': row[3] or '',
            'is_active': bool(row[4])
        } for row in rows]

        cursor.close()
        conn.close()

        return jsonify({'success': True, 'data': data, 'total': total})
    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/master-data/dim-date')
@login_required
def api_master_data_dim_date_list():
    """Get date dimension with pagination"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        month = request.args.get('month', '').strip()
        year_be = request.args.get('year_be', type=int)
        quarter = request.args.get('quarter', type=int)
        fiscal_year = request.args.get('fiscal_year', type=int)

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Build WHERE clause
        where_clauses = []
        params = []

        if month:
            where_clauses.append("full_date >= %s AND full_date < %s")
            # month is in format YYYY-MM
            start_date = f"{month}-01"
            # Calculate next month
            import datetime as dt
            date_obj = dt.datetime.strptime(month, '%Y-%m')
            next_month = date_obj.replace(day=28) + dt.timedelta(days=4)
            next_month = next_month.replace(day=1)
            end_date = next_month.strftime('%Y-%m-%d')
            params.extend([start_date, end_date])

        if year_be:
            # Convert BE to Gregorian year
            gregorian_year = year_be - 543
            where_clauses.append("year = %s")
            params.append(gregorian_year)

        if quarter:
            where_clauses.append("quarter = %s")
            params.append(quarter)

        if fiscal_year:
            # Convert BE to Gregorian year
            gregorian_fiscal = fiscal_year - 543
            where_clauses.append("fiscal_year = %s")
            params.append(gregorian_fiscal)

        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

        # Get total count
        cursor.execute(f"SELECT COUNT(*) FROM dim_date WHERE {where_sql}", params)
        total = cursor.fetchone()[0]

        # Get paginated data
        offset = (page - 1) * per_page
        cursor.execute(f"""
            SELECT full_date, year, fiscal_year, month_name_th, quarter,
                   week_of_year, day_name
            FROM dim_date
            WHERE {where_sql}
            ORDER BY full_date DESC
            LIMIT %s OFFSET %s
        """, params + [per_page, offset])

        rows = cursor.fetchall()
        data = []
        for row in rows:
            # Convert Gregorian date to Buddhist Era format
            full_date = row[0]
            year_be = row[1] + 543 if row[1] else None
            fiscal_year_be = row[2] + 543 if row[2] else None

            # Format date in Thai Buddhist calendar (YYYYMMDD format with BE year)
            if full_date:
                date_be_str = full_date.strftime(f'{year_be}%m%d') if year_be else ''
            else:
                date_be_str = ''

            data.append({
                'date_value': str(full_date) if full_date else '',
                'date_be': date_be_str,
                'year_be': year_be,
                'month_name_th': row[3] or '',
                'quarter': row[4],
                'fiscal_year_be': fiscal_year_be,
                'week_of_year': row[5],
                'day_name_th': row[6] or ''
            })

        cursor.close()
        conn.close()

        return jsonify({'success': True, 'data': data, 'total': total})
    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/master-data/dim-date/coverage')
@login_required
def api_master_data_dim_date_coverage():
    """Get date dimension coverage information"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Get date range
        cursor.execute("""
            SELECT MIN(full_date) as earliest, MAX(full_date) as latest, COUNT(*) as total
            FROM dim_date
        """)
        row = cursor.fetchone()

        if row and row[0] and row[1]:
            earliest = row[0]
            latest = row[1]
            total = row[2]

            # Convert to BE
            earliest_be = earliest.year + 543
            latest_be = latest.year + 543

            # Calculate coverage
            coverage_days = (latest - earliest).days + 1

            data = {
                'earliest_date': str(earliest),
                'latest_date': str(latest),
                'earliest_year_be': earliest_be,
                'latest_year_be': latest_be,
                'total_records': total,
                'coverage_days': coverage_days,
                'has_data': True
            }
        else:
            data = {
                'has_data': False,
                'total_records': 0
            }

        cursor.close()
        conn.close()

        return jsonify({'success': True, 'data': data})
    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/master-data/dim-date/generate', methods=['POST'])
@login_required
def api_master_data_dim_date_generate():
    """Generate date dimension data"""
    try:
        data = request.get_json() or {}
        target_year = data.get('target_year')

        if not target_year:
            return jsonify({'success': False, 'error': 'target_year is required'}), 400

        target_year = int(target_year)

        # Validate year
        current_year = datetime.now().year
        if target_year < current_year:
            return jsonify({'success': False, 'error': f'Target year must be >= {current_year}'}), 400
        if target_year > current_year + 20:
            return jsonify({'success': False, 'error': 'Target year must be within 20 years from now'}), 400

        # Import and run generator
        from utils.dim_date_generator import DimDateGenerator

        generator = DimDateGenerator()
        count = generator.extend_to_year(target_year)

        target_year_be = target_year + 543

        return jsonify({
            'success': True,
            'message': f'Generated {count} new records up to year {target_year} (BE {target_year_be})',
            'records_added': count,
            'target_year': target_year,
            'target_year_be': target_year_be
        })

    except Exception as e:
        logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


# =============================================================================
# Main Entry Point
# =============================================================================

if __name__ == '__main__':
    app.run(
        host='0.0.0.0',
        port=5001,
        debug=(os.getenv('FLASK_ENV') == 'development')
    )
