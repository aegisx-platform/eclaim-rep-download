"""Master Data and Health Offices API Blueprint"""

import io
import re
import time
import datetime as dt
import pandas as pd
from openpyxl import load_workbook
from flask import Blueprint, jsonify, request, current_app
from flask_login import login_required
from config.database import DB_TYPE
from utils.logging_config import safe_format_exception


# Create blueprint
master_data_api_bp = Blueprint('master_data_api', __name__)


def get_db_connection():
    """Get database connection from pool via app context"""
    from config.db_pool import get_connection as get_pooled_connection
    try:
        conn = get_pooled_connection()
        if conn is None:
            current_app.logger.error("Failed to get connection from pool")
        return conn
    except Exception as e:
        current_app.logger.error(f"Database connection error: {e}")
        return None


# =============================================================================
# Health Offices API Routes
# =============================================================================

@master_data_api_bp.route('/api/health-offices')
def api_health_offices_list():
    """Get health offices with filtering and pagination"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Get filter parameters
        search = request.args.get('search', '').strip()
        province = request.args.get('province', '')
        status = request.args.get('status', '')
        level = request.args.get('level', '')
        region = request.args.get('region', '')
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        offset = (page - 1) * per_page

        # Build query
        where_clauses = []
        params = []

        if search:
            like_op = "LIKE" if DB_TYPE == 'mysql' else "ILIKE"
            where_clauses.append(f"(name {like_op} %s OR hcode5 {like_op} %s OR hcode9 {like_op} %s)")
            params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])
        if province:
            where_clauses.append("province = %s")
            params.append(province)
        if status:
            where_clauses.append("status = %s")
            params.append(status)
        if level:
            where_clauses.append("hospital_level = %s")
            params.append(level)
        if region:
            where_clauses.append("health_region = %s")
            params.append(region)

        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

        # Get total count
        cursor.execute(f"SELECT COUNT(*) FROM health_offices WHERE {where_sql}", params)
        total = cursor.fetchone()[0]

        # Get data
        cursor.execute(f"""
            SELECT id, name, hcode5, hcode9, org_type, service_type, hospital_level,
                   actual_beds, status, health_region, province, district, address
            FROM health_offices
            WHERE {where_sql}
            ORDER BY name
            LIMIT %s OFFSET %s
        """, params + [per_page, offset])

        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        offices = []
        for row in rows:
            offices.append({
                'id': row[0],
                'name': row[1],
                'hcode5': row[2],
                'hcode9': row[3],
                'org_type': row[4],
                'service_type': row[5],
                'hospital_level': row[6],
                'actual_beds': row[7],
                'status': row[8],
                'health_region': row[9],
                'province': row[10],
                'district': row[11],
                'address': row[12]
            })

        return jsonify({
            'success': True,
            'data': offices,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': (total + per_page - 1) // per_page
        })

    except Exception as e:
        current_app.logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@master_data_api_bp.route('/api/health-offices/stats')
def api_health_offices_stats():
    """Get health offices statistics"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        stats = {}

        # Total count
        cursor.execute("SELECT COUNT(*) FROM health_offices")
        stats['total'] = cursor.fetchone()[0]

        # By status
        cursor.execute("""
            SELECT status, COUNT(*) FROM health_offices
            WHERE status IS NOT NULL
            GROUP BY status ORDER BY COUNT(*) DESC
        """)
        stats['by_status'] = [{'status': r[0], 'count': r[1]} for r in cursor.fetchall()]

        # By hospital level
        cursor.execute("""
            SELECT hospital_level, COUNT(*) FROM health_offices
            WHERE hospital_level IS NOT NULL
            GROUP BY hospital_level ORDER BY COUNT(*) DESC
        """)
        stats['by_level'] = [{'level': r[0], 'count': r[1]} for r in cursor.fetchall()]

        # By region
        cursor.execute("""
            SELECT health_region, COUNT(*) FROM health_offices
            WHERE health_region IS NOT NULL
            GROUP BY health_region ORDER BY health_region
        """)
        stats['by_region'] = [{'region': r[0], 'count': r[1]} for r in cursor.fetchall()]

        # By province (top 10)
        cursor.execute("""
            SELECT province, COUNT(*) FROM health_offices
            WHERE province IS NOT NULL
            GROUP BY province ORDER BY COUNT(*) DESC LIMIT 10
        """)
        stats['by_province_top10'] = [{'province': r[0], 'count': r[1]} for r in cursor.fetchall()]

        # Get distinct values for filters
        cursor.execute("SELECT DISTINCT province FROM health_offices WHERE province IS NOT NULL ORDER BY province")
        stats['provinces'] = [r[0] for r in cursor.fetchall()]

        cursor.execute("SELECT DISTINCT health_region FROM health_offices WHERE health_region IS NOT NULL ORDER BY health_region")
        stats['regions'] = [r[0] for r in cursor.fetchall()]

        cursor.execute("SELECT DISTINCT hospital_level FROM health_offices WHERE hospital_level IS NOT NULL ORDER BY hospital_level")
        stats['levels'] = [r[0] for r in cursor.fetchall()]

        # Last import info
        cursor.execute("""
            SELECT import_date, filename, total_records, imported, import_mode
            FROM health_offices_import_log
            ORDER BY import_date DESC LIMIT 1
        """)
        last_import = cursor.fetchone()
        if last_import:
            stats['last_import'] = {
                'date': last_import[0].strftime('%Y-%m-%d %H:%M') if last_import[0] else None,
                'filename': last_import[1],
                'total_records': last_import[2],
                'imported': last_import[3],
                'mode': last_import[4]
            }

        cursor.close()
        conn.close()

        return jsonify({'success': True, 'stats': stats})

    except Exception as e:
        current_app.logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@master_data_api_bp.route('/api/health-offices/import', methods=['POST'])
def api_health_offices_import():
    """Import health offices from uploaded Excel file"""

    def parse_formula_value(val):
        """Parse Excel formula value like ='32045' or =\"32045\" to plain value"""
        if val is None:
            return None
        val_str = str(val)
        # Match ="xxx" or ='xxx' format
        match = re.match(r'^=["\'](.*)["\']\s*$', val_str)
        if match:
            return match.group(1)
        return val_str if val_str.lower() not in ('none', 'nan', '') else None

    try:
        start_time = time.time()

        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400

        import_mode = request.form.get('mode', 'upsert')  # 'upsert' or 'replace'

        # Read Excel file using openpyxl to handle formula values
        file_bytes = io.BytesIO(file.read())
        wb = load_workbook(file_bytes, data_only=False)
        ws = wb.active

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        # Read data rows with formula parsing for code columns
        data = []
        formula_columns = {'รหัส 9 หลักใหม่', 'รหัส 9 หลัก', 'รหัส 5 หลัก', 'เลขอนุญาตให้ประกอบสถานบริการสุขภาพ 11 หลัก'}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            row_data = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers):
                    header = headers[col_idx]
                    if header in formula_columns:
                        row_data[header] = parse_formula_value(cell.value)
                    else:
                        row_data[header] = cell.value
            data.append(row_data)

        df = pd.DataFrame(data)

        # Column mapping
        column_map = {
            'ชื่อ': 'name',
            'รหัส 9 หลักใหม่': 'hcode9_new',
            'รหัส 9 หลัก': 'hcode9',
            'รหัส 5 หลัก': 'hcode5',
            'เลขอนุญาตให้ประกอบสถานบริการสุขภาพ 11 หลัก': 'license_no',
            'ประเภทองค์กร': 'org_type',
            'ประเภทหน่วยบริการสุขภาพ': 'service_type',
            'สังกัด': 'affiliation',
            'แผนก/กรม': 'department',
            'ระดับโรงพยาบาล': 'hospital_level',
            'เตียงที่ใช้จริง': 'actual_beds',
            'สถานะการใช้งาน': 'status',
            'เขตบริการ': 'health_region',
            'ที่อยู่': 'address',
            'รหัสจังหวัด': 'province_code',
            'จังหวัด': 'province',
            'รหัสอำเภอ': 'district_code',
            'อำเภอ/เขต': 'district',
            'รหัสตำบล': 'subdistrict_code',
            'ตำบล/แขวง': 'subdistrict',
            'หมู่': 'moo',
            'รหัสไปรษณีย์': 'postal_code',
            'แม่ข่าย': 'parent_code',
            'วันที่ก่อตั้ง': 'established_date',
            'วันที่ปิดบริการ': 'closed_date',
            'อัพเดตล่าสุด(เริ่ม 05/09/2566)': 'source_updated_at'
        }

        # Rename columns
        df = df.rename(columns=column_map)

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Clear existing data if replace mode
        if import_mode == 'replace':
            cursor.execute("TRUNCATE TABLE health_offices RESTART IDENTITY")

        imported = 0
        updated = 0
        skipped = 0
        errors = 0
        total_records = len(df)

        for idx, row in df.iterrows():
            try:
                # Convert codes to string (handle both formula-parsed strings and numbers)
                def clean_code(val):
                    if pd.isna(val) or val is None or str(val).lower() in ('none', 'nan', ''):
                        return None
                    # Remove any whitespace and convert to string
                    return str(val).strip()

                hcode5 = clean_code(row.get('hcode5'))
                hcode9 = clean_code(row.get('hcode9'))
                hcode9_new = clean_code(row.get('hcode9_new'))

                # Parse dates
                def parse_date(val):
                    if pd.isna(val):
                        return None
                    if isinstance(val, str):
                        try:
                            # Try dd/mm/yyyy format
                            parts = val.split('/')
                            if len(parts) == 3:
                                day, month, year = parts
                                if int(year) > 2500:
                                    year = str(int(year) - 543)
                                return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                        except:
                            pass
                    return None

                established = parse_date(row.get('established_date'))
                closed = parse_date(row.get('closed_date'))
                source_updated = parse_date(row.get('source_updated_at'))

                # Prepare values
                values = {
                    'name': row.get('name'),
                    'hcode9_new': hcode9_new,
                    'hcode9': hcode9,
                    'hcode5': hcode5,
                    'license_no': clean_code(row.get('license_no')),
                    'org_type': row.get('org_type'),
                    'service_type': row.get('service_type'),
                    'affiliation': row.get('affiliation'),
                    'department': row.get('department'),
                    'hospital_level': row.get('hospital_level'),
                    'actual_beds': int(row['actual_beds']) if pd.notna(row.get('actual_beds')) else 0,
                    'status': row.get('status'),
                    'health_region': row.get('health_region'),
                    'address': row.get('address'),
                    'province_code': str(int(row['province_code'])) if pd.notna(row.get('province_code')) else None,
                    'province': row.get('province'),
                    'district_code': str(int(row['district_code'])) if pd.notna(row.get('district_code')) else None,
                    'district': row.get('district'),
                    'subdistrict_code': str(int(row['subdistrict_code'])) if pd.notna(row.get('subdistrict_code')) else None,
                    'subdistrict': row.get('subdistrict'),
                    'moo': row.get('moo'),
                    'postal_code': row.get('postal_code'),
                    'parent_code': row.get('parent_code'),
                    'established_date': established,
                    'closed_date': closed,
                    'source_updated_at': source_updated
                }

                # Skip if no name
                if not values['name'] or pd.isna(values['name']):
                    skipped += 1
                    continue

                # Insert or update
                if import_mode == 'replace' or not hcode5:
                    # Insert only
                    cursor.execute("""
                        INSERT INTO health_offices (
                            name, hcode9_new, hcode9, hcode5, license_no, org_type, service_type,
                            affiliation, department, hospital_level, actual_beds, status, health_region,
                            address, province_code, province, district_code, district, subdistrict_code,
                            subdistrict, moo, postal_code, parent_code, established_date, closed_date,
                            source_updated_at
                        ) VALUES (
                            %(name)s, %(hcode9_new)s, %(hcode9)s, %(hcode5)s, %(license_no)s, %(org_type)s,
                            %(service_type)s, %(affiliation)s, %(department)s, %(hospital_level)s,
                            %(actual_beds)s, %(status)s, %(health_region)s, %(address)s, %(province_code)s,
                            %(province)s, %(district_code)s, %(district)s, %(subdistrict_code)s,
                            %(subdistrict)s, %(moo)s, %(postal_code)s, %(parent_code)s, %(established_date)s,
                            %(closed_date)s, %(source_updated_at)s
                        )
                    """, values)
                    imported += 1
                else:
                    # Upsert by hcode5 - use appropriate syntax based on database type
                    if DB_TYPE == 'mysql':
                        cursor.execute("""
                            INSERT INTO health_offices (
                                name, hcode9_new, hcode9, hcode5, license_no, org_type, service_type,
                                affiliation, department, hospital_level, actual_beds, status, health_region,
                                address, province_code, province, district_code, district, subdistrict_code,
                                subdistrict, moo, postal_code, parent_code, established_date, closed_date,
                                source_updated_at
                            ) VALUES (
                                %(name)s, %(hcode9_new)s, %(hcode9)s, %(hcode5)s, %(license_no)s, %(org_type)s,
                                %(service_type)s, %(affiliation)s, %(department)s, %(hospital_level)s,
                                %(actual_beds)s, %(status)s, %(health_region)s, %(address)s, %(province_code)s,
                                %(province)s, %(district_code)s, %(district)s, %(subdistrict_code)s,
                                %(subdistrict)s, %(moo)s, %(postal_code)s, %(parent_code)s, %(established_date)s,
                                %(closed_date)s, %(source_updated_at)s
                            )
                            ON DUPLICATE KEY UPDATE
                                name = VALUES(name),
                                hcode9_new = VALUES(hcode9_new),
                                hcode9 = VALUES(hcode9),
                                license_no = VALUES(license_no),
                                org_type = VALUES(org_type),
                                service_type = VALUES(service_type),
                                affiliation = VALUES(affiliation),
                                department = VALUES(department),
                                hospital_level = VALUES(hospital_level),
                                actual_beds = VALUES(actual_beds),
                                status = VALUES(status),
                                health_region = VALUES(health_region),
                                address = VALUES(address),
                                province_code = VALUES(province_code),
                                province = VALUES(province),
                                district_code = VALUES(district_code),
                                district = VALUES(district),
                                subdistrict_code = VALUES(subdistrict_code),
                                subdistrict = VALUES(subdistrict),
                                moo = VALUES(moo),
                                postal_code = VALUES(postal_code),
                                parent_code = VALUES(parent_code),
                                established_date = VALUES(established_date),
                                closed_date = VALUES(closed_date),
                                source_updated_at = VALUES(source_updated_at),
                                updated_at = CURRENT_TIMESTAMP
                        """, values)
                    else:
                        cursor.execute("""
                            INSERT INTO health_offices (
                                name, hcode9_new, hcode9, hcode5, license_no, org_type, service_type,
                                affiliation, department, hospital_level, actual_beds, status, health_region,
                                address, province_code, province, district_code, district, subdistrict_code,
                                subdistrict, moo, postal_code, parent_code, established_date, closed_date,
                                source_updated_at
                            ) VALUES (
                                %(name)s, %(hcode9_new)s, %(hcode9)s, %(hcode5)s, %(license_no)s, %(org_type)s,
                                %(service_type)s, %(affiliation)s, %(department)s, %(hospital_level)s,
                                %(actual_beds)s, %(status)s, %(health_region)s, %(address)s, %(province_code)s,
                                %(province)s, %(district_code)s, %(district)s, %(subdistrict_code)s,
                                %(subdistrict)s, %(moo)s, %(postal_code)s, %(parent_code)s, %(established_date)s,
                                %(closed_date)s, %(source_updated_at)s
                            )
                            ON CONFLICT (hcode5) DO UPDATE SET
                                name = EXCLUDED.name,
                                hcode9_new = EXCLUDED.hcode9_new,
                                hcode9 = EXCLUDED.hcode9,
                                license_no = EXCLUDED.license_no,
                                org_type = EXCLUDED.org_type,
                                service_type = EXCLUDED.service_type,
                                affiliation = EXCLUDED.affiliation,
                                department = EXCLUDED.department,
                                hospital_level = EXCLUDED.hospital_level,
                                actual_beds = EXCLUDED.actual_beds,
                                status = EXCLUDED.status,
                                health_region = EXCLUDED.health_region,
                                address = EXCLUDED.address,
                                province_code = EXCLUDED.province_code,
                                province = EXCLUDED.province,
                                district_code = EXCLUDED.district_code,
                                district = EXCLUDED.district,
                                subdistrict_code = EXCLUDED.subdistrict_code,
                                subdistrict = EXCLUDED.subdistrict,
                                moo = EXCLUDED.moo,
                                postal_code = EXCLUDED.postal_code,
                                parent_code = EXCLUDED.parent_code,
                                established_date = EXCLUDED.established_date,
                                closed_date = EXCLUDED.closed_date,
                                source_updated_at = EXCLUDED.source_updated_at,
                                updated_at = CURRENT_TIMESTAMP
                        """, values)
                    if cursor.rowcount > 0:
                        imported += 1

            except Exception as e:
                errors += 1
                if errors <= 5:
                    current_app.logger.error(f"Error importing row {idx}: {e}")

        # Log import
        duration = time.time() - start_time
        cursor.execute("""
            INSERT INTO health_offices_import_log
            (filename, total_records, imported, updated, skipped, errors, import_mode, duration_seconds)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (file.filename, total_records, imported, updated, skipped, errors, import_mode, duration))

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'message': f'Import completed',
            'total': total_records,
            'imported': imported,
            'updated': updated,
            'skipped': skipped,
            'errors': errors,
            'duration': round(duration, 2)
        })

    except Exception as e:
        current_app.logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@master_data_api_bp.route('/api/health-offices/clear', methods=['POST'])
def api_health_offices_clear():
    """Clear all health offices data"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()
        cursor.execute("TRUNCATE TABLE health_offices RESTART IDENTITY")
        conn.commit()

        cursor.execute("SELECT COUNT(*) FROM health_offices")
        count = cursor.fetchone()[0]

        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'message': 'All health offices data cleared',
            'remaining': count
        })

    except Exception as e:
        current_app.logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@master_data_api_bp.route('/api/health-offices/lookup/<code>')
def api_health_offices_lookup(code):
    """Lookup health office by code (hcode5 or hcode9)"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, name, hcode5, hcode9, hcode9_new, org_type, service_type,
                   hospital_level, actual_beds, status, health_region, province,
                   district, address
            FROM health_offices
            WHERE hcode5 = %s OR hcode9 = %s OR hcode5 = %s
            LIMIT 1
        """, (code, code, code.zfill(5)))

        row = cursor.fetchone()
        cursor.close()
        conn.close()

        if not row:
            return jsonify({'success': False, 'error': 'Not found'}), 404

        return jsonify({
            'success': True,
            'data': {
                'id': row[0],
                'name': row[1],
                'hcode5': row[2],
                'hcode9': row[3],
                'hcode9_new': row[4],
                'org_type': row[5],
                'service_type': row[6],
                'hospital_level': row[7],
                'actual_beds': row[8],
                'status': row[9],
                'health_region': row[10],
                'province': row[11],
                'district': row[12],
                'address': row[13]
            }
        })

    except Exception as e:
        current_app.logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


# =============================================================================
# Master Data Count APIs
# =============================================================================

@master_data_api_bp.route('/api/master-data/health-offices/count')
@login_required
def api_master_data_health_offices_count():
    """Get count of health offices"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM health_offices")
        count = cursor.fetchone()[0]
        cursor.close()
        conn.close()

        return jsonify({'success': True, 'count': count})
    except Exception as e:
        current_app.logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@master_data_api_bp.route('/api/master-data/error-codes/count')
@login_required
def api_master_data_error_codes_count():
    """Get count of NHSO error codes"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM nhso_error_codes")
        count = cursor.fetchone()[0]
        cursor.close()
        conn.close()

        return jsonify({'success': True, 'count': count})
    except Exception as e:
        current_app.logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@master_data_api_bp.route('/api/master-data/fund-types/count')
@login_required
def api_master_data_fund_types_count():
    """Get count of fund types"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM fund_types")
        count = cursor.fetchone()[0]
        cursor.close()
        conn.close()

        return jsonify({'success': True, 'count': count})
    except Exception as e:
        current_app.logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@master_data_api_bp.route('/api/master-data/service-types/count')
@login_required
def api_master_data_service_types_count():
    """Get count of service types"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM service_types")
        count = cursor.fetchone()[0]
        cursor.close()
        conn.close()

        return jsonify({'success': True, 'count': count})
    except Exception as e:
        current_app.logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@master_data_api_bp.route('/api/master-data/dim-date/count')
@login_required
def api_master_data_dim_date_count():
    """Get count of date dimension records"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM dim_date")
        count = cursor.fetchone()[0]
        cursor.close()
        conn.close()

        return jsonify({'success': True, 'count': count})
    except Exception as e:
        current_app.logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


# =============================================================================
# Master Data List APIs
# =============================================================================

@master_data_api_bp.route('/api/master-data/error-codes')
@login_required
def api_master_data_error_codes_list():
    """Get NHSO error codes with pagination"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        search = request.args.get('search', '').strip()
        category = request.args.get('category', '').strip()

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Build WHERE clause
        where_clauses = []
        params = []

        if search:
            like_op = "LIKE" if DB_TYPE == 'mysql' else "ILIKE"
            where_clauses.append(f"(code {like_op} %s OR type {like_op} %s OR description {like_op} %s)")
            params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])

        if category:
            where_clauses.append("type = %s")
            params.append(category)

        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

        # Get total count
        cursor.execute(f"SELECT COUNT(*) FROM nhso_error_codes WHERE {where_sql}", params)
        total = cursor.fetchone()[0]

        # Get paginated data
        offset = (page - 1) * per_page
        cursor.execute(f"""
            SELECT code, description, type, guide
            FROM nhso_error_codes
            WHERE {where_sql}
            ORDER BY code
            LIMIT %s OFFSET %s
        """, params + [per_page, offset])

        rows = cursor.fetchall()
        data = [{
            'error_code': row[0],
            'error_message': row[1] or '',
            'category': row[2],
            'description': row[3] or ''
        } for row in rows]

        cursor.close()
        conn.close()

        return jsonify({'success': True, 'data': data, 'total': total})
    except Exception as e:
        current_app.logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@master_data_api_bp.route('/api/master-data/error-codes/<code>')
@login_required
def api_master_data_error_code_by_code(code):
    """Get error code details by code"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        cursor.execute("""
            SELECT code, description, type, guide
            FROM nhso_error_codes
            WHERE code = %s
        """, [code])

        row = cursor.fetchone()
        cursor.close()
        conn.close()

        if not row:
            return jsonify({'success': False, 'error': 'Error code not found'}), 404

        data = {
            'error_code': row[0],
            'error_message': row[1] or '',
            'category': row[2] or '',
            'guide': row[3] or ''
        }

        return jsonify({'success': True, 'data': data})
    except Exception as e:
        current_app.logger.error(f"Error getting error code {code}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@master_data_api_bp.route('/api/master-data/fund-types')
@login_required
def api_master_data_fund_types_list():
    """Get fund types with pagination"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        search = request.args.get('search', '').strip()

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Build WHERE clause
        where_sql = "1=1"
        params = []

        if search:
            like_op = "LIKE" if DB_TYPE == 'mysql' else "ILIKE"
            where_sql = f"(fund_code {like_op} %s OR fund_name_th {like_op} %s)"
            params.extend([f'%{search}%', f'%{search}%'])

        # Get total count
        cursor.execute(f"SELECT COUNT(*) FROM fund_types WHERE {where_sql}", params)
        total = cursor.fetchone()[0]

        # Get paginated data
        offset = (page - 1) * per_page
        cursor.execute(f"""
            SELECT fund_code, fund_name_th, fund_name_en, description, is_active
            FROM fund_types
            WHERE {where_sql}
            ORDER BY fund_code
            LIMIT %s OFFSET %s
        """, params + [per_page, offset])

        rows = cursor.fetchall()
        data = [{
            'fund_code': row[0],
            'fund_name': row[1],
            'short_name': row[2] or '',
            'description': row[3] or '',
            'is_active': bool(row[4])
        } for row in rows]

        cursor.close()
        conn.close()

        return jsonify({'success': True, 'data': data, 'total': total})
    except Exception as e:
        current_app.logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@master_data_api_bp.route('/api/master-data/service-types')
@login_required
def api_master_data_service_types_list():
    """Get service types with pagination"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        search = request.args.get('search', '').strip()

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Build WHERE clause
        where_sql = "1=1"
        params = []

        if search:
            like_op = "LIKE" if DB_TYPE == 'mysql' else "ILIKE"
            where_sql = f"(service_code {like_op} %s OR service_name_th {like_op} %s)"
            params.extend([f'%{search}%', f'%{search}%'])

        # Get total count
        cursor.execute(f"SELECT COUNT(*) FROM service_types WHERE {where_sql}", params)
        total = cursor.fetchone()[0]

        # Get paginated data
        offset = (page - 1) * per_page
        cursor.execute(f"""
            SELECT service_code, service_name_th, service_name_en, description, is_active
            FROM service_types
            WHERE {where_sql}
            ORDER BY service_code
            LIMIT %s OFFSET %s
        """, params + [per_page, offset])

        rows = cursor.fetchall()
        data = [{
            'service_code': row[0],
            'service_name': row[1],
            'short_name': row[2] or '',
            'description': row[3] or '',
            'is_active': bool(row[4])
        } for row in rows]

        cursor.close()
        conn.close()

        return jsonify({'success': True, 'data': data, 'total': total})
    except Exception as e:
        current_app.logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@master_data_api_bp.route('/api/master-data/dim-date')
@login_required
def api_master_data_dim_date_list():
    """Get date dimension with pagination"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        month = request.args.get('month', '').strip()
        year_be = request.args.get('year_be', type=int)
        quarter = request.args.get('quarter', type=int)
        fiscal_year = request.args.get('fiscal_year', type=int)

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Build WHERE clause
        where_clauses = []
        params = []

        if month:
            where_clauses.append("full_date >= %s AND full_date < %s")
            # month is in format YYYY-MM
            start_date = f"{month}-01"
            # Calculate next month
            date_obj = dt.datetime.strptime(month, '%Y-%m')
            next_month = date_obj.replace(day=28) + dt.timedelta(days=4)
            next_month = next_month.replace(day=1)
            end_date = next_month.strftime('%Y-%m-%d')
            params.extend([start_date, end_date])

        if year_be:
            # Convert BE to Gregorian year
            gregorian_year = year_be - 543
            where_clauses.append("year = %s")
            params.append(gregorian_year)

        if quarter:
            where_clauses.append("quarter = %s")
            params.append(quarter)

        if fiscal_year:
            # Convert BE to Gregorian year
            gregorian_fiscal = fiscal_year - 543
            where_clauses.append("fiscal_year = %s")
            params.append(gregorian_fiscal)

        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

        # Get total count
        cursor.execute(f"SELECT COUNT(*) FROM dim_date WHERE {where_sql}", params)
        total = cursor.fetchone()[0]

        # Get paginated data
        offset = (page - 1) * per_page
        cursor.execute(f"""
            SELECT full_date, year, fiscal_year, month_name_th, quarter,
                   week_of_year, day_name
            FROM dim_date
            WHERE {where_sql}
            ORDER BY full_date DESC
            LIMIT %s OFFSET %s
        """, params + [per_page, offset])

        rows = cursor.fetchall()
        data = []
        for row in rows:
            # Convert Gregorian date to Buddhist Era format
            full_date = row[0]
            year_be = row[1] + 543 if row[1] else None
            fiscal_year_be = row[2] + 543 if row[2] else None

            # Format date in Thai Buddhist calendar (YYYYMMDD format with BE year)
            if full_date:
                date_be_str = full_date.strftime(f'{year_be}%m%d') if year_be else ''
            else:
                date_be_str = ''

            data.append({
                'date_value': str(full_date) if full_date else '',
                'date_be': date_be_str,
                'year_be': year_be,
                'month_name_th': row[3] or '',
                'quarter': row[4],
                'fiscal_year_be': fiscal_year_be,
                'week_of_year': row[5],
                'day_name_th': row[6] or ''
            })

        cursor.close()
        conn.close()

        return jsonify({'success': True, 'data': data, 'total': total})
    except Exception as e:
        current_app.logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@master_data_api_bp.route('/api/master-data/dim-date/coverage')
@login_required
def api_master_data_dim_date_coverage():
    """Get date dimension coverage information"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor()

        # Get date range
        cursor.execute("""
            SELECT MIN(full_date) as earliest, MAX(full_date) as latest, COUNT(*) as total
            FROM dim_date
        """)
        row = cursor.fetchone()

        if row and row[0] and row[1]:
            earliest = row[0]
            latest = row[1]
            total = row[2]

            # Convert to BE
            earliest_be = earliest.year + 543
            latest_be = latest.year + 543

            # Calculate coverage
            coverage_days = (latest - earliest).days + 1

            data = {
                'earliest_date': str(earliest),
                'latest_date': str(latest),
                'earliest_year_be': earliest_be,
                'latest_year_be': latest_be,
                'total_records': total,
                'coverage_days': coverage_days,
                'has_data': True
            }
        else:
            data = {
                'has_data': False,
                'total_records': 0
            }

        cursor.close()
        conn.close()

        return jsonify({'success': True, 'data': data})
    except Exception as e:
        current_app.logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500


@master_data_api_bp.route('/api/master-data/dim-date/generate', methods=['POST'])
@login_required
def api_master_data_dim_date_generate():
    """Generate date dimension data"""
    from datetime import datetime

    try:
        data = request.get_json() or {}
        target_year = data.get('target_year')

        if not target_year:
            return jsonify({'success': False, 'error': 'target_year is required'}), 400

        target_year = int(target_year)

        # Validate year
        current_year = datetime.now().year
        if target_year < current_year:
            return jsonify({'success': False, 'error': f'Target year must be >= {current_year}'}), 400
        if target_year > current_year + 20:
            return jsonify({'success': False, 'error': 'Target year must be within 20 years from now'}), 400

        # Import and run generator
        from utils.dim_date_generator import DimDateGenerator

        generator = DimDateGenerator()
        count = generator.extend_to_year(target_year)

        target_year_be = target_year + 543

        return jsonify({
            'success': True,
            'message': f'Generated {count} new records up to year {target_year} (BE {target_year_be})',
            'records_added': count,
            'target_year': target_year,
            'target_year_be': target_year_be
        })

    except Exception as e:
        current_app.logger.error(safe_format_exception())
        return jsonify({'success': False, 'error': str(e)}), 500
