#!/usr/bin/env python3
"""
Health Offices Seed Data Importer
Imports health offices from Excel file into the database.

Usage:
    python database/seeds/health_offices_importer.py [path_to_excel]

If no path provided, looks for:
    1. database/seeds/data/health_office.xlsx
    2. /app/database/seeds/data/health_office.xlsx (Docker)
"""

import os
import sys
from pathlib import Path
from datetime import datetime

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from config.database import get_db_config, DB_TYPE


# NHSO Hospital Codes mapping: (hospital_name, province) -> hcode5
# Source: https://hcode.moph.go.th/
NHSO_HOSPITAL_CODES = {
    # เขตบริการสุขภาพที่ 8
    ('โรงพยาบาลอุดรธานี', 'อุดรธานี'): '10670',
    ('โรงพยาบาลกุมภวาปี', 'อุดรธานี'): '10671',
    ('โรงพยาบาลสมเด็จพระยุพราชบ้านดุง', 'อุดรธานี'): '10672',
    ('โรงพยาบาลหนองคาย', 'หนองคาย'): '10676',
    ('โรงพยาบาลสกลนคร', 'สกลนคร'): '10681',
    ('โรงพยาบาลนครพนม', 'นครพนม'): '10686',
    ('โรงพยาบาลเลย', 'เลย'): '10692',
    ('โรงพยาบาลหนองบัวลำภู', 'หนองบัวลำภู'): '10697',
    ('โรงพยาบาลบึงกาฬ', 'บึงกาฬ'): '14861',
    # เขตบริการสุขภาพที่ 7
    ('โรงพยาบาลขอนแก่น', 'ขอนแก่น'): '10699',
    ('โรงพยาบาลมหาสารคาม', 'มหาสารคาม'): '10702',
    ('โรงพยาบาลร้อยเอ็ด', 'ร้อยเอ็ด'): '10707',
    ('โรงพยาบาลกาฬสินธุ์', 'กาฬสินธุ์'): '10713',
    # เขตบริการสุขภาพที่ 9
    ('โรงพยาบาลมหาราชนครราชสีมา', 'นครราชสีมา'): '10720',
    ('โรงพยาบาลชัยภูมิ', 'ชัยภูมิ'): '10726',
    ('โรงพยาบาลบุรีรัมย์', 'บุรีรัมย์'): '10732',
    ('โรงพยาบาลสุรินทร์', 'สุรินทร์'): '10738',
    # เขตบริการสุขภาพที่ 10
    ('โรงพยาบาลสรรพสิทธิประสงค์', 'อุบลราชธานี'): '10744',
    ('โรงพยาบาลศรีสะเกษ', 'ศรีสะเกษ'): '10750',
    ('โรงพยาบาลยโสธร', 'ยโสธร'): '10756',
    ('โรงพยาบาลอำนาจเจริญ', 'อำนาจเจริญ'): '10762',
    ('โรงพยาบาลมุกดาหาร', 'มุกดาหาร'): '10768',
    # เขตบริการสุขภาพที่ 1
    ('โรงพยาบาลนครพิงค์', 'เชียงใหม่'): '11070',
    ('โรงพยาบาลลำปาง', 'ลำปาง'): '10774',
    ('โรงพยาบาลลำพูน', 'ลำพูน'): '10780',
    ('โรงพยาบาลเชียงรายประชานุเคราะห์', 'เชียงราย'): '10786',
    ('โรงพยาบาลพะเยา', 'พะเยา'): '10792',
    ('โรงพยาบาลแม่ฮ่องสอน', 'แม่ฮ่องสอน'): '10798',
    ('โรงพยาบาลน่าน', 'น่าน'): '10804',
    ('โรงพยาบาลแพร่', 'แพร่'): '10810',
    # เขตบริการสุขภาพที่ 2
    ('โรงพยาบาลพุทธชินราช', 'พิษณุโลก'): '10816',
    ('โรงพยาบาลสุโขทัย', 'สุโขทัย'): '10822',
    ('โรงพยาบาลอุตรดิตถ์', 'อุตรดิตถ์'): '10828',
    ('โรงพยาบาลตาก', 'ตาก'): '10834',
    ('โรงพยาบาลเพชรบูรณ์', 'เพชรบูรณ์'): '10840',
    # เขตบริการสุขภาพที่ 3
    ('โรงพยาบาลสวรรค์ประชารักษ์', 'นครสวรรค์'): '10846',
    ('โรงพยาบาลสวรรค์ประชารักษ์ แห่งใหม่', 'นครสวรรค์'): '10846',
    ('โรงพยาบาลกำแพงเพชร', 'กำแพงเพชร'): '10852',
    ('โรงพยาบาลพิจิตร', 'พิจิตร'): '10858',
    ('โรงพยาบาลชัยนาท', 'ชัยนาท'): '10864',
    ('โรงพยาบาลชัยนาทนเรนทร', 'ชัยนาท'): '10864',
    ('โรงพยาบาลอุทัยธานี', 'อุทัยธานี'): '10870',
    # เขตบริการสุขภาพที่ 4
    ('โรงพยาบาลสระบุรี', 'สระบุรี'): '10876',
    ('โรงพยาบาลพระนารายณ์มหาราช', 'ลพบุรี'): '10882',
    ('โรงพยาบาลสิงห์บุรี', 'สิงห์บุรี'): '10888',
    ('โรงพยาบาลอ่างทอง', 'อ่างทอง'): '10894',
    ('โรงพยาบาลพระนครศรีอยุธยา', 'พระนครศรีอยุธยา'): '10900',
    ('โรงพยาบาลปทุมธานี', 'ปทุมธานี'): '10906',
    ('โรงพยาบาลนนทบุรี', 'นนทบุรี'): '10912',
    # เขตบริการสุขภาพที่ 5
    ('โรงพยาบาลราชบุรี', 'ราชบุรี'): '10918',
    ('โรงพยาบาลเพชรบุรี', 'เพชรบุรี'): '10924',
    ('โรงพยาบาลสมุทรสงคราม', 'สมุทรสงคราม'): '10930',
    ('โรงพยาบาลประจวบคีรีขันธ์', 'ประจวบคีรีขันธ์'): '10936',
    ('โรงพยาบาลกาญจนบุรี', 'กาญจนบุรี'): '10942',
    ('โรงพยาบาลสุพรรณบุรี', 'สุพรรณบุรี'): '10948',
    ('โรงพยาบาลนครปฐม', 'นครปฐม'): '10954',
    ('โรงพยาบาลสมุทรสาคร', 'สมุทรสาคร'): '10960',
    # เขตบริการสุขภาพที่ 6
    ('โรงพยาบาลชลบุรี', 'ชลบุรี'): '10966',
    ('โรงพยาบาลระยอง', 'ระยอง'): '10972',
    ('โรงพยาบาลพระปกเกล้า', 'จันทบุรี'): '10978',
    ('โรงพยาบาลตราด', 'ตราด'): '10984',
    ('โรงพยาบาลฉะเชิงเทรา', 'ฉะเชิงเทรา'): '10996',
    ('โรงพยาบาลปราจีนบุรี', 'ปราจีนบุรี'): '11002',
    ('โรงพยาบาลสมุทรปราการ', 'สมุทรปราการ'): '11008',
    ('โรงพยาบาลสระแก้ว', 'สระแก้ว'): '11014',
    # เขตบริการสุขภาพที่ 11
    ('โรงพยาบาลสุราษฎร์ธานี', 'สุราษฎร์ธานี'): '11020',
    ('โรงพยาบาลมหาราชนครศรีธรรมราช', 'นครศรีธรรมราช'): '11026',
    ('โรงพยาบาลชุมพร', 'ชุมพร'): '11032',
    ('โรงพยาบาลชุมพรเขตอุดมศักดิ์', 'ชุมพร'): '11032',
    ('โรงพยาบาลระนอง', 'ระนอง'): '11038',
    ('โรงพยาบาลกระบี่', 'กระบี่'): '11044',
    ('โรงพยาบาลพังงา', 'พังงา'): '11050',
    ('โรงพยาบาลวชิระภูเก็ต', 'ภูเก็ต'): '11056',
    # เขตบริการสุขภาพที่ 12
    ('โรงพยาบาลหาดใหญ่', 'สงขลา'): '11062',
    ('โรงพยาบาลตรัง', 'ตรัง'): '11068',
    ('โรงพยาบาลสตูล', 'สตูล'): '11074',
    ('โรงพยาบาลปัตตานี', 'ปัตตานี'): '11080',
    ('โรงพยาบาลยะลา', 'ยะลา'): '11086',
    ('โรงพยาบาลนราธิวาสราชนครินทร์', 'นราธิวาส'): '11092',
    ('โรงพยาบาลพัทลุง', 'พัทลุง'): '11098',
    # กรุงเทพมหานคร
    ('โรงพยาบาลศิริราช', 'กรุงเทพมหานคร'): '13756',
    ('โรงพยาบาลจุฬาลงกรณ์', 'กรุงเทพมหานคร'): '13757',
    ('โรงพยาบาลจุฬาลงกรณ์สภากาชาดไทย', 'กรุงเทพมหานคร'): '13757',
    ('โรงพยาบาลรามาธิบดี', 'กรุงเทพมหานคร'): '13758',
    ('โรงพยาบาลรามาธิบดี  มหาวิทยาลัยมหิดล', 'กรุงเทพมหานคร'): '13758',
    ('โรงพยาบาลราชวิถี', 'กรุงเทพมหานคร'): '13759',
    ('โรงพยาบาลพระมงกุฎเกล้า', 'กรุงเทพมหานคร'): '13760',
    ('สถาบันบำราศนราดูร', 'นนทบุรี'): '13761',
    ('โรงพยาบาลธรรมศาสตร์เฉลิมพระเกียรติ', 'ปทุมธานี'): '14060',
    ('โรงพยาบาลสงขลานครินทร์', 'สงขลา'): '28008',
}


# Column mapping: Excel column -> Database column
COLUMN_MAP = {
    'ชื่อ': 'name',
    'รหัส 9 หลักใหม่': 'hcode9_new',
    'รหัส 9 หลัก': 'hcode9',
    'รหัส 5 หลัก': 'hcode5',
    'เลขอนุญาตให้ประกอบสถานบริการสุขภาพ 11 หลัก': 'license_no',
    'ประเภทองค์กร': 'org_type',
    'ประเภทหน่วยบริการสุขภาพ': 'service_type',
    'สังกัด': 'affiliation',
    'แผนก/กรม': 'department',
    'ระดับโรงพยาบาล': 'hospital_level',
    'เตียงที่ใช้จริง': 'actual_beds',
    'สถานะการใช้งาน': 'status',
    'เขตบริการ': 'health_region',
    'ที่อยู่': 'address',
    'รหัสจังหวัด': 'province_code',
    'จังหวัด': 'province',
    'รหัสอำเภอ': 'district_code',
    'อำเภอ/เขต': 'district',
    'รหัสตำบล': 'subdistrict_code',
    'ตำบล/แขวง': 'subdistrict',
    'หมู่': 'moo',
    'รหัสไปรษณีย์': 'postal_code',
    'แม่ข่าย': 'parent_code',
    'วันที่ก่อตั้ง': 'established_date',
    'วันที่ปิดบริการ': 'closed_date',
    'อัพเดตล่าสุด(เริ่ม 05/09/2566)': 'source_updated_at',
}


def parse_excel_formula_value(value):
    """
    Parse Excel formula value like '="32045"' to extract actual value.

    Examples:
        '="32045"' -> '32045'
        '="CA0032045"' -> 'CA0032045'
        'normal' -> 'normal'
        None -> ''
    """
    if value is None:
        return ''

    val_str = str(value).strip()

    # Check for formula format: ="value" or ='value'
    if val_str.startswith('="') and val_str.endswith('"'):
        return val_str[2:-1]  # Remove =" and "
    if val_str.startswith("='") and val_str.endswith("'"):
        return val_str[2:-1]  # Remove =' and '

    return val_str


def import_health_offices(excel_path: str, batch_size: int = 1000) -> dict:
    """
    Import health offices from Excel file using openpyxl.

    Uses openpyxl instead of pandas to correctly read Excel formula values
    like '="32045"' which pandas cannot parse.

    Returns:
        dict with import statistics
    """
    from openpyxl import load_workbook
    import hashlib

    print(f"[seed] Reading Excel file with openpyxl: {excel_path}")
    wb = load_workbook(excel_path, read_only=True, data_only=False)
    ws = wb.active

    # Get headers from first row
    headers = []
    for cell in list(ws.rows)[0]:
        headers.append(cell.value)

    print(f"[seed] Found {len(headers)} columns")

    # Map Excel column index to database column name
    col_indices = {}
    for excel_col, db_col in COLUMN_MAP.items():
        if excel_col in headers:
            col_indices[headers.index(excel_col)] = db_col

    print(f"[seed] Mapped {len(col_indices)} columns")

    # Read all data rows
    records = []
    row_count = 0

    for row in ws.iter_rows(min_row=2):  # Skip header
        row_count += 1
        record = {}

        for col_idx, db_col in col_indices.items():
            cell = row[col_idx]
            value = cell.value if cell else None

            # Parse formula values like ="32045"
            if db_col in ['hcode5', 'hcode9', 'hcode9_new', 'postal_code', 'moo']:
                value = parse_excel_formula_value(value)
            elif value is None:
                value = ''
            else:
                value = str(value).strip() if not isinstance(value, (int, float)) else value

            record[db_col] = value

        records.append(record)

        if row_count % 10000 == 0:
            print(f"[seed] Read {row_count} rows...")

    wb.close()

    total = len(records)
    print(f"[seed] Found {total} records")

    # Show sample hcode5 values to verify parsing
    sample_codes = [r.get('hcode5', '') for r in records[:5] if r.get('hcode5')]
    print(f"[seed] Sample hcode5 values: {sample_codes}")

    # Clean and convert data
    valid_cols = list(set(col_indices.values()))

    for record in records:
        # Convert numeric columns
        for col in ['actual_beds', 'province_code', 'district_code', 'subdistrict_code']:
            if col in record:
                try:
                    val = record[col]
                    if val and val != '':
                        record[col] = int(float(str(val).replace(',', '')))
                    else:
                        record[col] = 0
                except (ValueError, TypeError):
                    record[col] = 0

        # Clean code columns (remove 'nan', '0', empty)
        for col in ['hcode5', 'hcode9', 'hcode9_new', 'postal_code', 'moo']:
            if col in record:
                val = str(record[col]).strip()
                if val in ['nan', 'None', '0', '0.0']:
                    record[col] = ''
                else:
                    record[col] = val

        # Convert date columns (Thai format DD/MM/YYYY to YYYY-MM-DD)
        for col in ['established_date', 'closed_date', 'source_updated_at']:
            if col in record:
                val = record[col]
                if not val or val in ['', 'nan', 'None', 'NaT']:
                    record[col] = None
                elif isinstance(val, str) and '/' in val:
                    try:
                        parts = val.split('/')
                        if len(parts) == 3:
                            day, month, year = parts
                            year = int(year)
                            if year > 2500:
                                year -= 543
                            record[col] = f"{year:04d}-{int(month):02d}-{int(day):02d}"
                        else:
                            record[col] = None
                    except:
                        record[col] = None
                else:
                    record[col] = None

    # Generate hcode5 for records without one
    # Priority: 1. Existing hcode5, 2. NHSO mapping, 3. Hash-based generated code
    def generate_hcode5(record, idx):
        # Check if already has hcode5
        if record.get('hcode5') and record['hcode5'] not in ['', 'nan']:
            return record['hcode5']

        # Try NHSO mapping by (name, province)
        name = str(record.get('name', '')).strip()
        province = str(record.get('province', '')).strip()

        # Exact match
        if (name, province) in NHSO_HOSPITAL_CODES:
            return NHSO_HOSPITAL_CODES[(name, province)]

        # Fallback: Generate unique code from hash
        key = f"{name}-{province}-{record.get('district', '')}-{record.get('address', '')}-{idx}"
        hash_val = hashlib.md5(key.encode()).hexdigest()[:8]
        return f"G{hash_val}"  # G prefix = Generated

    for idx, record in enumerate(records):
        record['hcode5'] = generate_hcode5(record, idx)

    # Count how many got file codes vs NHSO vs generated
    file_code_count = sum(1 for r in records if r.get('hcode5') and not r['hcode5'].startswith('G') and (r['hcode5'], ) != generate_hcode5({'name': r.get('name'), 'province': r.get('province')}, 0))
    nhso_count = sum(1 for r in records if r.get('hcode5') and not r['hcode5'].startswith('G'))
    generated_count = sum(1 for r in records if r.get('hcode5', '').startswith('G'))
    print(f"[seed] hcode5 from file: {nhso_count - generated_count}, Generated codes: {generated_count}")

    # Verify no duplicates
    hcode5_list = [r.get('hcode5', '') for r in records]
    seen_codes = {}
    for i, code in enumerate(hcode5_list):
        if code in seen_codes:
            seen_codes[code] += 1
            records[i]['hcode5'] = f"{code}_{seen_codes[code]}"
        else:
            seen_codes[code] = 0

    dupes = sum(1 for v in seen_codes.values() if v > 0)
    if dupes > 0:
        print(f"[seed] Warning: {dupes} duplicate hcode5 found, made unique with suffix")

    # Connect to database
    config = get_db_config()

    if DB_TYPE == 'postgresql':
        import psycopg2
        conn = psycopg2.connect(**config)
    else:
        import pymysql
        conn = pymysql.connect(**config)

    cursor = conn.cursor()

    # Stats
    stats = {
        'total': total,
        'imported': 0,
        'updated': 0,
        'skipped': 0,
        'errors': 0
    }

    start_time = datetime.now()

    # Build upsert query
    columns = valid_cols

    if DB_TYPE == 'postgresql':
        placeholders = ', '.join(['%s'] * len(columns))
        update_set = ', '.join([f"{c} = EXCLUDED.{c}" for c in columns if c != 'hcode5'])

        upsert_sql = f"""
            INSERT INTO health_offices ({', '.join(columns)})
            VALUES ({placeholders})
            ON CONFLICT (hcode5) DO UPDATE SET {update_set}
        """
    else:
        placeholders = ', '.join(['%s'] * len(columns))
        update_set = ', '.join([f"{c} = VALUES({c})" for c in columns if c != 'hcode5'])

        upsert_sql = f"""
            INSERT INTO health_offices ({', '.join(columns)})
            VALUES ({placeholders})
            ON DUPLICATE KEY UPDATE {update_set}
        """

    # Process in batches
    for i in range(0, len(records), batch_size):
        batch = records[i:i+batch_size]

        for idx, record in enumerate(batch):
            # Skip if no name (required field)
            if not record.get('name') or record.get('name') in ['', 'nan']:
                stats['skipped'] += 1
                continue

            try:
                values = [record.get(c, '') for c in columns]
                # Clean empty strings to None for proper NULL handling
                values = [None if v == '' else v for v in values]

                cursor.execute(upsert_sql, values)
                stats['imported'] += 1

            except Exception as e:
                stats['errors'] += 1
                if stats['errors'] <= 5:
                    print(f"[seed] Error row {i + idx}: {e}")
                    print(f"[seed] Values: {values[:5]}...")
                    print(f"[seed] hcode5 value: {record.get('hcode5')}")

        conn.commit()
        print(f"[seed] Processed {min(i + batch_size, len(records))}/{len(records)} records...")

    # Log import
    duration = (datetime.now() - start_time).total_seconds()

    log_sql = """
        INSERT INTO health_offices_import_log
        (filename, total_records, imported, updated, skipped, errors, import_mode, status, duration_seconds)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    cursor.execute(log_sql, (
        Path(excel_path).name,
        stats['total'],
        stats['imported'],
        stats['updated'],
        stats['skipped'],
        stats['errors'],
        'seed',
        'completed',
        duration
    ))
    conn.commit()

    cursor.close()
    conn.close()

    stats['duration_seconds'] = duration
    return stats


def main():
    # Find Excel file
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        # Default locations
        candidates = [
            Path(__file__).parent / 'data' / 'health_office.xlsx',
            Path('/app/database/seeds/data/health_office.xlsx'),
            Path(__file__).parent.parent.parent / 'downloads' / 'health_office.xlsx',
        ]

        excel_path = None
        for p in candidates:
            if p.exists():
                excel_path = str(p)
                break

        if not excel_path:
            print("[seed] ERROR: health_office.xlsx not found")
            print("[seed] Please provide path: python health_offices_importer.py /path/to/file.xlsx")
            print("[seed] Or place file in: database/seeds/data/health_office.xlsx")
            sys.exit(1)

    if not Path(excel_path).exists():
        print(f"[seed] ERROR: File not found: {excel_path}")
        sys.exit(1)

    print(f"[seed] Starting health offices import...")
    print(f"[seed] Database type: {DB_TYPE}")

    stats = import_health_offices(excel_path)

    print(f"\n[seed] ===== Import Complete =====")
    print(f"[seed] Total records: {stats['total']}")
    print(f"[seed] Imported: {stats['imported']}")
    print(f"[seed] Skipped (no hcode5): {stats['skipped']}")
    print(f"[seed] Errors: {stats['errors']}")
    print(f"[seed] Duration: {stats['duration_seconds']:.2f}s")


if __name__ == '__main__':
    main()
